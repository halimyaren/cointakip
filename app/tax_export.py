"""
CoinTakip — Vergi-hazır dışa aktarım.

Bu modül **rapor değil, dışa aktarım** üretir. Aradaki fark bilinçli ve
projenin en başından beri kayıtlı bir karardır (`CLAUDE_CODE_HANDOFF.md`,
27 Ağustos yol haritası, 5. madde):

    "Vergi-hazır dışa aktarım — rapor değil, dışa aktarım. Hesaplanmış
     yükümlülük üretmek sorumluluk doğurur ve Türkiye'de kripto
     vergilendirmesi oturmuş değil."

Yani burada **vergi hesaplanmaz.** Matrah, istisna, oran, mahsup, TRY kuru —
hiçbiri bu dosyada yoktur ve olmamalıdır. Üretilen şey, defterdeki
gerçekleşmiş olayların mali müşavirin denetleyebileceği düz bir tablosudur.

## Neden TRY yok

Sistemin tamamı USD üzerinden çalışır. Türkiye'de beyan TRY üzerinden yapılır
ve **işlem tarihindeki** kur gerekir. Kuru buraya koymak üç ayrı karar vermek
demektir: hangi kurum (TCMB, borsa), hangi kur (alış / satış / efektif) ve
tatil günlerinde hangi günün kuru. Bunlar bizim vereceğimiz kararlar değildir;
yanlış kur, doğru veriden yanlış beyan üretir. Bu yüzden her sütun USD kalır
ve dosya bunu kendi üstünde açıkça yazar.

## Neden "eksik veri" sayfası var

Bir vergi dışa aktarımının en tehlikeli hatası, sessizce satır düşürmektir.
Defterde kapanmış ama çıkış fiyatı yazılmamış kayıtlar olabilir — büyük
ihtimalle satılmış ama satış fiyatı girilmemiş pozisyonlardır. Bunları
"gerçekleşmiş K/Z üretmiyor" diye atmak, hem kazancı hem zararı eksik
gösterir. Bu yüzden düşürülmezler; ayrı bir sayfada, neden dışarıda
kaldıkları yazılarak listelenirler.

Aynı sebeple transferler ve mutabakat kapanışları da görünür kalır: defterdeki
her kayıt dört kümeden birine düşer (gerçekleşmiş / eksik veri / kapsam dışı /
hâlâ açık) ve toplamları defterin tamamını verir. Denetlenebilirlik budur.
"""

import csv
import io
from collections import Counter, OrderedDict
from datetime import datetime

from log_config import get_logger

logger = get_logger(__name__)


# Dışa aktarımın kendi sürümü. Sütun düzeni değişirse artırılır; mali müşavir
# elindeki iki dosyanın aynı şemada olup olmadığını buradan görür.
SCHEMA_VERSION = "1.0"

PARA_BIRIMI = "USD"

# Olay türleri. Bir vergi tablosunda "satış" ile "değersiz yazımı" ve
# "toplu mutabakat özeti" aynı satır gibi görünmemelidir: ilki bir alım-satım
# kararının sonucu, ikincisi bir kaybın kabulü, üçüncüsü ise TEK BİR İŞLEM
# DEĞİL, kapanmış birçok işlemin özetidir.
OLAY_SATIS = "Satış"
OLAY_YAZIM = "Yazım (değersiz)"
OLAY_MUTABAKAT = "Mutabakat özeti (toplu)"

# Kapsam dışı bırakma nedenleri.
DISI_TRANSFER = "Transfer — satış değildir, maliyet tabanı korunur"
DISI_REBUILD = "Mutabakat kapanışı — hatalı kayıt düzeltmesi, satış değildir"
DISI_ACIK = "Pozisyon hâlâ açık"

# Eksik veri nedeni.
EKSIK_CIKIS_YOK = "Kapanmış ama çıkış fiyatı/tarihi yok"


def _sayi(deger, varsayilan=0.0):
    """Boş, None veya bozuk değeri patlamadan sayıya çevirir."""
    if deger is None or deger == "":
        return varsayilan
    try:
        return float(deger)
    except (TypeError, ValueError):
        return varsayilan


def _yil(tarih):
    """`2026-03-14` → `2026`. Tanınmayan biçimde boş döner."""
    metin = str(tarih or "").strip()
    if len(metin) >= 4 and metin[:4].isdigit():
        return metin[:4]
    return ""


def _temel_sembol(coin):
    """`ETHUSDT` → `ETH`. `data_manager.base_symbol` ile aynı kural."""
    try:
        from data_manager import base_symbol
        return base_symbol(coin)
    except Exception:
        return str(coin or "").upper().strip()


def _olay_turu(tx):
    """Kapanmış bir kaydın vergi açısından ne olduğunu söyler."""
    neden = tx.get("close_reason")
    if neden == "write_off":
        return OLAY_YAZIM
    if neden == "rebuild_realized":
        return OLAY_MUTABAKAT
    return OLAY_SATIS


def _yazim_aciklamasi(tx):
    try:
        from data_manager import WRITE_OFF_REASONS
        return WRITE_OFF_REASONS.get(tx.get("write_off_reason") or "", "")
    except Exception:
        return ""


def _gerceklesmis_mi(tx):
    """
    Kaydın gerçekleşmiş bir olay olup olmadığı.

    Ölçüt `calculate_realized_metrics` ile birebir aynı tutulur; iki yerin
    farklı cevap vermesi, ekranda görünen K/Z ile dışa aktarılan K/Z'nin
    tutmaması demektir.
    """
    return tx.get("exit_price") is not None or tx.get("realized_pnl_usd") is not None


def _kapali_mi(tx):
    return str(tx.get("status") or "").startswith("Kapand")


def _satir_kur(tx):
    """Gerçekleşmiş bir kaydı dışa aktarım satırına çevirir."""
    miktar = _sayi(tx.get("qty") or tx.get("amount"))
    birim_maliyet = _sayi(tx.get("cost") or tx.get("buy_price"))
    toplam_maliyet = miktar * birim_maliyet

    birim_cikis = _sayi(tx.get("exit_price"), birim_maliyet)
    hasilat = _sayi(tx.get("exit_value"), miktar * birim_cikis)
    komisyon = _sayi(tx.get("fee_usd"))

    kz_ham = tx.get("realized_pnl_usd")
    if kz_ham is not None:
        kz = _sayi(kz_ham)
    else:
        kz = (hasilat - toplam_maliyet) - komisyon

    tur = _olay_turu(tx)
    aciklama = str(tx.get("notes") or "").strip()
    if tur == OLAY_YAZIM:
        etiket = _yazim_aciklamasi(tx)
        if etiket:
            aciklama = f"{etiket}. {aciklama}".strip()

    cikis_tarihi = tx.get("exit_date") or tx.get("date") or ""

    return OrderedDict([
        ("Kayıt No", tx.get("id")),
        ("Varlık", _temel_sembol(tx.get("coin") or tx.get("symbol"))),
        ("Kayıt Sembolü", tx.get("coin") or tx.get("symbol") or ""),
        ("Konum", tx.get("exchange") or ""),
        ("Olay Türü", tur),
        ("Alış Tarihi", tx.get("date") or ""),
        ("Çıkış Tarihi", cikis_tarihi),
        ("Miktar", round(miktar, 12)),
        (f"Birim Alış Fiyatı ({PARA_BIRIMI})", round(birim_maliyet, 12)),
        (f"Toplam Maliyet ({PARA_BIRIMI})", round(toplam_maliyet, 2)),
        (f"Birim Çıkış Fiyatı ({PARA_BIRIMI})", round(birim_cikis, 12)),
        (f"Toplam Hasılat ({PARA_BIRIMI})", round(hasilat, 2)),
        (f"Komisyon ({PARA_BIRIMI})", round(komisyon, 4)),
        (f"Gerçekleşmiş K/Z ({PARA_BIRIMI})", round(kz, 2)),
        ("Maliyet Yöntemi", tx.get("cost_method") or ""),
        ("Kapanış Nedeni", tx.get("close_reason") or ""),
        ("Açıklama", aciklama),
    ])


def _eksik_satir(tx):
    miktar = _sayi(tx.get("qty") or tx.get("amount"))
    birim_maliyet = _sayi(tx.get("cost") or tx.get("buy_price"))
    return OrderedDict([
        ("Kayıt No", tx.get("id")),
        ("Varlık", _temel_sembol(tx.get("coin") or tx.get("symbol"))),
        ("Konum", tx.get("exchange") or ""),
        ("Alış Tarihi", tx.get("date") or ""),
        ("Miktar", round(miktar, 12)),
        (f"Birim Alış Fiyatı ({PARA_BIRIMI})", round(birim_maliyet, 12)),
        (f"Toplam Maliyet ({PARA_BIRIMI})", round(miktar * birim_maliyet, 2)),
        ("Durum", tx.get("status") or ""),
        ("Neden Dışarıda", EKSIK_CIKIS_YOK),
        ("Açıklama", str(tx.get("notes") or "").strip()),
    ])


def _kapsam_disi_satir(tx, neden):
    miktar = _sayi(tx.get("qty") or tx.get("amount"))
    birim_maliyet = _sayi(tx.get("cost") or tx.get("buy_price"))
    return OrderedDict([
        ("Kayıt No", tx.get("id")),
        ("Varlık", _temel_sembol(tx.get("coin") or tx.get("symbol"))),
        ("Konum", tx.get("exchange") or ""),
        ("İşlem Türü", tx.get("type") or ""),
        ("Tarih", tx.get("date") or ""),
        ("Miktar", round(miktar, 12)),
        (f"Toplam Maliyet ({PARA_BIRIMI})", round(miktar * birim_maliyet, 2)),
        ("Durum", tx.get("status") or ""),
        ("Neden Kapsam Dışı", neden),
    ])


def build_tax_rows(data, year=None):
    """
    Defteri dört kümeye ayırır ve dışa aktarım govdesini üretir.

    `year` verilirse yalnızca o yılda **kapanmış** olaylar `rows` içine girer;
    eksik veri ve kapsam dışı listeleri yıla göre süzülmez, çünkü bunların
    amacı "şu yıl ne oldu" değil, "defterin tamamı nereye düştü" sorusuna
    cevap vermektir.

    Hiçbir şey yazmaz; defteri yalnızca okur.
    """
    islemler = list((data or {}).get("transactions", []) or [])

    satirlar = []
    eksikler = []
    kapsam_disi = []
    mevcut_yillar = set()

    for tx in islemler:
        if _gerceklesmis_mi(tx):
            satir = _satir_kur(tx)
            mevcut_yillar.add(_yil(satir["Çıkış Tarihi"]))
            satirlar.append(satir)
            continue

        if not _kapali_mi(tx):
            kapsam_disi.append(_kapsam_disi_satir(tx, DISI_ACIK))
            continue

        if str(tx.get("type") or "").upper() == "TRANSFER":
            kapsam_disi.append(_kapsam_disi_satir(tx, DISI_TRANSFER))
        elif tx.get("close_reason") == "rebuild":
            kapsam_disi.append(_kapsam_disi_satir(tx, DISI_REBUILD))
        else:
            eksikler.append(_eksik_satir(tx))

    mevcut_yillar.discard("")
    yillar = sorted(mevcut_yillar, reverse=True)

    secili = str(year).strip() if year not in (None, "") else None
    if secili:
        satirlar = [s for s in satirlar if _yil(s["Çıkış Tarihi"]) == secili]

    satirlar.sort(key=lambda s: (str(s["Çıkış Tarihi"]), str(s["Varlık"])))

    kz_alani = f"Gerçekleşmiş K/Z ({PARA_BIRIMI})"
    maliyet_alani = f"Toplam Maliyet ({PARA_BIRIMI})"
    hasilat_alani = f"Toplam Hasılat ({PARA_BIRIMI})"
    komisyon_alani = f"Komisyon ({PARA_BIRIMI})"

    kar = sum(s[kz_alani] for s in satirlar if s[kz_alani] > 0)
    zarar = sum(-s[kz_alani] for s in satirlar if s[kz_alani] < 0)

    toplamlar = {
        "olay_sayisi": len(satirlar),
        "toplam_maliyet": round(sum(s[maliyet_alani] for s in satirlar), 2),
        "toplam_hasilat": round(sum(s[hasilat_alani] for s in satirlar), 2),
        "toplam_komisyon": round(sum(s[komisyon_alani] for s in satirlar), 4),
        "toplam_kz": round(sum(s[kz_alani] for s in satirlar), 2),
        "toplam_kar": round(kar, 2),
        "toplam_zarar": round(zarar, 2),
    }

    tur_dagilimi = Counter(s["Olay Türü"] for s in satirlar)

    return {
        "schema_version": SCHEMA_VERSION,
        "currency": PARA_BIRIMI,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "year": secili,
        "available_years": yillar,
        "ledger_tx_count": len(islemler),
        "rows": satirlar,
        "gaps": eksikler,
        "excluded": kapsam_disi,
        "totals": toplamlar,
        "event_types": dict(tur_dagilimi),
    }


# =============================================================
# KAPSAM BEYANI — dosyanın kendi üstünde taşıdığı sınırlar
# =============================================================
def coverage_notes(govde):
    """
    Dosyanın en üstüne yazılan beyan. Bir dışa aktarımın ne OLMADIĞINI
    söylemesi, ne olduğunu söylemesi kadar önemlidir: bu satırlar olmadan
    tablo, olmadığı bir şey (hesaplanmış bir vergi raporu) sanılabilir.
    """
    toplam = govde["totals"]
    notlar = [
        "Bu dosya bir VERGİ BEYANI VEYA VERGİ HESABI DEĞİLDİR. "
        "Matrah, istisna, oran veya mahsup içermez.",
        f"Tüm tutarlar {PARA_BIRIMI} cinsindendir. TRY kuru UYGULANMAMIŞTIR; "
        "beyan için işlem tarihindeki kuru mali müşaviriniz uygulamalıdır.",
        "Fiyatlar üçüncü taraf kaynaklardan gelir ve hatalı ya da gecikmeli "
        "olabilir. Maliyet tabanı kullanıcı tarafından elle girilmiştir.",
        "Transferler satış sayılmaz: bir varlığı kendi cüzdanınıza taşımak "
        "gerçekleşmiş kâr/zarar üretmez ve 'Kapsam Dışı' sayfasında listelenir.",
    ]

    if govde["gaps"]:
        notlar.append(
            f"DİKKAT: {len(govde['gaps'])} kayıt kapanmış görünüyor ama çıkış "
            "fiyatı taşımıyor. Bunlar sessizce atılmadı; 'Eksik Veri' "
            "sayfasındalar. Aralarında satılmış ama satış fiyatı girilmemiş "
            "pozisyonlar varsa bu tablo hem kazancı hem zararı eksik gösterir."
        )

    if govde["event_types"].get(OLAY_MUTABAKAT):
        notlar.append(
            f"{govde['event_types'][OLAY_MUTABAKAT]} satır 'Mutabakat özeti' "
            "türündedir: tek bir işlem değil, borsa dosyasından yeniden "
            "kurulmuş birçok kapanmış işlemin TOPLU sonucudur."
        )

    if govde["event_types"].get(OLAY_YAZIM):
        notlar.append(
            f"{govde['event_types'][OLAY_YAZIM]} satır 'Yazım' türündedir: "
            "değersizleşmiş pozisyonun sıfıra kapatılmasıdır, hasılatı yoktur. "
            "Bunun gider yazılabilirliği mali müşavirinizin kararıdır."
        )

    notlar.append(
        f"Defterdeki {govde['ledger_tx_count']} kaydın tamamı dört kümeden "
        f"birindedir: gerçekleşmiş {toplam['olay_sayisi']}, eksik veri "
        f"{len(govde['gaps'])}, kapsam dışı {len(govde['excluded'])}."
    )
    return notlar


# =============================================================
# CSV
# =============================================================
def export_tax_csv(data, year=None):
    """
    Taşınabilir düz metin ciktisı. Excel'in Türkçe karakterleri doğru
    açması için BOM ile UTF-8 yazılır; bölümler boş satırla ayrılır.
    """
    govde = build_tax_rows(data, year)
    tampon = io.StringIO()
    yazici = csv.writer(tampon, delimiter=";", lineterminator="\r\n")

    yazici.writerow(["CoinTakip — Vergi-Hazır Dışa Aktarım"])
    yazici.writerow(["Şema sürümü", govde["schema_version"]])
    yazici.writerow(["Oluşturulma", govde["generated_at"]])
    yazici.writerow(["Dönem", govde["year"] or "Tüm yıllar"])
    yazici.writerow(["Para birimi", govde["currency"]])
    yazici.writerow([])

    for not_ in coverage_notes(govde):
        yazici.writerow([not_])
    yazici.writerow([])

    def bolum(baslik, satirlar):
        yazici.writerow([baslik])
        if not satirlar:
            yazici.writerow(["(kayıt yok)"])
            yazici.writerow([])
            return
        yazici.writerow(list(satirlar[0].keys()))
        for s in satirlar:
            yazici.writerow(list(s.values()))
        yazici.writerow([])

    bolum("GERÇEKLEŞMİŞ İŞLEMLER", govde["rows"])

    t = govde["totals"]
    yazici.writerow(["TOPLAMLAR"])
    yazici.writerow(["Olay sayısı", t["olay_sayisi"]])
    yazici.writerow([f"Toplam maliyet ({PARA_BIRIMI})", t["toplam_maliyet"]])
    yazici.writerow([f"Toplam hasılat ({PARA_BIRIMI})", t["toplam_hasilat"]])
    yazici.writerow([f"Toplam komisyon ({PARA_BIRIMI})", t["toplam_komisyon"]])
    yazici.writerow([f"Toplam kâr ({PARA_BIRIMI})", t["toplam_kar"]])
    yazici.writerow([f"Toplam zarar ({PARA_BIRIMI})", t["toplam_zarar"]])
    yazici.writerow([f"Net gerçekleşmiş K/Z ({PARA_BIRIMI})", t["toplam_kz"]])
    yazici.writerow([])

    bolum("EKSİK VERİ — İNCELENMELİ", govde["gaps"])
    bolum("KAPSAM DIŞI", govde["excluded"])

    return ("﻿" + tampon.getvalue()).encode("utf-8")


# =============================================================
# XLSX
# =============================================================
def _sutun_genislet(ws, satirlar, ek_baslik=None):
    from openpyxl.utils import get_column_letter
    if not satirlar:
        return
    basliklar = list(satirlar[0].keys())
    for idx, baslik in enumerate(basliklar, 1):
        en = len(str(baslik))
        for s in satirlar[:400]:
            en = max(en, len(str(s.get(baslik, ""))))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(en + 2, 10), 46)


def export_tax_excel(data, year=None):
    """
    Mali müşavire verilecek asıl dosya. Dört sayfa:
    Özet / Gerçekleşmiş İşlemler / Eksik Veri / Kapsam Dışı.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    govde = build_tax_rows(data, year)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
    font_regular = Font(name="Segoe UI", size=9.5, color="1E293B")
    font_warn = Font(name="Segoe UI", size=9.5, color="9A3412")
    font_green = Font(name="Segoe UI", size=9.5, bold=True, color="059669")
    font_red = Font(name="Segoe UI", size=9.5, bold=True, color="DC2626")

    fill_title = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_warn = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    ince = Side(border_style="thin", color="CBD5E1")
    border_cell = Border(left=ince, right=ince, top=ince, bottom=ince)
    sol = Alignment(horizontal="left", vertical="center", wrap_text=True)
    orta = Alignment(horizontal="center", vertical="center")
    sag = Alignment(horizontal="right", vertical="center")

    # ---------- SAYFA 1: Özet ----------
    ws = wb.create_sheet(title="Ozet")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 78

    ws.merge_cells("A1:B1")
    ws["A1"] = "COINTAKIP — VERGI-HAZIR DISA AKTARIM"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = orta
    ws.row_dimensions[1].height = 30

    satir = 3
    for etiket, deger in [
        ("Şema sürümü", govde["schema_version"]),
        ("Oluşturulma", govde["generated_at"]),
        ("Dönem", govde["year"] or "Tüm yıllar"),
        ("Para birimi", govde["currency"]),
        ("Defterdeki toplam kayıt", govde["ledger_tx_count"]),
    ]:
        ws.cell(row=satir, column=1, value=etiket).font = font_bold
        ws.cell(row=satir, column=2, value=deger).font = font_regular
        satir += 1

    satir += 1
    ws.cell(row=satir, column=1, value="KAPSAM BEYANI").font = font_bold
    satir += 1
    for not_ in coverage_notes(govde):
        hucre = ws.cell(row=satir, column=1, value=not_)
        ws.merge_cells(start_row=satir, start_column=1, end_row=satir, end_column=2)
        hucre.font = font_warn
        hucre.fill = fill_warn
        hucre.alignment = sol
        ws.row_dimensions[satir].height = 30
        satir += 1

    satir += 1
    ws.cell(row=satir, column=1, value="TOPLAMLAR").font = font_bold
    satir += 1
    t = govde["totals"]
    for etiket, deger, para in [
        ("Olay sayısı", t["olay_sayisi"], False),
        (f"Toplam maliyet ({PARA_BIRIMI})", t["toplam_maliyet"], True),
        (f"Toplam hasılat ({PARA_BIRIMI})", t["toplam_hasilat"], True),
        (f"Toplam komisyon ({PARA_BIRIMI})", t["toplam_komisyon"], True),
        (f"Toplam kâr ({PARA_BIRIMI})", t["toplam_kar"], True),
        (f"Toplam zarar ({PARA_BIRIMI})", t["toplam_zarar"], True),
        (f"Net gerçekleşmiş K/Z ({PARA_BIRIMI})", t["toplam_kz"], True),
    ]:
        ws.cell(row=satir, column=1, value=etiket).font = font_bold
        hucre = ws.cell(row=satir, column=2, value=deger)
        hucre.font = font_regular
        hucre.alignment = sag
        if para:
            hucre.number_format = "$#,##0.00"
        satir += 1

    # ---------- Tablo sayfaları ----------
    def tablo_sayfasi(baslik, satirlar, bos_mesaj, para_sutunlari=(), kz_sutunu=None):
        sayfa = wb.create_sheet(title=baslik)
        if not satirlar:
            sayfa.column_dimensions["A"].width = 90
            hucre = sayfa.cell(row=1, column=1, value=bos_mesaj)
            hucre.font = font_regular
            hucre.alignment = sol
            return sayfa

        basliklar = list(satirlar[0].keys())
        for idx, h in enumerate(basliklar, 1):
            hucre = sayfa.cell(row=1, column=idx, value=h)
            hucre.font = font_header
            hucre.fill = fill_header
            hucre.alignment = orta
            hucre.border = border_cell
        sayfa.row_dimensions[1].height = 24

        for r_idx, s in enumerate(satirlar, 2):
            for c_idx, h in enumerate(basliklar, 1):
                deger = s.get(h, "")
                hucre = sayfa.cell(row=r_idx, column=c_idx, value=deger)
                hucre.border = border_cell
                hucre.font = font_regular
                if isinstance(deger, (int, float)) and h != "Kayıt No":
                    hucre.alignment = sag
                    if h in para_sutunlari:
                        hucre.number_format = "$#,##0.00"
                    elif h.startswith("Miktar"):
                        hucre.number_format = "#,##0.00000000"
                    elif h.startswith("Birim"):
                        hucre.number_format = "$#,##0.00000000"
                else:
                    hucre.alignment = sol
                if kz_sutunu and h == kz_sutunu and isinstance(deger, (int, float)):
                    hucre.font = font_green if deger >= 0 else font_red
                if r_idx % 2 == 0:
                    hucre.fill = fill_zebra

        sayfa.freeze_panes = "A2"
        sayfa.auto_filter.ref = sayfa.dimensions
        _sutun_genislet(sayfa, satirlar)
        return sayfa

    kz_alani = f"Gerçekleşmiş K/Z ({PARA_BIRIMI})"
    para = {
        f"Toplam Maliyet ({PARA_BIRIMI})",
        f"Toplam Hasılat ({PARA_BIRIMI})",
        f"Komisyon ({PARA_BIRIMI})",
        kz_alani,
    }

    donem = govde["year"] or "seçili dönem"
    tablo_sayfasi(
        "Gerceklesmis Islemler",
        govde["rows"],
        f"{donem} için gerçekleşmiş işlem kaydı yok. "
        "Bu, hiç işlem yapılmadığı anlamına gelebileceği gibi, satışların "
        "deftere çıkış fiyatıyla girilmemiş olması da olabilir — "
        "'Eksik Veri' sayfasına bakın.",
        para_sutunlari=para,
        kz_sutunu=kz_alani,
    )
    tablo_sayfasi(
        "Eksik Veri",
        govde["gaps"],
        "Eksik veri yok: kapanmış her kaydın çıkış bilgisi var.",
        para_sutunlari={f"Toplam Maliyet ({PARA_BIRIMI})"},
    )
    tablo_sayfasi(
        "Kapsam Disi",
        govde["excluded"],
        "Kapsam dışı kayıt yok.",
        para_sutunlari={f"Toplam Maliyet ({PARA_BIRIMI})"},
    )

    cikti = io.BytesIO()
    wb.save(cikti)
    cikti.seek(0)
    logger.info(
        "Vergi disa aktarimi olusturuldu: donem=%s, olay=%d, eksik=%d, kapsam_disi=%d",
        govde["year"] or "tum", len(govde["rows"]),
        len(govde["gaps"]), len(govde["excluded"]),
    )
    return cikti.getvalue()


def tax_summary(data):
    """
    Arayüzün indirmeden önce gösterebileceği özet. Dosya üretmez.
    """
    govde = build_tax_rows(data, None)
    yillik = Counter(_yil(s["Çıkış Tarihi"]) for s in govde["rows"])
    return {
        "available_years": govde["available_years"],
        "year_counts": {y: yillik.get(y, 0) for y in govde["available_years"]},
        "total_events": len(govde["rows"]),
        "gap_count": len(govde["gaps"]),
        "excluded_count": len(govde["excluded"]),
        "ledger_tx_count": govde["ledger_tx_count"],
        "event_types": govde["event_types"],
        "totals": govde["totals"],
        "currency": PARA_BIRIMI,
        "schema_version": SCHEMA_VERSION,
    }

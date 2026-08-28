import os
import json
import copy
import shutil
import hashlib
import secrets
import hmac
import base64
from datetime import datetime
from collections import defaultdict

from log_config import get_logger

logger = get_logger("data_manager")

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
DATA_FILE = os.path.join(DATA_DIR, "portfolio.json")
BACKUP_DIR = os.path.join(DATA_DIR, "backups")

DEFAULT_CATEGORIES = {
    "BTCUSDT": "Majör / L1",
    "ETHUSDT": "Majör / L1",
    "SOLUSDT": "Majör / L1",
    "BNBUSDT": "Majör / L1",
    "AVAXUSDT": "Majör / L1",
    "SUIUSDT": "Majör / L1",
    "XAUTUSDT": "Emtia / Altın",
    "PAXGUSDT": "Emtia / Altın",
    "UNIUSDT": "DeFi / DEX",
    "AAVEUSDT": "DeFi / Lending",
    "MKRUSDT": "DeFi / Lending",
    "LINKUSDT": "Oracle / Infra",
    "ARBUSDT": "Layer 2",
    "OPUSDT": "Layer 2",
    "MATICUSDT": "Layer 2",
    "POLUSDT": "Layer 2",
    "GALAUSDT": "Gaming / NFT",
    "SANDUSDT": "Gaming / Metaverse",
    "MANAUSDT": "Gaming / Metaverse",
    "AXSUSDT": "Gaming / NFT",
    "DOGEUSDT": "Meme",
    "SHIBUSDT": "Meme",
    "PEPEUSDT": "Meme",
    "FETUSDT": "Yapay Zeka (AI)",
    "RENDERUSDT": "Yapay Zeka (AI)",
    "TAOUSDT": "Yapay Zeka (AI)",
    "ONDOUSDT": "RWA",
    "OMUSDT": "RWA",
    "PENDLEUSDT": "DeFi / Yield",
    "ENSUSDT": "Web3 / Domain",
    "ETHFIUSDT": "Restaking",
    "BBUSDT": "Restaking / L2",
    "NOTUSDT": "Gaming / TON",
    "TRUMPUSDT": "Meme / PolitiFi",
    "RDNTUSDT": "Lending",
    "RDNT": "Lending",
    "CATERPILLAR": "Meme / DEX",
    "CPL": "Meme / DEX"
}

def ensure_data_dir():
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR, exist_ok=True)
    if not os.path.exists(BACKUP_DIR):
        os.makedirs(BACKUP_DIR, exist_ok=True)

def initialize_portfolio_if_missing():
    ensure_data_dir()
    if os.path.exists(DATA_FILE):
        return

    initial_transactions = []

    initial_data = {
        "wallets": {
            "usdt_cash": 500.0,
            "exchange_cash": {
                "BINANCE": 400.0,
                "MEXC": 100.0,
                "GATE.IO": 0.0,
                "DEX": 0.0
            },
            "futures_balance": 0.0,
            "margin_balance": 0.0
        },
        "settings": {
            "currency": "USD",
            "theme": "slate",
            "refresh_interval_sec": 4,
            "default_exchange": "BINANCE"
        },
        "transactions": initial_transactions,
        "next_tx_id": len(initial_transactions) + 1,
        "hedges": [],
        "next_hedge_id": 1
    }

    save_portfolio(initial_data)


def _ensure_schema(data):
    """
    Eski portföy dosyalarına yeni bölümleri ekler (geriye dönük uyumluluk).

    FAZ E hedge kayıtlarını `transactions` içine DEĞİL, ayrı bir `hedges`
    listesine yazar. Kaldıraçlı pozisyonun veri modeli spot alımdan farklıdır
    (yön, kaldıraç, marj, giriş fiyatı) ve ikisini aynı listede tutmak maliyet
    tabanı hesabını bozardı.
    """
    if not isinstance(data, dict):
        return data
    if "hedges" not in data or not isinstance(data.get("hedges"), list):
        data["hedges"] = []
    if "next_hedge_id" not in data:
        mevcut = [int(h.get("id", 0) or 0) for h in data["hedges"]]
        data["next_hedge_id"] = (max(mevcut) + 1) if mevcut else 1

    # FAZ F1 — Transferler de hedge gibi AYRI bir listede tutulur.
    # `transactions` içine sahte bir "çıkış kaydı" yazmıyoruz; çünkü transfer
    # ne bir satıştır ne de gerçekleşmiş bir K/Z'dir. Burada tutulan kayıt hem
    # denetim izidir hem de geri almayı mümkün kılar (hangi lot ne kadar
    # tüketildi bilgisi saklanır).
    if "transfers" not in data or not isinstance(data.get("transfers"), list):
        data["transfers"] = []
    if "next_transfer_id" not in data:
        mevcut_t = [int(t.get("id", 0) or 0) for t in data["transfers"]]
        data["next_transfer_id"] = (max(mevcut_t) + 1) if mevcut_t else 1

    # FAZ F5 — Mutabakat düzeltmeleri de ayrı bir denetim defterinde tutulur.
    # Eski lotlar SİLİNMEZ, kapatılır; kayıt hangi lotun kapandığını ve hangi
    # lotların oluştuğunu taşır, böylece düzeltme birebir geri alınabilir.
    if "rebuilds" not in data or not isinstance(data.get("rebuilds"), list):
        data["rebuilds"] = []
    if "next_rebuild_id" not in data:
        mevcut_r = [int(r.get("id", 0) or 0) for r in data["rebuilds"]]
        data["next_rebuild_id"] = (max(mevcut_r) + 1) if mevcut_r else 1
    return data


def load_portfolio():
    ensure_data_dir()
    if not os.path.exists(DATA_FILE):
        initialize_portfolio_if_missing()

    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return _ensure_schema(json.load(f))
    except Exception as e:
        logger.error("portfolio.json okunamadı: %s — yedekten kurtarma deneniyor.", e)
        backups = sorted(os.listdir(BACKUP_DIR))
        if backups:
            recovered = backups[-1]
            logger.warning("Portföy en son yedekten geri yüklendi: %s", recovered)
            with open(os.path.join(BACKUP_DIR, recovered), "r", encoding="utf-8") as bf:
                return _ensure_schema(json.load(bf))
        logger.error("Kullanılabilir yedek bulunamadı — BOŞ portföy ile devam ediliyor.")
        return _ensure_schema({"wallets": {"usdt_cash": 500.0}, "transactions": [], "next_tx_id": 1})

def save_portfolio(data):
    ensure_data_dir()
    temp_file = DATA_FILE + ".tmp"
    with open(temp_file, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(temp_file, DATA_FILE)
    
    today_str = datetime.now().strftime("%Y%m%d")
    backup_file = os.path.join(BACKUP_DIR, f"portfolio_backup_{today_str}.json")
    if not os.path.exists(backup_file):
        try:
            shutil.copyfile(DATA_FILE, backup_file)
            logger.info("Günlük portföy yedeği oluşturuldu: %s", os.path.basename(backup_file))
        except Exception as e:
            # Yedek alınamaması kaydetmeyi engellemez, ama sessiz kalmamalı.
            logger.warning("Günlük yedek oluşturulamadı (%s): %s", os.path.basename(backup_file), e)

# -------------------------------------------------------------
# Fiyat çözümleme yardımcısı
# -------------------------------------------------------------
def resolve_price_info(live_prices, lookup_symbol, raw_coin, base_clean, cost):
    """
    Bir işlem için fiyat kaydını bulur.

    ÖNEMLİ — FAZ B++ öncesi davranış: hiçbir kaynakta bulunamayan coin için
    sessizce `{"price": cost}` döndürülüyordu. Bu, pozisyonu arayüzde
    kalıcı olarak "canlı fiyat = maliyet, K/Z %0.00" gösteriyordu; kullanıcı
    başabaş sandığı pozisyonun aslında hiç fiyat almadığını göremiyordu
    (gerçek örnek: yalnızca WhiteBIT'te işlem gören SCM).

    Şimdi de değerleme maliyet üzerinden yapılır — aksi halde pozisyon
    sıfır değerlenir ve toplam kasa yanlış çöker — ancak kayda `no_source`
    bayrağı eklenir. Arayüz bu bayrağı görüp fiyat yerine "—" ve bir uyarı
    rozeti gösterir. Yani sayı artık bir iddia değil, etiketli bir varsayım.
    """
    info = (
        live_prices.get(lookup_symbol)
        or live_prices.get(raw_coin)
        or live_prices.get(base_clean)
    )
    if info and not info.get("no_source"):
        return info, True

    return {
        "price": cost,
        "open_price": cost,
        "change_pct": 0.0,
        "source": "Kaynak Yok",
        "no_source": True,
    }, False


# -------------------------------------------------------------
# Portfolio Calculation Engine (Multi-Exchange KPI Engine)
# -------------------------------------------------------------
def calculate_portfolio_metrics(data, live_prices):
    transactions = data.get("transactions", [])
    wallets = data.get("wallets", {"usdt_cash": 500.0})
    
    total_usdt_cash = float(wallets.get("usdt_cash", 500.0))
    exchange_cash = wallets.get("exchange_cash", {})
    if not exchange_cash:
        exchange_cash = {
            "BINANCE": total_usdt_cash * 0.8,
            "MEXC": total_usdt_cash * 0.2,
            "GATE.IO": 0.0,
            "DEX": 0.0
        }
    
    futures_balance = float(wallets.get("futures_balance", 0.0))
    margin_balance = float(wallets.get("margin_balance", 0.0))

    active_txs = [t for t in transactions if t.get("status") == "Aktif"]
    # FAZ F1 — Transferle taşınan ve yazımla silinen kayıtlar "izlenen eski
    # pozisyon" değildir; bkz. _defter_artigi_mi().
    closed_txs = [t for t in transactions
                  if t.get("status") == "Kapandı / İzleme" and not _defter_artigi_mi(t)]

    # 1. Consolidated Active Coins Map grouped by (Symbol, Exchange)
    coins_map = {}
    total_active_invested = 0.0
    total_active_current_value = 0.0
    total_24h_diff_usd = 0.0

    # Exchange tracking dictionary
    exchange_stats = defaultdict(lambda: {
        "spot_invested": 0.0,
        "spot_current_value": 0.0,
        "daily_diff_24h_usd": 0.0,
        "active_coins_count": 0,
        "active_tx_count": 0
    })

    for tx in active_txs:
        raw_coin = tx["coin"].upper().strip()
        lookup_symbol = raw_coin if raw_coin.endswith("USDT") else f"{raw_coin}USDT"
        base_clean = raw_coin.replace("USDT", "").replace("/WBNB", "").replace("/USD", "").replace("/USDT", "").strip()
        tx_exchange = tx.get("exchange", "BINANCE").upper().strip()
        
        pos_key = f"{lookup_symbol}@{tx_exchange}"
        
        qty = float(tx["qty"])
        cost = float(tx["cost"])
        tot_cost = qty * cost

        total_active_invested += tot_cost

        price_info, has_source = resolve_price_info(
            live_prices, lookup_symbol, raw_coin, base_clean, cost
        )
        live_price = float(price_info.get("price", cost))
        open_price = float(price_info.get("open_price", live_price))
        change_24h_pct = float(price_info.get("change_pct", 0.0))
        
        cur_val = qty * live_price
        daily_diff = (live_price - open_price) * qty
        total_24h_diff_usd += daily_diff

        # Update exchange stats.
        # FAZ F1c — Eskiden buradaki kural `"UNI" in ex_norm` idi; içinde bu üç
        # harf geçen her konumu (kullanıcının kendi cüzdan adı dahil) DEX
        # kovasına atıyordu. Artık tek bir kanonik fonksiyon kullanılıyor.
        ex_norm = normalize_location(tx_exchange)
        exchange_stats[ex_norm]["spot_invested"] += tot_cost
        exchange_stats[ex_norm]["spot_current_value"] += cur_val
        exchange_stats[ex_norm]["daily_diff_24h_usd"] += daily_diff
        exchange_stats[ex_norm]["active_tx_count"] += 1

        if pos_key not in coins_map:
            coins_map[pos_key] = {
                "pos_key": pos_key,
                "symbol": lookup_symbol,
                "display_name": raw_coin,
                "exchange": tx_exchange,
                "source": price_info.get("source", tx_exchange),
                "category": tx.get("category") or DEFAULT_CATEGORIES.get(lookup_symbol, DEFAULT_CATEGORIES.get(raw_coin, DEFAULT_CATEGORIES.get(base_clean, "Altcoin"))),
                # Grafik türü artık fiyatın GERÇEKTE nereden geldiğine bakar.
                # Eskiden buraya CPL'in kontrat adresi üç kez gömülüydü ve
                # "DEX" yazan her işlem zincir üstü sayılıyordu; cüzdanda
                # tutulan BNB/SOL/ETH de öyle işaretleniyordu.
                "is_dex": bool(price_info.get("is_dex")) or "DEX" in (price_info.get("source") or "").upper(),
                "dex_url": price_info.get("dex_url"),
                "dex_embed_url": price_info.get("embed_url"),
                "dextools_url": price_info.get("dextools_url"),
                "pair_address": price_info.get("pair_address"),
                "chain_id": price_info.get("chain_id"),
                # Fiyat kaynağı bulunamadıysa arayüz bunu açıkça göstermeli.
                "no_source": not has_source,
                "price_source_id": price_info.get("pinned_source"),
                "is_manual_price": bool(price_info.get("is_manual")),
                # Zincir üstü eşleşme yalnızca sembol adına dayanıyorsa bu
                # başka bir zincirdeki aynı adlı token olabilir. Arayüz uyarır.
                "price_match_by": price_info.get("match_by"),
                "dca_count": 0,
                "total_qty": 0.0,
                "total_invested": 0.0,
                "live_price": live_price,
                "change_24h_pct": change_24h_pct,
                "current_value": 0.0,
                "pnl_usd": 0.0,
                "pnl_pct": 0.0,
                "avg_cost": 0.0,
                "breakeven_req_rise_pct": 0.0
            }
            exchange_stats[ex_norm]["active_coins_count"] += 1

        cm = coins_map[pos_key]
        cm["dca_count"] += 1
        cm["total_qty"] += qty
        cm["total_invested"] += tot_cost
        cm["current_value"] += cur_val
        cm["live_price"] = live_price
        cm["open_price"] = open_price
        cm["daily_diff_usd"] = cm.get("daily_diff_usd", 0.0) + daily_diff
        cm["change_24h_pct"] = change_24h_pct

    consolidated_coins = []
    for pos_key, cm in coins_map.items():
        if cm["total_qty"] > 0:
            cm["avg_cost"] = cm["total_invested"] / cm["total_qty"]
        else:
            cm["avg_cost"] = 0.0

        # Kaynağı olmayan pozisyonda "fiyat" her lotun kendi maliyetiydi;
        # birden fazla alım varsa konsolide satır son lotun maliyetini
        # gösteriyordu. Toplam doğruydu ama tek satırlık fiyat tutarsızdı.
        # Excel ve AI gibi ham veriyi okuyan tarafların da tutarlı bir sayı
        # görmesi için ortalama maliyete sabitliyoruz.
        if cm.get("no_source"):
            cm["live_price"] = cm["avg_cost"]
            cm["open_price"] = cm["avg_cost"]

        cm["pnl_usd"] = cm["current_value"] - cm["total_invested"]
        if cm["total_invested"] > 0:
            cm["pnl_pct"] = (cm["pnl_usd"] / cm["total_invested"]) * 100.0
        else:
            cm["pnl_pct"] = 0.0

        if cm["live_price"] > 0 and cm["avg_cost"] > cm["live_price"]:
            cm["breakeven_req_rise_pct"] = ((cm["avg_cost"] - cm["live_price"]) / cm["live_price"]) * 100.0
            cm["profit_margin_pct"] = 0.0
        else:
            cm["breakeven_req_rise_pct"] = 0.0
            cm["profit_margin_pct"] = ((cm["live_price"] - cm["avg_cost"]) / cm["avg_cost"] * 100.0) if cm["avg_cost"] > 0 else 0.0

        # Target Calculation (Faz 1: Hedef Fiyat & Kâr Alma Simülasyonu)
        targets = data.get("targets", {})
        if pos_key in targets:
            tgt = targets[pos_key]
            tgt_price = float(tgt.get("target_price", 0.0))
            tgt_sell_pct = float(tgt.get("target_sell_pct", 100.0))
            tgt_notes = tgt.get("notes", "")

            live_p = cm["live_price"]
            avg_c = cm["avg_cost"]
            tot_q = cm["total_qty"]
            sell_qty = tot_q * (tgt_sell_pct / 100.0)

            tgt_cash_return = sell_qty * tgt_price
            tgt_pnl_usd = (tgt_price - avg_c) * sell_qty
            tgt_pnl_pct = ((tgt_price - avg_c) / avg_c * 100.0) if avg_c > 0 else 0.0

            req_rise_pct = ((tgt_price - live_p) / live_p * 100.0) if live_p > 0 else 0.0
            progress_pct = (live_p / tgt_price * 100.0) if tgt_price > 0 else 0.0
            reached = live_p >= tgt_price if (tgt_price > 0 and live_p > 0) else False

            cm["target"] = {
                "pos_key": pos_key,
                "target_price": tgt_price,
                "target_sell_pct": tgt_sell_pct,
                "notes": tgt_notes,
                "sell_qty": sell_qty,
                "cash_return": tgt_cash_return,
                "pnl_usd": tgt_pnl_usd,
                "pnl_pct": tgt_pnl_pct,
                "req_rise_pct": req_rise_pct,
                "progress_pct": min(max(progress_pct, 0.0), 100.0),
                "reached": reached
            }
        else:
            cm["target"] = None

        # Faz 4: 7 Günlük Mini Trend Grafiği (Sparklines)
        try:
            from price_service import price_service
            sp_data = price_service.get_sparkline_7d(
                cm.get("symbol") or cm.get("display_name"),
                live_price=cm.get("live_price", 0.0),
                change_24h=cm.get("change_24h_pct", 0.0)
            )
            cm["sparkline_7d"] = sp_data.get("points", [])
            cm["change_7d_pct"] = sp_data.get("change_7d_pct", 0.0)
            cm["sparkline_min"] = sp_data.get("min_price", 0.0)
            cm["sparkline_max"] = sp_data.get("max_price", 0.0)
        except Exception:
            lp = cm.get("live_price", 0.0)
            cm["sparkline_7d"] = [lp] * 7
            cm["change_7d_pct"] = 0.0
            cm["sparkline_min"] = lp
            cm["sparkline_max"] = lp

        total_active_current_value += cm["current_value"]
        consolidated_coins.append(cm)

    consolidated_coins.sort(key=lambda x: x["pnl_usd"], reverse=True)

    for cm in consolidated_coins:
        if total_active_current_value > 0:
            cm["portfolio_share_pct"] = (cm["current_value"] / total_active_current_value) * 100.0
        else:
            cm["portfolio_share_pct"] = 0.0

    net_pnl_usd = total_active_current_value - total_active_invested
    net_pnl_pct = (net_pnl_usd / total_active_invested * 100.0) if total_active_invested > 0 else 0.0
    total_portfolio_equity = total_active_current_value + total_usdt_cash + futures_balance + margin_balance

    # 2. Build Multi-Exchange KPI Breakdown
    exchange_kpis = {
        "ALL": {
            "name": "TÜM PORTFÖY",
            "total_kasa": total_portfolio_equity,
            "spot_invested": total_active_invested,
            "spot_current_value": total_active_current_value,
            "net_pnl_usd": net_pnl_usd,
            "net_pnl_pct": net_pnl_pct,
            "daily_diff_24h_usd": total_24h_diff_usd,
            "usdt_cash": total_usdt_cash,
            "active_coins_count": len(consolidated_coins),
            "active_tx_count": len(active_txs),
            # Fiyat kaynağı bulunamayan pozisyon sayısı. Sıfırdan büyükse
            # toplam kasa o pozisyonları maliyet üzerinden sayıyor demektir
            # ve arayüz bunu kullanıcıya bildirir.
            "no_source_count": sum(1 for c in consolidated_coins if c.get("no_source")),
            # FAZ F1 — Yalnızca "kaç tane" demek yetmiyordu. Kullanıcı toplam
            # kasasının ne kadarının doğrulanamamış fiyata dayandığını görmeli;
            # bu tutar maliyet üzerinden sayılıyor, yani bir varsayım.
            "no_source_value_usd": sum(c.get("current_value", 0.0)
                                       for c in consolidated_coins if c.get("no_source")),
            "kasa_share_pct": 100.0
        }
    }

    # Kullanıcının kendi konumları (METAMASK, LEDGER…) da KPI üretmeli;
    # aksi halde varlık orada durur ama Kasa ekranında hiç görünmez.
    known_exchanges = set(list(exchange_stats.keys()) + list(DEFAULT_LOCATIONS))
    for exch in known_exchanges:
        st = exchange_stats[exch]
        c_cash = float(exchange_cash.get(exch, 0.0))
        e_spot_val = st["spot_current_value"]
        e_spot_inv = st["spot_invested"]
        e_kasa = e_spot_val + c_cash
        e_pnl_usd = e_spot_val - e_spot_inv
        e_pnl_pct = (e_pnl_usd / e_spot_inv * 100.0) if e_spot_inv > 0 else 0.0
        e_share = (e_kasa / total_portfolio_equity * 100.0) if total_portfolio_equity > 0 else 0.0

        exchange_kpis[exch] = {
            "name": exch,
            "total_kasa": e_kasa,
            "spot_invested": e_spot_inv,
            "spot_current_value": e_spot_val,
            "net_pnl_usd": e_pnl_usd,
            "net_pnl_pct": e_pnl_pct,
            "daily_diff_24h_usd": st["daily_diff_24h_usd"],
            "usdt_cash": c_cash,
            "active_coins_count": st["active_coins_count"],
            "active_tx_count": st["active_tx_count"],
            "kasa_share_pct": e_share
        }

    # 3. Process Closed / Simulated Transactions with Category Groups
    simulated_txs = []
    sim_categories_set = set()
    total_simulated_old_invested = 0.0
    total_simulated_current_value = 0.0

    for tx in closed_txs:
        raw_coin = tx["coin"].upper().strip()
        lookup_symbol = raw_coin if raw_coin.endswith("USDT") else f"{raw_coin}USDT"
        base_clean = raw_coin.replace("USDT", "").replace("/WBNB", "").replace("/USD", "").replace("/USDT", "").strip()
        qty = float(tx["qty"])
        cost = float(tx["cost"])
        tot_cost = qty * cost

        price_info, has_source = resolve_price_info(
            live_prices, lookup_symbol, raw_coin, base_clean, cost
        )
        live_price = float(price_info.get("price", cost))
        change_24h_pct = float(price_info.get("change_pct", 0.0))
        cur_val = qty * live_price
        pnl_usd = cur_val - tot_cost
        pnl_pct = (pnl_usd / tot_cost * 100.0) if tot_cost > 0 else 0.0

        total_simulated_old_invested += tot_cost
        total_simulated_current_value += cur_val

        cat = tx.get("category") or DEFAULT_CATEGORIES.get(lookup_symbol, DEFAULT_CATEGORIES.get(raw_coin, DEFAULT_CATEGORIES.get(base_clean, "Gözlem / Eski")))
        sim_categories_set.add(cat)

        simulated_txs.append({
            "id": tx["id"],
            "date": tx.get("date", ""),
            "coin": raw_coin,
            "exchange": tx.get("exchange", "BINANCE"),
            "qty": qty,
            "old_cost": cost,
            "old_invested": tot_cost,
            "live_price": live_price,
            "sim_value": cur_val,
            "sim_pnl_usd": pnl_usd,
            "sim_pnl_pct": pnl_pct,
            "change_24h_pct": change_24h_pct,
            "no_source": not has_source,
            "notes": tx.get("notes", ""),
            "category": cat
        })

    # 4. Process Full Enriched DCA Transaction Ledger
    enriched_transactions = []
    for tx in transactions:
        raw_coin = tx["coin"].upper().strip()
        lookup_symbol = raw_coin if raw_coin.endswith("USDT") else f"{raw_coin}USDT"
        base_clean = raw_coin.replace("USDT", "").replace("/WBNB", "").replace("/USD", "").replace("/USDT", "").strip()
        qty = float(tx["qty"])
        cost = float(tx["cost"])
        tot_cost = qty * cost

        price_info, has_source = resolve_price_info(
            live_prices, lookup_symbol, raw_coin, base_clean, cost
        )
        live_price = float(price_info.get("price", cost))
        cur_val = qty * live_price
        pnl_usd = cur_val - tot_cost
        pnl_pct = (pnl_usd / tot_cost * 100.0) if tot_cost > 0 else 0.0

        enriched_transactions.append({
            "id": tx["id"],
            "date": tx.get("date", ""),
            "coin": raw_coin,
            "exchange": tx.get("exchange", "BINANCE"),
            "qty": qty,
            "cost": cost,
            "tot_cost": tot_cost,
            "live_price": live_price,
            "no_source": not has_source,
            "cur_val": cur_val,
            "pnl_usd": pnl_usd,
            "pnl_pct": pnl_pct,
            "status": tx.get("status", "Aktif"),
            "notes": tx.get("notes", ""),
            "category": tx.get("category") or DEFAULT_CATEGORIES.get(lookup_symbol, DEFAULT_CATEGORIES.get(raw_coin, DEFAULT_CATEGORIES.get(base_clean, "Altcoin")))
        })

    enriched_transactions.sort(key=lambda x: x["id"], reverse=False)

    # FAZ E — Hedge katmanı. Spot hesabı bittikten sonra çalışır ve toplam
    # kasaya açık hedge'lerin gerçekleşmemiş K/Z'sini ekler.
    hedge_bilgi = calculate_hedge_metrics(data, live_prices, consolidated_coins)
    hedge_pnl = hedge_bilgi["hedge_kpis"]["unrealized_pnl_usd"]
    for kpi in exchange_kpis.values():
        kpi["total_kasa"] = kpi.get("total_kasa", 0.0)
    exchange_kpis["ALL"]["total_kasa"] += hedge_pnl
    exchange_kpis["ALL"]["hedge_unrealized_pnl_usd"] = hedge_pnl
    exchange_kpis["ALL"]["hedge_margin_usd"] = hedge_bilgi["hedge_kpis"]["margin_usd"]
    exchange_kpis["ALL"]["open_hedge_count"] = hedge_bilgi["hedge_kpis"]["open_count"]

    return {
        "kpis": exchange_kpis["ALL"],
        "exchange_kpis": exchange_kpis,
        # Arayüzün konum listesi tek kaynaktan gelsin — Kasa sekmesi, cüzdan
        # modalı ve transfer hedefi aynı listeyi kullanır.
        "locations": known_locations(data),
        "consolidated_coins": consolidated_coins,
        "hedges": hedge_bilgi["hedges"],
        "hedge_kpis": hedge_bilgi["hedge_kpis"],
        "exposures": hedge_bilgi["exposures"],
        "transactions": enriched_transactions,
        "simulations": simulated_txs,
        "sim_categories": sorted(list(sim_categories_set)),
        "sim_summary": {
            "total_old_invested": total_simulated_old_invested,
            "total_sim_value": total_simulated_current_value,
            "total_sim_pnl_usd": total_simulated_current_value - total_simulated_old_invested,
            "total_sim_pnl_pct": ((total_simulated_current_value - total_simulated_old_invested) / total_simulated_old_invested * 100.0) if total_simulated_old_invested > 0 else 0.0
        }
    }


# =====================================================================
# FAZ E: HEDGE / KALDIRAÇLI POZİSYON KATMANI
# =====================================================================
# Spot alımlardan AYRI bir veri modelidir (`portfolio["hedges"]`).
# Kapsam bilinçli olarak dardır: yön, miktar, giriş fiyatı, kaldıraç ve
# mark-to-market K/Z. Fonlama birikimi, likidasyon fiyatı ve çapraz marj
# KAPSAM DIŞIDIR — onlar bir borsa motoru inşa etmek demektir.
#
# ÖNEMLİ KAVRAM: Kaldıraç, USD cinsinden kâr/zararı DEĞİŞTİRMEZ.
# K/Z = fiyat farkı × miktar. Kaldıraç yalnızca kilitlenen marjı ve
# dolayısıyla marj getirisini (ROE) belirler. 2X ile açılan bir pozisyon
# aynı miktarda 1X pozisyonla aynı doları kazanır, sadece yarısı kadar
# teminat bağlar. Bu ayrım karıştırıldığı için açıkça yazıldı.

HEDGE_DIRECTIONS = ("SHORT", "LONG")
HEDGE_STATUS_OPEN = "Açık"
HEDGE_STATUS_CLOSED = "Kapandı"


def _hedge_lookup_symbol(coin):
    raw = str(coin or "").upper().strip()
    return raw if raw.endswith("USDT") else f"{raw}USDT"


def validate_hedge_payload(payload):
    """Yeni hedge kaydını doğrular. (temiz, hata) döndürür."""
    if not isinstance(payload, dict):
        return None, "Hedge tanımı bir nesne olmalı."

    coin = str(payload.get("coin", "")).upper().strip()
    if not coin:
        return None, "Coin seçilmeli."

    direction = str(payload.get("direction", "SHORT")).upper().strip()
    if direction not in HEDGE_DIRECTIONS:
        return None, f"Yön 'SHORT' veya 'LONG' olmalı (gelen: '{direction}')."

    # DİKKAT: kaldıraç için `or 1` KULLANILMAZ. Kullanıcı 0 yazdığında
    # `0 or 1` sessizce 1'e dönüşüyordu; geçersiz girdi kabul edilmiş gibi
    # davranmak yerine reddedilmeli. Yalnızca alan hiç gönderilmediğinde
    # varsayılan uygulanır.
    ham_kaldirac = payload.get("leverage", 1)
    if ham_kaldirac is None or ham_kaldirac == "":
        ham_kaldirac = 1
    try:
        qty = float(payload.get("qty", 0) or 0)
        entry_price = float(payload.get("entry_price", 0) or 0)
        leverage = float(ham_kaldirac)
        fee_usd = float(payload.get("fee_usd", 0) or 0)
    except (ValueError, TypeError):
        return None, "Miktar, giriş fiyatı ve kaldıraç sayı olmalı."

    if entry_price <= 0:
        return None, "Giriş fiyatı sıfırdan büyük olmalı."
    if leverage < 1 or leverage > 125:
        return None, "Kaldıraç 1 ile 125 arasında olmalı."
    if fee_usd < 0:
        return None, "Komisyon negatif olamaz."

    # Kullanıcı pozisyonu genellikle "100$ teminatla 2X" diye düşünür, coin
    # miktarı diye değil. Miktar verilmediyse teminattan türetilir:
    #   nominal = teminat × kaldıraç,  miktar = nominal / giriş fiyatı
    if qty <= 0 and payload.get("margin_usd") not in (None, ""):
        try:
            margin_usd = float(payload.get("margin_usd") or 0)
        except (ValueError, TypeError):
            return None, "Teminat sayı olmalı."
        if margin_usd <= 0:
            return None, "Teminat sıfırdan büyük olmalı."
        qty = (margin_usd * leverage) / entry_price

    if qty <= 0:
        return None, "Miktar (veya teminat) sıfırdan büyük olmalı."

    return {
        "coin": coin,
        "exchange": str(payload.get("exchange", "BINANCE")).upper().strip() or "BINANCE",
        "direction": direction,
        "qty": qty,
        "entry_price": entry_price,
        "leverage": leverage,
        "fee_usd": fee_usd,
        "open_date": str(payload.get("open_date") or datetime.now().strftime("%Y-%m-%d")),
        "notes": str(payload.get("notes", "")),
    }, None


def hedge_pnl(direction, qty, entry_price, exit_price):
    """
    Kaldıraçtan BAĞIMSIZ kâr/zarar. SHORT fiyat düşünce kazanır.
    """
    if str(direction).upper() == "SHORT":
        return (float(entry_price) - float(exit_price)) * float(qty)
    return (float(exit_price) - float(entry_price)) * float(qty)


def calculate_hedge_metrics(data, live_prices, consolidated_coins=None):
    """
    Açık hedge'leri canlı fiyatla değerler ve coin bazında net maruziyeti
    çıkarır. Kullanıcının asıl görmek istediği şey bu: spot pozisyonunun
    ne kadarının korunduğu ve toplam varlığa etkisi.
    """
    hedges_raw = (data or {}).get("hedges", []) or []

    # Spot miktarları — maruziyet hesabı için
    spot_qty = defaultdict(float)
    spot_price = {}
    for coin in (consolidated_coins or []):
        sym = _hedge_lookup_symbol(coin.get("display_name") or coin.get("symbol"))
        spot_qty[sym] += float(coin.get("total_qty", 0.0) or 0.0)
        if coin.get("live_price"):
            spot_price[sym] = float(coin["live_price"])

    enriched = []
    total_unrealized = 0.0
    total_margin = 0.0
    open_count = 0
    hedge_by_coin = defaultdict(lambda: {"short_qty": 0.0, "long_qty": 0.0, "pnl_usd": 0.0})

    for h in hedges_raw:
        coin = str(h.get("coin", "")).upper().strip()
        sym = _hedge_lookup_symbol(coin)
        qty = float(h.get("qty", 0.0) or 0.0)
        entry = float(h.get("entry_price", 0.0) or 0.0)
        leverage = float(h.get("leverage", 1.0) or 1.0) or 1.0
        direction = str(h.get("direction", "SHORT")).upper()
        status = h.get("status", HEDGE_STATUS_OPEN)
        fee = float(h.get("fee_usd", 0.0) or 0.0)

        notional = qty * entry
        margin = notional / leverage if leverage else notional

        info = live_prices.get(sym) or live_prices.get(coin) or {}
        live_price = float(info.get("price") or 0.0)
        if not live_price or info.get("no_source"):
            live_price = spot_price.get(sym, 0.0)

        row = dict(h)
        row.update({
            "coin": coin,
            "notional_usd": notional,
            "margin_usd": margin,
            "live_price": live_price,
            "price_source": info.get("source"),
        })

        if status == HEDGE_STATUS_CLOSED:
            row["unrealized_pnl_usd"] = 0.0
            row["roe_pct"] = 0.0
            row["price_change_pct"] = 0.0
        else:
            open_count += 1
            pnl = hedge_pnl(direction, qty, entry, live_price) - fee if live_price else 0.0
            row["unrealized_pnl_usd"] = pnl
            row["roe_pct"] = (pnl / margin * 100.0) if margin > 0 else 0.0
            row["price_change_pct"] = ((live_price - entry) / entry * 100.0) if entry > 0 and live_price else 0.0
            total_unrealized += pnl
            total_margin += margin

            bucket = hedge_by_coin[sym]
            if direction == "SHORT":
                bucket["short_qty"] += qty
            else:
                bucket["long_qty"] += qty
            bucket["pnl_usd"] += pnl

        enriched.append(row)

    enriched.sort(key=lambda r: (r.get("status") != HEDGE_STATUS_OPEN, -int(r.get("id", 0) or 0)))

    # Coin bazında net maruziyet
    exposures = []
    for sym in sorted(set(list(hedge_by_coin.keys()) + [s for s in spot_qty if spot_qty[s] > 0])):
        s_qty = spot_qty.get(sym, 0.0)
        bucket = hedge_by_coin.get(sym, {"short_qty": 0.0, "long_qty": 0.0, "pnl_usd": 0.0})
        if s_qty <= 0 and bucket["short_qty"] <= 0 and bucket["long_qty"] <= 0:
            continue
        price = spot_price.get(sym) or float((live_prices.get(sym) or {}).get("price") or 0.0)
        net_qty = s_qty - bucket["short_qty"] + bucket["long_qty"]
        coverage = (bucket["short_qty"] / s_qty * 100.0) if s_qty > 0 else 0.0

        # Yalnızca hedge'i olan ya da anlamlı spot'u olan coinleri göster
        if bucket["short_qty"] <= 0 and bucket["long_qty"] <= 0:
            continue

        exposures.append({
            "symbol": sym,
            "display_name": sym[:-4] if sym.endswith("USDT") else sym,
            "spot_qty": s_qty,
            "short_qty": bucket["short_qty"],
            "long_qty": bucket["long_qty"],
            "net_qty": net_qty,
            "live_price": price,
            "spot_value_usd": s_qty * price,
            "net_value_usd": net_qty * price,
            "coverage_pct": coverage,
            "hedge_pnl_usd": bucket["pnl_usd"],
        })

    return {
        "hedges": enriched,
        "exposures": exposures,
        "hedge_kpis": {
            "unrealized_pnl_usd": total_unrealized,
            "margin_usd": total_margin,
            "open_count": open_count,
            "closed_count": sum(1 for h in hedges_raw if h.get("status") == HEDGE_STATUS_CLOSED),
            "realized_pnl_usd": sum(float(h.get("realized_pnl_usd") or 0.0)
                                    for h in hedges_raw if h.get("status") == HEDGE_STATUS_CLOSED),
        },
    }


def open_hedge(payload):
    clean, err = validate_hedge_payload(payload)
    if err:
        return None, err

    data = load_portfolio()
    hedge_id = int(data.get("next_hedge_id", 1))
    record = {
        "id": hedge_id,
        **clean,
        "status": HEDGE_STATUS_OPEN,
        "close_price": None,
        "close_date": None,
        "realized_pnl_usd": None,
    }
    data.setdefault("hedges", []).append(record)
    data["next_hedge_id"] = hedge_id + 1
    save_portfolio(data)
    logger.info("Hedge açıldı: #%s %s %s %s @ %s (%sX)",
                hedge_id, clean["direction"], clean["qty"], clean["coin"],
                clean["entry_price"], clean["leverage"])
    return record, None


def close_hedge(hedge_id, close_price, fee_usd=0.0, close_date=None):
    """
    Hedge'i kapatır ve gerçekleşmiş K/Z'yi `futures_balance`'a ekler.

    Neden futures_balance: hedge marjı vadeli cüzdanda durur; kapanışta kâr
    veya zarar oraya yansır. Böylece toplam kasa iki kez saymaz — pozisyon
    kapandığı an gerçekleşmemiş K/Z sıfırlanır, tutar bakiyeye geçer.
    """
    try:
        close_price = float(close_price)
        fee_usd = float(fee_usd or 0.0)
    except (ValueError, TypeError):
        return None, "Kapanış fiyatı sayı olmalı."
    if close_price <= 0:
        return None, "Kapanış fiyatı sıfırdan büyük olmalı."
    if fee_usd < 0:
        return None, "Komisyon negatif olamaz."

    data = load_portfolio()
    target = next((h for h in data.get("hedges", []) if int(h.get("id", 0)) == int(hedge_id)), None)
    if not target:
        return None, "Hedge kaydı bulunamadı."
    if target.get("status") == HEDGE_STATUS_CLOSED:
        return None, "Bu hedge zaten kapatılmış."

    realized = hedge_pnl(target.get("direction"), target.get("qty"),
                         target.get("entry_price"), close_price)
    realized -= float(target.get("fee_usd", 0.0) or 0.0) + fee_usd

    target["status"] = HEDGE_STATUS_CLOSED
    target["close_price"] = close_price
    target["close_date"] = str(close_date or datetime.now().strftime("%Y-%m-%d"))
    target["realized_pnl_usd"] = realized
    target["close_fee_usd"] = fee_usd

    wallets = data.setdefault("wallets", {})
    wallets["futures_balance"] = float(wallets.get("futures_balance", 0.0) or 0.0) + realized

    save_portfolio(data)
    logger.info("Hedge kapatıldı: #%s @ %s → gerçekleşmiş K/Z $%.2f", hedge_id, close_price, realized)
    return target, None


def delete_hedge(hedge_id):
    data = load_portfolio()
    hedges = data.get("hedges", [])
    kalan = [h for h in hedges if int(h.get("id", 0)) != int(hedge_id)]
    if len(kalan) == len(hedges):
        return False
    data["hedges"] = kalan
    save_portfolio(data)
    logger.info("Hedge kaydı silindi: #%s", hedge_id)
    return True


def hedge_scenario(data, live_prices, consolidated_coins, move_pct):
    """
    "BTC %X hareket ederse ne olur?" sorusunu cevaplar.

    Spot tarafı ve hedge tarafı ayrı ayrı hesaplanır ki kullanıcı korumanın
    ne kadarını sağladığını görebilsin.
    """
    try:
        move = float(move_pct)
    except (ValueError, TypeError):
        move = 0.0

    bilgi = calculate_hedge_metrics(data, live_prices, consolidated_coins)
    satirlar = []
    toplam_spot_delta = 0.0
    toplam_hedge_delta = 0.0

    for exp in bilgi["exposures"]:
        price = exp["live_price"]
        yeni_fiyat = price * (1.0 + move / 100.0)
        spot_delta = (yeni_fiyat - price) * exp["spot_qty"]
        hedge_delta = ((price - yeni_fiyat) * exp["short_qty"]) + ((yeni_fiyat - price) * exp["long_qty"])
        toplam_spot_delta += spot_delta
        toplam_hedge_delta += hedge_delta
        satirlar.append({
            "symbol": exp["symbol"],
            "display_name": exp["display_name"],
            "new_price": yeni_fiyat,
            "spot_delta_usd": spot_delta,
            "hedge_delta_usd": hedge_delta,
            "net_delta_usd": spot_delta + hedge_delta,
        })

    return {
        "move_pct": move,
        "rows": satirlar,
        "total_spot_delta_usd": toplam_spot_delta,
        "total_hedge_delta_usd": toplam_hedge_delta,
        "total_net_delta_usd": toplam_spot_delta + toplam_hedge_delta,
    }


def save_target(pos_key, target_price, target_sell_pct=100.0, notes=""):
    data = load_portfolio()
    if "targets" not in data or not isinstance(data["targets"], dict):
        data["targets"] = {}
    
    data["targets"][pos_key] = {
        "target_price": float(target_price),
        "target_sell_pct": float(target_sell_pct),
        "notes": notes,
        "updated_at": datetime.now().isoformat()
    }
    save_portfolio(data)
    return data["targets"][pos_key]


def delete_target(pos_key):
    data = load_portfolio()
    if "targets" in data and pos_key in data["targets"]:
        del data["targets"][pos_key]
        save_portfolio(data)
        return True
    return False


def execute_target_sale(pos_key: str, sell_price: float = None, sell_qty: float = None, fee_amount: float = 0.0, fee_asset: str = "USDT", fee_usd: float = 0.0, cost_method: str = "Konsolide Ortalama"):
    data = load_portfolio()
    tgt = data.get("targets", {}).get(pos_key, {})

    target_price = float(tgt.get("target_price", 0.0)) if sell_price is None else float(sell_price)
    target_sell_pct = float(tgt.get("target_sell_pct", 100.0))

    if "@" in pos_key:
        symbol, exch = pos_key.split("@", 1)
    elif "_" in pos_key:
        symbol, exch = pos_key.rsplit("_", 1)
    else:
        symbol, exch = pos_key, "BINANCE"

    def match_tx(tx):
        if tx.get("status") != "Aktif":
            return False
        tx_c = tx.get("coin", "").upper().strip()
        tx_ex = tx.get("exchange", "BINANCE").upper().strip()
        match_coin = (tx_c == symbol) or (tx_c.replace("USDT", "") == symbol.replace("USDT", ""))
        match_exch = (tx_ex == exch) or (exch.startswith("DEX") and tx_ex.startswith("DEX"))
        return match_coin and match_exch

    active_txs = [tx for tx in data.get("transactions", []) if match_tx(tx)]
    if not active_txs:
        raise ValueError(f"Bu varlığa ait aktif işlem bulunamadı: {pos_key}")

    tot_qty = sum(float(tx["qty"]) for tx in active_txs)
    tot_invested = sum(float(tx["qty"]) * float(tx["cost"]) for tx in active_txs)
    avg_cost = (tot_invested / tot_qty) if tot_qty > 0 else 0.0

    if sell_qty is not None and float(sell_qty) > 0:
        qty_to_sell = min(float(sell_qty), tot_qty)
    else:
        qty_to_sell = tot_qty * (target_sell_pct / 100.0)

    total_proceeds = qty_to_sell * target_price
    fee_val_usd = float(fee_usd or 0.0)
    fee_amt = float(fee_amount or 0.0)
    fee_ast = (fee_asset or "USDT").upper().strip()
    date_str = datetime.now().strftime("%Y-%m-%d")

    method_clean = "FIFO" if (cost_method or "").upper() == "FIFO" else "Konsolide Ortalama"

    if method_clean == "Konsolide Ortalama":
        # Proportionally reduce active lots or close them all
        if tot_qty - qty_to_sell <= 1e-8:
            for tx in active_txs:
                tx["status"] = "Kapandı / İzleme"
        else:
            ratio = (tot_qty - qty_to_sell) / tot_qty
            for tx in active_txs:
                tx["qty"] = round(float(tx["qty"]) * ratio, 8)

        new_id = max((t["id"] for t in data["transactions"]), default=0) + 1
        net_pnl = ((target_price - avg_cost) * qty_to_sell) - fee_val_usd
        fee_info_str = f" | Komisyon: {fee_amt} {fee_ast} (${fee_val_usd:,.2f})" if fee_amt > 0 else ""

        closed_record = {
            "id": new_id,
            "date": date_str,
            "coin": symbol,
            "exchange": exch,
            "qty": qty_to_sell,
            "cost": round(avg_cost, 4),
            "status": "Kapandı / İzleme",
            "exit_price": target_price,
            "exit_date": date_str,
            "exit_value": total_proceeds,
            "realized_pnl_usd": round(net_pnl, 2),
            "fee_amount": fee_amt,
            "fee_asset": fee_ast,
            "fee_usd": round(fee_val_usd, 4),
            "cost_method": "Konsolide Ortalama",
            "notes": f"Kısmi Satış (Konsolide Ortalama @${avg_cost:,.2f}) | Gelir: +${total_proceeds:,.2f}{fee_info_str} | Net K/Z: {'+' if net_pnl >= 0 else ''}${net_pnl:,.2f}",
            "category": active_txs[0].get("category", "Altcoin") if active_txs else "Altcoin"
        }
        data["transactions"].append(closed_record)

    else:
        # FIFO Lot-by-Lot Consumption
        remaining_sell = qty_to_sell
        for tx in active_txs:
            if remaining_sell <= 0:
                break
            tx_qty = float(tx["qty"])
            tx_cost = float(tx["cost"])
            
            if tx_qty <= remaining_sell + 1e-8:
                portion = tx_qty
                lot_exit_value = portion * target_price
                lot_fee = fee_val_usd * (lot_exit_value / total_proceeds) if total_proceeds > 0 else 0.0
                tx["status"] = "Kapandı / İzleme"
                tx["exit_price"] = target_price
                tx["exit_date"] = date_str
                tx["exit_value"] = lot_exit_value
                net_pnl = ((target_price - tx_cost) * portion) - lot_fee
                tx["realized_pnl_usd"] = round(net_pnl, 2)
                tx["fee_amount"] = round(fee_amt * (lot_exit_value / total_proceeds), 6) if total_proceeds > 0 else 0.0
                tx["fee_asset"] = fee_ast
                tx["fee_usd"] = round(lot_fee, 4)
                tx["cost_method"] = "FIFO"
                tx["notes"] = (tx.get("notes", "") + f" [FIFO Satış @${target_price:.4f} | Maliyet: ${tx_cost:.4f}]").strip()
                remaining_sell -= portion
            else:
                portion = remaining_sell
                tx["qty"] = tx_qty - portion
                
                new_id = max((t["id"] for t in data["transactions"]), default=0) + 1
                lot_exit_value = portion * target_price
                lot_fee = fee_val_usd * (lot_exit_value / total_proceeds) if total_proceeds > 0 else 0.0
                net_pnl = ((target_price - tx_cost) * portion) - lot_fee
                data["transactions"].append({
                    "id": new_id,
                    "date": date_str,
                    "coin": tx["coin"],
                    "exchange": tx.get("exchange", "BINANCE"),
                    "qty": portion,
                    "cost": tx_cost,
                    "status": "Kapandı / İzleme",
                    "exit_price": target_price,
                    "exit_date": date_str,
                    "exit_value": lot_exit_value,
                    "realized_pnl_usd": round(net_pnl, 2),
                    "fee_amount": round(fee_amt * (lot_exit_value / total_proceeds), 6) if total_proceeds > 0 else 0.0,
                    "fee_asset": fee_ast,
                    "fee_usd": round(lot_fee, 4),
                    "cost_method": "FIFO",
                    "notes": f"Kısmi Satış (FIFO Lot #{tx['id']} @${tx_cost:.4f})",
                    "category": tx.get("category", "Altcoin")
                })
                remaining_sell = 0

    # Add cash to wallet
    cash_to_add = total_proceeds - fee_val_usd if fee_ast == "USDT" else total_proceeds
    wallets = data.get("wallets", {})
    exchange_cash = wallets.get("exchange_cash", {})
    # FAZ F1c — Eskiden bilinmeyen her konum "BINANCE"a düşüyordu. Transfer
    # özelliğiyle birlikte kullanıcı artık METAMASK gibi kendi konumlarını
    # yaratabiliyor; oradan yapılan bir satışın gelirini Binance'e yazmak
    # nakit dağılımını sessizce bozardı. Konum artık olduğu gibi kullanılır.
    ex_norm = normalize_location(exch)
    current_cash = float(exchange_cash.get(ex_norm, 0.0))
    exchange_cash[ex_norm] = current_cash + cash_to_add
    wallets["exchange_cash"] = exchange_cash
    wallets["usdt_cash"] = sum(float(v) for v in exchange_cash.values())
    data["wallets"] = wallets

    if tot_qty - qty_to_sell <= 1e-8:
        if pos_key in data.get("targets", {}):
            del data["targets"][pos_key]

    save_portfolio(data)
    return {
        "proceeds": total_proceeds,
        "cash_added": cash_to_add,
        "qty_sold": qty_to_sell,
        "target_price": target_price,
        "cost_method": method_clean,
        "fee_usd": fee_val_usd,
        "pos_key": pos_key
    }


# =====================================================================
# FAZ F1: DEĞER KAYBI YAZIMI (MEZARLIK) VE TRANSFER KATMANI
# =====================================================================
# İki farklı gerçek olayı birbirinden ayırmak için var:
#
#   1. YAZIM (write-off) — Coin öldü. Delist edildi, proje çöktü ya da
#      cüzdana erişim kayboldu. Ekonomik olarak zaten olan şey kayıttır:
#      pozisyon 0'dan kapanır, maliyetin tamamı gerçekleşmiş ZARAR olur.
#      KASAYA NAKİT EKLENMEZ — satış değildir, gelir yoktur.
#
#   2. TRANSFER — Coin yaşıyor, sadece yer değiştirdi (Binance → MetaMask).
#      Bu bir satış DEĞİLDİR. Gerçekleşmiş K/Z üretmez, nakit hareketi
#      yaratmaz ve MALİYET TABANI KORUNUR. Hedef borsada aynı maliyetle
#      yeni lotlar açılır; her kaynak lot kendi maliyetiyle taşınır, böylece
#      ileride FIFO hâlâ doğru çalışır.
#
# Bu ikisi karıştırılırsa portföy matematiği bozulur: transferi satış saymak
# sahte kâr/zarar üretir, yazımı satış saymak da olmayan bir nakit yaratır.

WRITE_OFF_REASONS = {
    "delist": "Borsadan çıkarıldı (delist)",
    "rug": "Proje çöktü / rug pull",
    "lost": "Erişim kaybı (cüzdan veya anahtar)",
    "worthless": "Değersizleşti",
    "other": "Diğer",
}

CLOSED_STATUS = "Kapandı / İzleme"
ACTIVE_STATUS = "Aktif"

# Kutudan çıktığı hâliyle bilinen konumlar. Bu liste bir SINIR DEĞİL, yalnızca
# arayüzün boş portföyde bile gösterebileceği başlangıç kümesidir; kullanıcı
# transferle kendi konumunu (METAMASK, LEDGER, TRUST WALLET…) yaratabilir ve
# sistemin her katmanı onu eşit vatandaş olarak görmelidir.
DEFAULT_LOCATIONS = ("BINANCE", "MEXC", "GATE.IO", "DEX")


def normalize_location(name: str) -> str:
    """
    Konum adını kanonik hâle getirir.

    Zincir üstü işlemler için kullanılan çeşitli adlar tek bir "DEX" kovasında
    toplanır (kullanıcının eski kayıtları "DEX", "PANCAKESWAP" gibi farklı
    adlar taşıyor). Bunun dışındaki her ad OLDUĞU GİBİ korunur — bilinmeyeni
    varsayılana düşürmek, kullanıcının kendi cüzdanını yok saymak demektir.
    """
    temiz = (name or "").upper().strip()
    if not temiz:
        return "BINANCE"
    if temiz.startswith("DEX") or "PANCAKE" in temiz or "UNISWAP" in temiz:
        return "DEX"
    return temiz


def known_locations(data=None) -> list:
    """
    Portföyde fiilen kullanılan tüm konumlar + varsayılanlar.

    Arayüzün her yeri (Kasa sekmesi, cüzdan modalı, transfer hedefi) bu tek
    kaynaktan beslenir; böylece bir yere eklenen konum diğerlerinde kaybolmaz.
    """
    if data is None:
        data = load_portfolio()
    bulunan = set(DEFAULT_LOCATIONS)
    for tx in data.get("transactions", []):
        if tx.get("status") == ACTIVE_STATUS:
            bulunan.add(normalize_location(tx.get("exchange")))
    for anahtar in (data.get("wallets", {}).get("exchange_cash") or {}):
        bulunan.add(normalize_location(anahtar))
    for t in data.get("transfers", []):
        bulunan.add(normalize_location(t.get("from_exchange")))
        bulunan.add(normalize_location(t.get("to_exchange")))

    # Varsayılanlar bilinen sırayla önde, kullanıcının kendi konumları alfabetik.
    ekstra = sorted(bulunan - set(DEFAULT_LOCATIONS))
    return [x for x in DEFAULT_LOCATIONS if x in bulunan] + ekstra


def _split_pos_key(pos_key: str):
    """`BTCUSDT@BINANCE` → ("BTCUSDT", "BINANCE"). Eski `_` biçimini de kabul eder."""
    pos_key = (pos_key or "").strip()
    if "@" in pos_key:
        symbol, exch = pos_key.split("@", 1)
    elif "_" in pos_key:
        symbol, exch = pos_key.rsplit("_", 1)
    else:
        symbol, exch = pos_key, "BINANCE"
    return symbol.upper().strip(), exch.upper().strip()


def _lot_matches(tx, symbol, exch):
    """execute_target_sale ile aynı eşleştirme kuralı — davranış tutarlı kalsın."""
    if tx.get("status") != ACTIVE_STATUS:
        return False
    tx_c = (tx.get("coin") or "").upper().strip()
    tx_ex = (tx.get("exchange") or "BINANCE").upper().strip()
    match_coin = (tx_c == symbol) or (tx_c.replace("USDT", "") == symbol.replace("USDT", ""))
    match_exch = (tx_ex == exch) or (exch.startswith("DEX") and tx_ex.startswith("DEX"))
    return match_coin and match_exch


def _aktif_lotlar(data, pos_key):
    symbol, exch = _split_pos_key(pos_key)
    lots = [tx for tx in data.get("transactions", []) if _lot_matches(tx, symbol, exch)]
    # FIFO anlamlı olsun diye eski lot önce gelir.
    lots.sort(key=lambda t: (t.get("date", ""), int(t.get("id", 0) or 0)))
    return symbol, exch, lots


def _sonraki_tx_id(data):
    return max((int(t.get("id", 0) or 0) for t in data.get("transactions", [])), default=0) + 1


def _defter_artigi_mi(tx):
    """
    Bu kayıt "izlenen eski pozisyon" listesinde GÖRÜNMEMELİ mi?

    Transferle taşınmış ya da yazımla silinmiş lotlar kapanmış olsalar da
    simülasyon listesine ait değiller: ilki başka borsada hâlâ duruyor
    (iki kez sayılırdı), ikincisi ise zaten sıfırlandı ve canlı fiyat
    bulunamadığında maliyet üzerinden değerlenip "para hâlâ orada" izlenimi
    verirdi. İkisi de gerçekleşmiş K/Z tarafında doğru şekilde görünüyor.

    FAZ F5 — Mutabakat düzeltmesiyle kapatılan lotlar da buraya girer: onların
    yerine aynı pozisyonun borsa kayıtlarından kurulmuş hâli duruyor. Listede
    göstermek aynı varlığı iki kez saydırırdı.
    """
    return bool(
        tx.get("transfer_out_id")
        or tx.get("write_off_ref")
        or tx.get("rebuild_out_id")
        or tx.get("close_reason") in ("write_off", "rebuild")
    )


def write_off_position(pos_key: str, reason: str = "worthless", note: str = ""):
    """
    Bir pozisyonun tamamını 0 fiyattan kapatır ve maliyeti zarar olarak yazar.

    Nakit eklenmez. Kapatılan lotlar silinmez — geri alınabilsin diye
    `write_off_ref` ile özet kayda bağlanır.
    """
    data = load_portfolio()
    symbol, exch, lots = _aktif_lotlar(data, pos_key)
    if not lots:
        raise ValueError(f"Bu varlığa ait aktif işlem bulunamadı: {pos_key}")

    reason_key = reason if reason in WRITE_OFF_REASONS else "other"
    tot_qty = sum(float(tx.get("qty") or 0.0) for tx in lots)
    tot_invested = sum(float(tx.get("qty") or 0.0) * float(tx.get("cost") or 0.0) for tx in lots)
    avg_cost = (tot_invested / tot_qty) if tot_qty > 0 else 0.0

    new_id = _sonraki_tx_id(data)
    affected_ids = []
    for tx in lots:
        tx["status"] = CLOSED_STATUS
        tx["write_off_ref"] = new_id
        affected_ids.append(int(tx.get("id", 0) or 0))

    date_str = datetime.now().strftime("%Y-%m-%d")
    aciklama = WRITE_OFF_REASONS[reason_key]
    kullanici_notu = f" | Not: {note.strip()}" if (note or "").strip() else ""

    ozet = {
        "id": new_id,
        "date": date_str,
        "coin": symbol,
        "exchange": exch,
        "qty": tot_qty,
        "cost": round(avg_cost, 8),
        "status": CLOSED_STATUS,
        "type": "YAZIM",
        "exit_price": 0.0,
        "exit_date": date_str,
        "exit_value": 0.0,
        "realized_pnl_usd": -round(tot_invested, 2),
        "fee_amount": 0.0,
        "fee_asset": "USDT",
        "fee_usd": 0.0,
        "cost_method": "Değer Kaybı Yazımı",
        # Vergi/raporlama tarafında yazımı gerçek satıştan ayırmayı sağlar.
        "close_reason": "write_off",
        "write_off_reason": reason_key,
        "write_off_affected_ids": affected_ids,
        "notes": f"Değer kaybı yazımı ({aciklama}) | {len(lots)} lot, "
                 f"maliyet ${tot_invested:,.2f} zarar yazıldı{kullanici_notu}",
        "category": lots[0].get("category", "Altcoin"),
    }
    data["transactions"].append(ozet)

    if pos_key in data.get("targets", {}):
        del data["targets"][pos_key]

    save_portfolio(data)
    return {
        "write_off_id": new_id,
        "pos_key": pos_key,
        "coin": symbol,
        "exchange": exch,
        "qty": tot_qty,
        "lot_count": len(lots),
        "realized_loss_usd": round(tot_invested, 2),
        "reason": reason_key,
        "reason_label": aciklama,
    }


def undo_write_off(write_off_id: int):
    """Yanlışlıkla yazılan bir pozisyonu geri açar."""
    data = load_portfolio()
    ozet = next((t for t in data.get("transactions", [])
                 if int(t.get("id", 0) or 0) == int(write_off_id)
                 and t.get("close_reason") == "write_off"), None)
    if not ozet:
        raise ValueError(f"Yazım kaydı bulunamadı: {write_off_id}")

    geri_gelen = set(int(i) for i in ozet.get("write_off_affected_ids", []))
    sayac = 0
    for tx in data.get("transactions", []):
        if int(tx.get("id", 0) or 0) in geri_gelen:
            tx["status"] = ACTIVE_STATUS
            tx.pop("write_off_ref", None)
            sayac += 1

    data["transactions"] = [t for t in data["transactions"]
                            if int(t.get("id", 0) or 0) != int(write_off_id)]
    save_portfolio(data)
    return {"restored_lots": sayac, "write_off_id": int(write_off_id)}


def validate_transfer_payload(payload: dict):
    """(temiz_veri, hata_mesajı) döndürür."""
    if not isinstance(payload, dict):
        return None, "Geçersiz istek gövdesi."

    pos_key = (payload.get("pos_key") or "").strip()
    if not pos_key:
        return None, "Kaynak pozisyon belirtilmedi."

    to_exchange = (payload.get("to_exchange") or "").upper().strip()
    if not to_exchange:
        return None, "Hedef konum belirtilmedi."

    _, from_exchange = _split_pos_key(pos_key)
    if to_exchange == from_exchange:
        return None, "Kaynak ve hedef konum aynı olamaz."

    try:
        qty = float(payload.get("qty") or 0.0)
    except (TypeError, ValueError):
        return None, "Miktar sayısal olmalı."
    if qty <= 0:
        return None, "Miktar sıfırdan büyük olmalı."

    try:
        fee_qty = float(payload.get("fee_qty") or 0.0)
    except (TypeError, ValueError):
        return None, "Ağ ücreti sayısal olmalı."
    if fee_qty < 0:
        return None, "Ağ ücreti negatif olamaz."
    if fee_qty >= qty:
        return None, "Ağ ücreti transfer miktarından küçük olmalı."

    tarih = (payload.get("date") or "").strip() or datetime.now().strftime("%Y-%m-%d")

    return {
        "pos_key": pos_key,
        "to_exchange": to_exchange,
        "qty": qty,
        "fee_qty": fee_qty,
        "date": tarih,
        "note": (payload.get("note") or "").strip(),
    }, None


def transfer_position(payload: dict):
    """
    Bir varlığı borsalar/cüzdanlar arasında taşır.

    Satış DEĞİLDİR: nakit hareketi yok, gerçekleşmiş K/Z yok. Kaynak lotlar
    FIFO sırasıyla tüketilir ve her biri hedefte KENDİ maliyetiyle yeniden
    açılır. Ağ ücreti olarak yanan miktar gerçekten kaybedilmiş coindir; o
    kadarlık maliyet ayrı bir yazım kaydıyla zarara geçer.
    """
    temiz, hata = validate_transfer_payload(payload)
    if hata:
        raise ValueError(hata)

    data = load_portfolio()
    symbol, from_exchange, lots = _aktif_lotlar(data, temiz["pos_key"])
    if not lots:
        raise ValueError(f"Bu varlığa ait aktif işlem bulunamadı: {temiz['pos_key']}")

    mevcut_qty = sum(float(tx.get("qty") or 0.0) for tx in lots)
    if temiz["qty"] > mevcut_qty + 1e-8:
        raise ValueError(
            f"Yetersiz bakiye: {from_exchange} üzerinde {mevcut_qty:.8f} {symbol} var, "
            f"{temiz['qty']:.8f} transfer edilemez."
        )

    transfer_id = int(data.get("next_transfer_id", 1))
    kalan = min(temiz["qty"], mevcut_qty)
    transfer_qty = kalan
    # Ücret oranı: her lottan aynı oranda yanar, böylece maliyet dağılımı bozulmaz.
    ucret_orani = (temiz["fee_qty"] / transfer_qty) if transfer_qty > 0 else 0.0

    tuketilen = []
    olusan_ids = []
    yanan_maliyet = 0.0
    next_id = _sonraki_tx_id(data)

    for tx in lots:
        if kalan <= 1e-12:
            break
        lot_qty = float(tx.get("qty") or 0.0)
        if lot_qty <= 0:
            continue
        lot_cost = float(tx.get("cost") or 0.0)
        alinan = min(lot_qty, kalan)

        tam_kapandi = (lot_qty - alinan) <= 1e-8
        if tam_kapandi:
            tx["status"] = CLOSED_STATUS
            tx["transfer_out_id"] = transfer_id
            tx["exit_date"] = temiz["date"]
        else:
            tx["qty"] = round(lot_qty - alinan, 12)

        tuketilen.append({
            "tx_id": int(tx.get("id", 0) or 0),
            "qty": alinan,
            "cost": lot_cost,
            "closed": tam_kapandi,
        })

        varan = alinan * (1.0 - ucret_orani)
        yanan_maliyet += (alinan - varan) * lot_cost

        if varan > 0:
            data["transactions"].append({
                "id": next_id,
                # Orijinal alım tarihi korunur — transfer yeni bir alım değildir,
                # elde tutma süresi ve maliyet tabanı devam eder.
                "date": tx.get("date", temiz["date"]),
                "coin": symbol,
                "exchange": temiz["to_exchange"],
                "qty": varan,
                "cost": lot_cost,
                "status": ACTIVE_STATUS,
                "type": "TRANSFER",
                "transfer_in_id": transfer_id,
                "cost_method": tx.get("cost_method", "Konsolide Ortalama"),
                "notes": f"Transfer: {from_exchange} → {temiz['to_exchange']} "
                         f"(kaynak lot #{tx.get('id')}, maliyet korundu)",
                "category": tx.get("category", "Altcoin"),
            })
            olusan_ids.append(next_id)
            next_id += 1

        kalan -= alinan

    ucret_tx_id = None
    if temiz["fee_qty"] > 0 and yanan_maliyet > 0:
        ucret_tx_id = next_id
        data["transactions"].append({
            "id": ucret_tx_id,
            "date": temiz["date"],
            "coin": symbol,
            "exchange": from_exchange,
            "qty": temiz["fee_qty"],
            "cost": round(yanan_maliyet / temiz["fee_qty"], 8) if temiz["fee_qty"] > 0 else 0.0,
            "status": CLOSED_STATUS,
            "type": "YAZIM",
            "exit_price": 0.0,
            "exit_date": temiz["date"],
            "exit_value": 0.0,
            "realized_pnl_usd": -round(yanan_maliyet, 2),
            "fee_amount": temiz["fee_qty"],
            "fee_asset": symbol.replace("USDT", "") or symbol,
            "fee_usd": round(yanan_maliyet, 4),
            "cost_method": "Değer Kaybı Yazımı",
            "close_reason": "write_off",
            "write_off_reason": "other",
            "write_off_affected_ids": [],
            "transfer_fee_of": transfer_id,
            "notes": f"Transfer ağ ücreti: {temiz['fee_qty']} {symbol} yandı "
                     f"({from_exchange} → {temiz['to_exchange']})",
            "category": lots[0].get("category", "Altcoin"),
        })
        next_id += 1

    kayit = {
        "id": transfer_id,
        "date": temiz["date"],
        "coin": symbol,
        "from_exchange": from_exchange,
        "to_exchange": temiz["to_exchange"],
        "qty": transfer_qty,
        "fee_qty": temiz["fee_qty"],
        "received_qty": transfer_qty - temiz["fee_qty"],
        "fee_cost_usd": round(yanan_maliyet, 4),
        "note": temiz["note"],
        "consumed": tuketilen,
        "created_tx_ids": olusan_ids,
        "fee_tx_id": ucret_tx_id,
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    data["transfers"].append(kayit)
    data["next_transfer_id"] = transfer_id + 1

    # Kaynak pozisyon tamamen boşaldıysa oradaki hedef fiyat kaydı anlamsızlaşır.
    if (mevcut_qty - transfer_qty) <= 1e-8:
        eski_key = f"{symbol}@{from_exchange}"
        if eski_key in data.get("targets", {}):
            del data["targets"][eski_key]

    save_portfolio(data)
    return kayit


def undo_transfer(transfer_id: int):
    """Transferi geri alır: hedefteki lotları siler, kaynak lotları eski hâline getirir."""
    data = load_portfolio()
    kayit = next((t for t in data.get("transfers", [])
                  if int(t.get("id", 0) or 0) == int(transfer_id)), None)
    if not kayit:
        raise ValueError(f"Transfer kaydı bulunamadı: {transfer_id}")

    silinecek = set(int(i) for i in kayit.get("created_tx_ids", []))
    if kayit.get("fee_tx_id"):
        silinecek.add(int(kayit["fee_tx_id"]))

    # Hedefte oluşan lotlar sonradan satıldıysa geri alma veriyi bozar.
    kalanlar = {int(t.get("id", 0) or 0): t for t in data.get("transactions", [])}
    for tx_id in kayit.get("created_tx_ids", []):
        tx = kalanlar.get(int(tx_id))
        if tx is None:
            raise ValueError(
                "Bu transfer geri alınamaz: hedefte oluşan lotlardan biri silinmiş."
            )
        if tx.get("status") != ACTIVE_STATUS:
            raise ValueError(
                "Bu transfer geri alınamaz: transfer edilen varlık hedefte "
                "satılmış veya kapatılmış. Önce o işlemi geri alın."
            )

    data["transactions"] = [t for t in data.get("transactions", [])
                            if int(t.get("id", 0) or 0) not in silinecek]

    kalanlar = {int(t.get("id", 0) or 0): t for t in data["transactions"]}
    for parca in kayit.get("consumed", []):
        tx = kalanlar.get(int(parca.get("tx_id", 0) or 0))
        if tx is None:
            continue
        if parca.get("closed"):
            tx["status"] = ACTIVE_STATUS
            tx.pop("transfer_out_id", None)
            tx.pop("exit_date", None)
        else:
            tx["qty"] = round(float(tx.get("qty") or 0.0) + float(parca.get("qty") or 0.0), 12)

    data["transfers"] = [t for t in data.get("transfers", [])
                         if int(t.get("id", 0) or 0) != int(transfer_id)]
    save_portfolio(data)
    return {"transfer_id": int(transfer_id), "restored_lots": len(kayit.get("consumed", []))}


def list_transfers(data=None):
    """Transfer defterini yeniden eskiye doğru döndürür."""
    if data is None:
        data = load_portfolio()
    kayitlar = list(data.get("transfers", []))
    kayitlar.sort(key=lambda t: int(t.get("id", 0) or 0), reverse=True)
    return kayitlar


def list_write_offs(data=None):
    """Yazım kayıtlarını yeniden eskiye doğru döndürür."""
    if data is None:
        data = load_portfolio()
    kayitlar = [t for t in data.get("transactions", [])
                if t.get("close_reason") == "write_off"]
    kayitlar.sort(key=lambda t: int(t.get("id", 0) or 0), reverse=True)
    return kayitlar


# =====================================================================
# FAZ F5: MUTABAKAT DÜZELTMESİ
# =====================================================================
"""
NE YAPAR
--------
Bir pozisyonun defterdeki lotlarını, borsa dışa aktarım dosyalarından FIFO ile
yeniden kurulmuş lotlarla DEĞİŞTİRİR. Öneriyi `reconcile.build_rebuild_plan`
üretir; burası yalnızca uygular.

DÖRT KURAL
----------
1. **Sessiz düzeltme yok.** Her düzeltme `rebuilds` defterinde ayrı, gerekçeli
   ve geri alınabilir bir kayıttır. Eski lotlar silinmez, kapatılır.

2. **Toplu içe aktarma yok.** Uygulama pozisyon başınadır; kullanıcı her birini
   ayrı görür ve ayrı onaylar.

3. **Gerçekleşmiş K/Z üretilmez.** Bu bir ekonomik olay değil, kayıt
   düzeltmesidir. Kapatılan lotlara `exit_price` veya `realized_pnl_usd`
   YAZILMAZ; `calculate_realized_metrics` yalnızca bu iki alandan birini
   taşıyan kayıtları saydığı için düzeltme K/Z tablosunu kirletmez.

4. **Sunucu istemciye güvenmez.** Uygulanacak lotlar istek gövdesinden
   alınmaz; sunucu planı kendisi yeniden üretir ve istemcinin gönderdiği imzayı
   doğrular. Böylece ne uydurma bir maliyet enjekte edilebilir ne de ekranda
   duran eski bir öneri sessizce başka bir şeye dönüşebilir.
"""

REBUILD_CLOSE_REASON = "rebuild"


def get_rebuild_plan(data=None, root=None):
    """Düzeltme önerilerini döndürür. Salt okunur."""
    import reconcile
    if data is None:
        data = load_portfolio()
    return reconcile.build_rebuild_plan(data, root)


def apply_rebuild(pos_key: str, signature: str = None, note: str = "", root=None):
    """
    Bir pozisyonun lotlarını borsa kayıtlarından kurulmuş hâliyle değiştirir.

    Nakit hareketi yok, gerçekleşmiş K/Z yok. Geri alınabilir.
    """
    import reconcile
    data = load_portfolio()
    plan = reconcile.build_rebuild_plan(data, root)
    satir = next((r for r in plan["rows"] if r["pos_key"] == pos_key), None)
    if satir is None:
        raise ValueError(f"Bu pozisyon için düzeltme önerisi yok: {pos_key}")

    if satir["status"] == "blocked":
        raise ValueError("Bu pozisyon düzeltilemez — " + " ".join(satir["blockers"]))
    if satir["status"] == "identical":
        raise ValueError("Bu pozisyon zaten borsa kayıtlarıyla aynı; "
                         "değiştirilecek bir şey yok.")
    if signature and signature != satir["signature"]:
        raise ValueError("Öneri değişmiş — dışa aktarım dosyaları güncellenmiş olabilir. "
                         "Listeyi yenileyip yeniden bakın.")

    symbol, exch, eski_lotlar = _aktif_lotlar(data, pos_key)
    rid = int(data.get("next_rebuild_id", 1))
    bugun = datetime.now().strftime("%Y-%m-%d")

    onceki_qty = sum(float(tx.get("qty") or 0.0) for tx in eski_lotlar)
    onceki_maliyet = sum(float(tx.get("qty") or 0.0) * float(tx.get("cost") or 0.0)
                         for tx in eski_lotlar)

    kapatilan_ids = []
    for tx in eski_lotlar:
        tx["status"] = CLOSED_STATUS
        tx["rebuild_out_id"] = rid
        # Gerçekleşmiş K/Z alanları BİLEREK yazılmıyor (bkz. kural 3).
        tx["close_reason"] = REBUILD_CLOSE_REASON
        tx["exit_date"] = bugun
        kapatilan_ids.append(int(tx.get("id", 0) or 0))

    kategori = eski_lotlar[0].get("category", "Altcoin") if eski_lotlar else "Altcoin"
    coin_adi = satir["coin"]
    next_id = _sonraki_tx_id(data)
    olusan_ids = []

    for lot in satir["proposed_lots"]:
        qty = round(float(lot["qty"]), 12)
        if qty <= 0:
            continue
        maliyet = round(float(lot["cost"]), 12)
        data["transactions"].append({
            "id": next_id,
            # Gerçek alım tarihi korunur — düzeltmenin bütün değeri bu.
            "date": lot["date"] or bugun,
            "coin": coin_adi,
            "exchange": exch,
            "qty": qty,
            "cost": maliyet,
            "status": ACTIVE_STATUS,
            "type": "DÜZELTME",
            "rebuild_in_id": rid,
            "cost_method": reconcile.REBUILD_COST_METHOD,
            "notes": f"Borsa mutabakatı ile yeniden kuruldu (#{rid}): "
                     f"{lot['date']} tarihli alım, {qty:.8f} @ ${maliyet:,.8f} — {exch}",
            "category": kategori,
        })
        olusan_ids.append(next_id)
        next_id += 1

    # --- Geçmişteki satışların gerçekleşmiş K/Z'si ---
    # FIFO'da hayatta kalan lotlar en son (ve düşen bir coinde en ucuz)
    # alımlardır; aradaki fark kaybolmadı, geçmiş satışlarda gerçekleşti.
    # Bunu yazmazsak düzeltme pozisyonu ucuzlatır ve tabloyu olduğundan
    # İYİ gösterir. Tek bir özet kayıt olarak yazılıyor: satış satış değil,
    # çünkü amaç işlem geçmişini içe aktarmak değil, sonucu doğru göstermek.
    kz_tx_id = None
    if satir.get("will_book_realized"):
        kz_qty = float(satir["realized_qty"])
        kz_maliyet = float(satir["realized_cost_usd"])
        kz_hasilat = float(satir["realized_proceeds_usd"])
        kz_tx_id = next_id
        data["transactions"].append({
            "id": kz_tx_id,
            "date": satir.get("coverage_start") or bugun,
            "coin": coin_adi,
            "exchange": exch,
            "qty": round(kz_qty, 12),
            "cost": round(kz_maliyet / kz_qty, 12) if kz_qty > 0 else 0.0,
            "status": CLOSED_STATUS,
            "type": "MUTABAKAT",
            "exit_price": round(kz_hasilat / kz_qty, 12) if kz_qty > 0 else 0.0,
            "exit_date": satir.get("realized_last_date") or bugun,
            "exit_value": round(kz_hasilat, 6),
            "realized_pnl_usd": round(kz_hasilat - kz_maliyet, 2),
            "fee_amount": 0.0,
            "fee_asset": "USDT",
            "fee_usd": 0.0,
            "cost_method": reconcile.REBUILD_COST_METHOD,
            "close_reason": "rebuild_realized",
            "rebuild_in_id": rid,
            "notes": f"Borsa mutabakatı (#{rid}): {exch} üzerinde kapanmış "
                     f"{kz_qty:,.8f} {satir['asset']} alım-satımının özeti — "
                     f"maliyet ${kz_maliyet:,.2f}, hasılat ${kz_hasilat:,.2f}. "
                     "Tek tek işlemler içe aktarılmadı; yalnızca sonuç yazıldı.",
            "category": kategori,
        })
        next_id += 1

    kayit = {
        "id": rid,
        "date": bugun,
        "pos_key": pos_key,
        "asset": satir["asset"],
        "coin": coin_adi,
        "exchange": exch,
        "before": {
            "qty": onceki_qty,
            "invested": round(onceki_maliyet, 6),
            "avg_cost": round(onceki_maliyet / onceki_qty, 12) if onceki_qty > 0 else 0.0,
            "lot_count": len(eski_lotlar),
        },
        "after": {
            "qty": satir["proposed_qty"],
            "invested": round(satir["proposed_invested"], 6),
            "avg_cost": round(satir["proposed_avg_cost"], 12),
            "lot_count": len(olusan_ids),
        },
        "diff_qty": satir["proposed_qty"] - onceki_qty,
        "diff_invested": round(satir["proposed_invested"] - onceki_maliyet, 6),
        "realized": {
            "booked": bool(kz_tx_id),
            "tx_id": kz_tx_id,
            "qty": satir.get("realized_qty", 0.0),
            "proceeds_usd": round(float(satir.get("realized_proceeds_usd") or 0.0), 2),
            "cost_usd": round(float(satir.get("realized_cost_usd") or 0.0), 2),
            "pnl_usd": round(float(satir.get("realized_pnl_usd") or 0.0), 2),
        },
        "closed_tx_ids": kapatilan_ids,
        "created_tx_ids": olusan_ids,
        "signature": satir["signature"],
        "coverage_start": satir.get("coverage_start"),
        "applied_warnings": satir.get("warnings", []),
        "sources": [k["name"] for k in plan.get("sources", [])],
        "note": (note or "").strip(),
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    data["rebuilds"].append(kayit)
    data["next_rebuild_id"] = rid + 1

    # Pozisyon tamamen kapandıysa oradaki hedef fiyat kaydı anlamsızlaşır.
    if not olusan_ids and pos_key in data.get("targets", {}):
        del data["targets"][pos_key]

    save_portfolio(data)
    return kayit


def undo_rebuild(rebuild_id: int):
    """Düzeltmeyi geri alır: oluşan lotları siler, eski lotları geri açar."""
    data = load_portfolio()
    kayit = next((r for r in data.get("rebuilds", [])
                  if int(r.get("id", 0) or 0) == int(rebuild_id)), None)
    if not kayit:
        raise ValueError(f"Düzeltme kaydı bulunamadı: {rebuild_id}")

    # Düzeltmeyle oluşan bir lot sonradan satıldıysa geri alma veriyi bozar.
    mevcut = {int(t.get("id", 0) or 0): t for t in data.get("transactions", [])}
    for tx_id in kayit.get("created_tx_ids", []):
        tx = mevcut.get(int(tx_id))
        if tx is None:
            raise ValueError("Bu düzeltme geri alınamaz: oluşan lotlardan biri silinmiş.")
        if tx.get("status") != ACTIVE_STATUS:
            raise ValueError("Bu düzeltme geri alınamaz: düzeltmeyle oluşan lotlardan biri "
                             "satılmış veya kapatılmış. Önce o işlemi geri alın.")

    silinecek = {int(i) for i in kayit.get("created_tx_ids", [])}
    # K/Z özeti de düzeltmeyle birlikte gitmeli; yoksa geri alma gerçekleşmiş
    # K/Z tablosunda sahibi olmayan bir kayıt bırakırdı.
    kz_id = (kayit.get("realized") or {}).get("tx_id")
    if kz_id:
        silinecek.add(int(kz_id))
    data["transactions"] = [t for t in data.get("transactions", [])
                            if int(t.get("id", 0) or 0) not in silinecek]

    mevcut = {int(t.get("id", 0) or 0): t for t in data["transactions"]}
    geri_gelen = 0
    for tx_id in kayit.get("closed_tx_ids", []):
        tx = mevcut.get(int(tx_id))
        if tx is None:
            continue
        tx["status"] = ACTIVE_STATUS
        tx.pop("rebuild_out_id", None)
        tx.pop("exit_date", None)
        if tx.get("close_reason") == REBUILD_CLOSE_REASON:
            tx.pop("close_reason", None)
        geri_gelen += 1

    data["rebuilds"] = [r for r in data.get("rebuilds", [])
                        if int(r.get("id", 0) or 0) != int(rebuild_id)]
    save_portfolio(data)
    return {"rebuild_id": int(rebuild_id),
            "restored_lots": geri_gelen,
            "removed_lots": len(silinecek)}


def list_rebuilds(data=None):
    """Düzeltme defterini yeniden eskiye doğru döndürür."""
    if data is None:
        data = load_portfolio()
    kayitlar = list(data.get("rebuilds", []))
    kayitlar.sort(key=lambda r: int(r.get("id", 0) or 0), reverse=True)
    return kayitlar


# =====================================================================
# FAZ 5: SETTINGS & HEALTH CHECK MANAGER
# =====================================================================
SETTINGS_FILE = os.path.join(DATA_DIR, "settings.json")

DEFAULT_SETTINGS = {
    "api_urls": {
        "binance_ticker": "https://api.binance.com/api/v3/ticker/24hr",
        "binance_ping": "https://api.binance.com/api/v3/ping",
        "mexc_ticker": "https://api.mexc.com/api/v3/ticker/24hr",
        "mexc_ping": "https://api.mexc.com/api/v3/ping",
        "whitebit_ticker": "https://whitebit.com/api/v4/public/ticker",
        "gateio_ticker": "https://api.gateio.ws/api/v4/spot/tickers",
        "dex_screener": "https://api.dexscreener.com/latest/dex/search"
    },
    # Fiyat kaynağı kayıt defteri (FAZ B++).
    # Hangi kademe açık ve hangi sırada denenecek. Kullanıcı arayüzden değiştirir.
    # Sözlük olarak tutulur ki yeni bir kaynak eklendiğinde load_settings'in
    # sığ birleştirmesi onu kullanıcının kayıtlı tercihlerini bozmadan eklesin.
    "price_sources": {
        "binance": {"enabled": True, "order": 1},
        "mexc": {"enabled": True, "order": 2},
        "whitebit": {"enabled": True, "order": 3},
        "gateio": {"enabled": False, "order": 4},
        "dex": {"enabled": True, "order": 5}
    },
    # Sembole özel kaynak tanımları. Kayıt defteri bir coini bulamadığında
    # (ya da yanlış coini bulduğunda) kullanıcı burada kaynağı sabitler.
    # Varsayılan olarak BOŞTUR — hiçbir kullanıcının coini koda gömülmez.
    #   { "SCM": {"type": "cex", "source": "whitebit", "market": "SCM_USDT"} }
    #   { "RDNT": {"type": "dex", "query": "RDNT"} }
    #   { "XYZ":  {"type": "manual", "price": 0.0012} }
    "symbol_sources": {},
    "api_keys": {
        "gemini_api_key": "",
        "telegram_bot_token": "",
        "telegram_chat_id": ""
    },
    "preferences": {
        "refresh_interval_sec": 3.5,
        "default_tab": "dashboard",
        "sound_alerts": True,
        "theme": "dark"
    },
    "security": {
        "pin_enabled": False,
        "pin_hash": "",
        "salt": "",
        "auto_lock_minutes": 15,
        "privacy_mode": False
    }
}

def _obfuscate_key(plaintext: str) -> str:
    """Base64 ile API anahtarını obfuscate eder."""
    if not plaintext:
        return ""
    return base64.b64encode(plaintext.encode("utf-8")).decode("utf-8")

def _deobfuscate_key(encoded: str) -> str:
    """Base64 obfuscation'ı çözer. Düz metin ise olduğu gibi döner (geriye uyumluluk)."""
    if not encoded:
        return ""
    try:
        decoded = base64.b64decode(encoded).decode("utf-8")
        return decoded
    except Exception:
        return encoded  # Eski açık metin format — olduğu gibi döndür

_OBFUSCATED_KEYS = ["gemini_api_key", "telegram_bot_token"]

def load_settings():
    ensure_data_dir()
    # ÖNEMLİ: deepcopy şart. dict() sığ kopya yapar ve iç içe sözlükler
    # (security, api_keys...) DEFAULT_SETTINGS ile paylaşılır. Çağıran taraf
    # bunları değiştirdiğinde modül sabiti kalıcı olarak kirlenir — örneğin
    # yeni kurulumda ilk set_pin() çağrısı PIN hash'ini global şablona yazardı.
    if not os.path.exists(SETTINGS_FILE):
        save_settings(DEFAULT_SETTINGS)
        return copy.deepcopy(DEFAULT_SETTINGS)
    try:
        with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            # Merge with default settings in case new keys were added
            merged = copy.deepcopy(DEFAULT_SETTINGS)
            for k in DEFAULT_SETTINGS:
                if k in data and isinstance(data[k], dict):
                    merged[k] = {**DEFAULT_SETTINGS[k], **data[k]}
                elif k in data:
                    merged[k] = data[k]
            # Deobfuscate API keys for runtime use
            api_keys = merged.get("api_keys", {})
            for key_name in _OBFUSCATED_KEYS:
                if key_name in api_keys:
                    api_keys[key_name] = _deobfuscate_key(api_keys[key_name])
            merged["api_keys"] = api_keys
            return merged
    except Exception as e:
        logger.error("settings.json okunamadı, varsayılan ayarlarla devam ediliyor: %s", e)
        return copy.deepcopy(DEFAULT_SETTINGS)

def merge_settings(partial: dict) -> dict:
    """
    Kısmi ayar güncellemesini mevcut dosyayla birleştirir.

    KRİTİK — gerçek veri kaybı buradan çıktı: `POST /api/settings` gelen gövdeyi
    dosyanın TAMAMININ yerine yazıyordu. Arayüz yalnızca `api_urls`, `api_keys`
    ve `preferences` gönderdiği için, kullanıcı Ayarlar ekranında "Kaydet"e her
    bastığında `security` bölümü (PIN hash'i, salt, kurtarma anahtarı) ve
    fiyat kaynağı tanımları sessizce siliniyordu. PIN koruması kendiliğinden
    kapanıyordu ve kullanıcı bunu göremiyordu.

    Artık yalnızca gönderilen bölümler güncellenir; gönderilmeyen her şey
    olduğu gibi korunur.
    """
    current = load_settings()
    if not isinstance(partial, dict):
        return current

    for key, value in partial.items():
        if isinstance(value, dict) and isinstance(current.get(key), dict):
            merged = dict(current[key])
            merged.update(value)
            current[key] = merged
        else:
            current[key] = value
    return current


def save_settings(new_settings):
    ensure_data_dir()
    # Deep copy to avoid mutating the input
    to_save = json.loads(json.dumps(new_settings))
    # Obfuscate API keys before writing to disk
    api_keys = to_save.get("api_keys", {})
    for key_name in _OBFUSCATED_KEYS:
        if key_name in api_keys and api_keys[key_name]:
            api_keys[key_name] = _obfuscate_key(api_keys[key_name])
    to_save["api_keys"] = api_keys
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(to_save, f, indent=2, ensure_ascii=False)
    return True


# =====================================================================
# FAZ B++: FİYAT KAYNAĞI YÖNETİMİ
# =====================================================================
VALID_SOURCE_TYPES = ("cex", "dex", "manual")


def normalize_symbol_key(symbol: str) -> str:
    """
    Sembolü kaynak defterinde kullanılan anahtara indirger.

    Kullanıcı "SCM", "SCMUSDT" veya "scm_usdt" yazsa da aynı kayda düşsün diye
    ayraçlar atılır ve USDT eki kaldırılır.
    """
    s = str(symbol or "").upper().replace("_", "").replace("-", "").replace("/", "").strip()
    if s.endswith("USDT") and len(s) > 4:
        s = s[:-4]
    return s


def validate_source_spec(spec: dict) -> tuple:
    """
    Kaynak tanımını doğrular. (temiz_tanim, hata_mesaji) döndürür.
    Hata varsa temiz_tanim None olur.
    """
    if not isinstance(spec, dict):
        return None, "Kaynak tanımı bir nesne olmalı."

    stype = str(spec.get("type", "")).lower().strip()
    if stype not in VALID_SOURCE_TYPES:
        return None, f"Geçersiz kaynak türü: '{stype}'. Beklenen: {', '.join(VALID_SOURCE_TYPES)}."

    if stype == "cex":
        source = str(spec.get("source", "")).lower().strip()
        market = str(spec.get("market", "")).strip()
        if not source:
            return None, "Merkezi borsa kaynağı için borsa seçilmeli."
        if not market:
            return None, "Merkezi borsa kaynağı için market adı gerekli (örn. SCM_USDT)."
        return {"type": "cex", "source": source, "market": market}, None

    if stype == "dex":
        query = str(spec.get("query") or spec.get("contract") or "").strip()
        if not query:
            return None, "Zincir üstü kaynak için kontrat adresi veya sembol gerekli."
        return {"type": "dex", "query": query}, None

    # manual
    try:
        price = float(spec.get("price", 0) or 0)
    except (ValueError, TypeError):
        return None, "Manuel fiyat sayı olmalı."
    if price <= 0:
        return None, "Manuel fiyat sıfırdan büyük olmalı."
    return {"type": "manual", "price": price}, None


def get_symbol_sources() -> dict:
    return (load_settings() or {}).get("symbol_sources") or {}


def set_symbol_source(symbol: str, spec: dict) -> tuple:
    """
    Bir sembol için fiyat kaynağını kaydeder. (basarili, hata) döndürür.
    """
    key = normalize_symbol_key(symbol)
    if not key:
        return False, "Sembol boş olamaz."

    clean, err = validate_source_spec(spec)
    if err:
        return False, err

    settings = load_settings()
    sources = dict(settings.get("symbol_sources") or {})
    sources[key] = clean
    settings["symbol_sources"] = sources
    save_settings(settings)
    logger.info("Fiyat kaynağı tanımlandı: %s → %s", key, clean)
    return True, None


def delete_symbol_source(symbol: str) -> bool:
    key = normalize_symbol_key(symbol)
    settings = load_settings()
    sources = dict(settings.get("symbol_sources") or {})
    if key not in sources:
        return False
    sources.pop(key)
    settings["symbol_sources"] = sources
    save_settings(settings)
    logger.info("Fiyat kaynağı kaldırıldı: %s", key)
    return True


def set_price_sources(registry: dict) -> tuple:
    """
    Kademe kayıt defterini günceller (hangi kaynak açık, hangi sırada).
    """
    if not isinstance(registry, dict) or not registry:
        return False, "Kaynak listesi boş olamaz."

    settings = load_settings()
    current = dict(settings.get("price_sources") or {})
    for sid, row in registry.items():
        if not isinstance(row, dict):
            continue
        base = current.get(sid, {"enabled": False, "order": 99})
        try:
            order = int(row.get("order", base.get("order", 99)))
        except (ValueError, TypeError):
            order = base.get("order", 99)
        current[sid] = {"enabled": bool(row.get("enabled", base.get("enabled", False))), "order": order}

    if not any(r.get("enabled") for r in current.values()):
        return False, "En az bir fiyat kaynağı açık kalmalı."

    settings["price_sources"] = current
    save_settings(settings)
    logger.info("Fiyat kaynağı kayıt defteri güncellendi.")
    return True, None


# =====================================================================
# FAZ 8: SECURITY & PIN AUTHENTICATION HELPERS
# =====================================================================
def hash_pin(pin: str, salt: str = None) -> tuple:
    if not salt:
        salt = secrets.token_hex(16)
    combined = f"{salt}:{pin}".encode("utf-8")
    h = hashlib.sha256(combined).hexdigest()
    return h, salt

def verify_pin(pin: str) -> bool:
    settings = load_settings()
    sec = settings.get("security", {})
    if not sec.get("pin_enabled", False):
        return True
    stored_hash = sec.get("pin_hash", "")
    stored_salt = sec.get("salt", "")
    if not stored_hash or not stored_salt:
        return True
    computed_hash, _ = hash_pin(str(pin).strip(), stored_salt)
    return hmac.compare_digest(stored_hash, computed_hash)

def set_pin(pin: str, auto_lock_minutes: int = 15) -> dict:
    """PIN kurulumu yapar ve 12 haneli kurtarma anahtarı üretir."""
    settings = load_settings()
    h, salt = hash_pin(str(pin).strip())
    sec = settings.get("security", {})
    sec["pin_enabled"] = True
    sec["pin_hash"] = h
    sec["salt"] = salt
    sec["auto_lock_minutes"] = int(auto_lock_minutes or 15)

    # Recovery key üret (ilk kurulumda veya her PIN kurulumunda yeniden üret)
    recovery_key = secrets.token_hex(6).upper()  # 12 hex karakter
    recovery_hash, recovery_salt = hash_pin(recovery_key)
    sec["recovery_hash"] = recovery_hash
    sec["recovery_salt"] = recovery_salt

    settings["security"] = sec
    save_settings(settings)
    return {"success": True, "recovery_key": recovery_key}

def verify_recovery_key(key: str) -> bool:
    """Kurtarma anahtarını doğrular."""
    settings = load_settings()
    sec = settings.get("security", {})
    stored_hash = sec.get("recovery_hash", "")
    stored_salt = sec.get("recovery_salt", "")
    if not stored_hash or not stored_salt:
        return False
    computed_hash, _ = hash_pin(str(key).strip().upper(), stored_salt)
    return hmac.compare_digest(stored_hash, computed_hash)

def reset_pin_with_recovery(recovery_key: str, new_pin: str, auto_lock_minutes: int = 15) -> dict:
    """Kurtarma anahtarı ile PIN sıfırlar ve yeni recovery key üretir."""
    if not verify_recovery_key(recovery_key):
        return {"success": False, "error": "Geçersiz kurtarma anahtarı."}
    return set_pin(new_pin, auto_lock_minutes)

def disable_pin(current_pin: str) -> bool:
    if not verify_pin(current_pin):
        return False
    settings = load_settings()
    sec = settings.get("security", {})
    sec["pin_enabled"] = False
    sec["pin_hash"] = ""
    sec["salt"] = ""
    settings["security"] = sec
    save_settings(settings)
    return True

def change_pin(current_pin: str, new_pin: str) -> bool:
    if not verify_pin(current_pin):
        return False
    settings = load_settings()
    h, salt = hash_pin(str(new_pin).strip())
    sec = settings.get("security", {})
    sec["pin_enabled"] = True
    sec["pin_hash"] = h
    sec["salt"] = salt
    settings["security"] = sec
    save_settings(settings)
    return True

def update_security_settings(auto_lock_minutes: int = None, privacy_mode: bool = None) -> dict:
    settings = load_settings()
    sec = settings.get("security", {})
    if auto_lock_minutes is not None:
        sec["auto_lock_minutes"] = int(auto_lock_minutes)
    if privacy_mode is not None:
        sec["privacy_mode"] = bool(privacy_mode)
    settings["security"] = sec
    save_settings(settings)
    return sec

def ping_services():
    import urllib.request
    import time
    settings = load_settings()
    urls = settings.get("api_urls", DEFAULT_SETTINGS["api_urls"])

    targets = [
        {"id": "binance", "name": "Binance Spot API", "url": urls.get("binance_ping", "https://api.binance.com/api/v3/ping")},
        {"id": "mexc", "name": "MEXC Global API", "url": urls.get("mexc_ping", "https://api.mexc.com/api/v3/ping")},
        {"id": "dex", "name": "DexScreener On-Chain API", "url": "https://api.dexscreener.com/latest/dex/search?q=BTC"}
    ]

    results = {}
    for t in targets:
        t0 = time.time()
        try:
            req = urllib.request.Request(
                t["url"],
                headers={
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                    "Accept": "application/json"
                }
            )
            with urllib.request.urlopen(req, timeout=4) as resp:
                code = resp.getcode()
                ms = round((time.time() - t0) * 1000)
                results[t["id"]] = {
                    "name": t["name"],
                    "status": "online",
                    "code": code,
                    "ms": ms,
                    "checked_at": datetime.now().strftime("%H:%M:%S")
                }
        except Exception as e:
            ms = round((time.time() - t0) * 1000)
            results[t["id"]] = {
                "name": t["name"],
                "status": "error",
                "code": 500,
                "ms": ms,
                "error": str(e),
                "checked_at": datetime.now().strftime("%H:%M:%S")
            }
    return results


# =============================================================
# FAZ 7: GERÇEKLEŞMİŞ KÂR/ZARAR (REALIZED PNL) VE İŞLEM KARNESİ
# =============================================================
def calculate_realized_metrics(data):
    transactions = data.get("transactions", [])
    closed_txs = []
    total_realized_pnl = 0.0
    total_profit = 0.0
    total_loss = 0.0
    total_fees_usd = 0.0
    win_count = 0
    loss_count = 0
    
    best_trade = None
    worst_trade = None
    # FAZ F1 — Yazımlar gerçek birer zarardır ve toplam gerçekleşmiş K/Z'ye
    # dahildir. Ancak bir alım-satım kararının sonucu ile ölmüş bir coinin
    # silinmesi aynı şey değil; kullanıcı ayrımı görebilsin diye ayrıca sayılır.
    total_write_off_usd = 0.0
    write_off_count = 0
    monthly_map = defaultdict(lambda: {"month": "", "pnl_usd": 0.0, "profit_usd": 0.0, "loss_usd": 0.0, "fees_usd": 0.0, "wins": 0, "total": 0})

    for t in transactions:
        fee_usd = float(t.get("fee_usd") or 0.0)

        status = t.get("status", "")
        realized_pnl = t.get("realized_pnl_usd")
        has_exit_price = t.get("exit_price") is not None
        
        # Sadece gerçek satış kaydı olan (exit_price veya realized_pnl_usd içeren) işlemleri hesaba kat
        is_realized_trade = has_exit_price or (realized_pnl is not None)
        if not is_realized_trade:
            continue
            
        total_fees_usd += fee_usd
        coin_name = t.get("coin") or t.get("symbol") or "Bilinmeyen"
        qty = float(t.get("qty") or t.get("amount") or 0.0)
        entry_price = float(t.get("cost") or t.get("buy_price") or 0.0)
        entry_cost = qty * entry_price
        
        exit_price = float(t.get("exit_price", entry_price))
        exit_val = float(t.get("exit_value", qty * exit_price))
        
        if realized_pnl is not None:
            pnl_usd = float(realized_pnl)
        else:
            pnl_usd = (exit_val - entry_cost) - fee_usd
            
        pnl_pct = ((pnl_usd / entry_cost) * 100.0) if entry_cost > 0 else 0.0
        
        exit_date = t.get("exit_date") or t.get("date") or datetime.now().strftime("%Y-%m-%d")
        month_key = exit_date[:7] if len(exit_date) >= 7 else datetime.now().strftime("%Y-%m")
        
        enriched_tx = {
            "id": t.get("id"),
            "coin": coin_name,
            "symbol": coin_name,
            "display_name": coin_name,
            "exchange": t.get("exchange", "BINANCE"),
            "category": t.get("category", "Altcoin"),
            "type": t.get("type", "ALIM"),
            "amount": qty,
            "qty": qty,
            "entry_price": entry_price,
            "cost": entry_price,
            "entry_cost": entry_cost,
            "exit_price": exit_price,
            "exit_value": exit_val,
            "entry_date": t.get("date", ""),
            "exit_date": exit_date,
            "realized_pnl_usd": round(pnl_usd, 2),
            "realized_pnl_pct": round(pnl_pct, 2),
            "fee_amount": float(t.get("fee_amount") or 0.0),
            "fee_asset": t.get("fee_asset", "USDT"),
            "fee_usd": round(fee_usd, 4),
            "cost_method": t.get("cost_method", "Konsolide Ortalama"),
            "close_reason": t.get("close_reason"),
            "write_off_reason": t.get("write_off_reason"),
            "write_off_reason_label": WRITE_OFF_REASONS.get(t.get("write_off_reason") or ""),
            "notes": t.get("notes", "")
        }
        closed_txs.append(enriched_tx)

        if t.get("close_reason") == "write_off":
            total_write_off_usd += pnl_usd
            write_off_count += 1

        total_realized_pnl += pnl_usd
        if pnl_usd > 0:
            total_profit += pnl_usd
            win_count += 1
            if best_trade is None or pnl_usd > best_trade["realized_pnl_usd"]:
                best_trade = enriched_tx
        elif pnl_usd < 0:
            total_loss += abs(pnl_usd)
            loss_count += 1
            if worst_trade is None or pnl_usd < worst_trade["realized_pnl_usd"]:
                worst_trade = enriched_tx
                
        monthly_map[month_key]["month"] = month_key
        monthly_map[month_key]["pnl_usd"] += pnl_usd
        monthly_map[month_key]["fees_usd"] += fee_usd
        if pnl_usd > 0:
            monthly_map[month_key]["profit_usd"] += pnl_usd
            monthly_map[month_key]["wins"] += 1
        elif pnl_usd < 0:
            monthly_map[month_key]["loss_usd"] += abs(pnl_usd)
        monthly_map[month_key]["total"] += 1
        
    total_closed = len(closed_txs)
    if win_count > 0:
        win_rate_pct = round((win_count / total_closed * 100.0), 1)
    else:
        win_rate_pct = 0.0
        
    if total_loss > 0:
        profit_factor = round((total_profit / total_loss), 2)
    elif total_profit > 0:
        profit_factor = round(total_profit, 2)
    else:
        profit_factor = 0.0
    
    monthly_list = sorted(monthly_map.values(), key=lambda x: x["month"], reverse=True)
    for m in monthly_list:
        m["pnl_usd"] = round(m["pnl_usd"], 2)
        m["profit_usd"] = round(m["profit_usd"], 2)
        m["loss_usd"] = round(m["loss_usd"], 2)
        m["fees_usd"] = round(m["fees_usd"], 2)
        m["win_rate_pct"] = round((m["wins"] / m["total"] * 100.0), 1) if (m["total"] > 0 and m["wins"] > 0) else 0.0

    return {
        "total_realized_pnl_usd": round(total_realized_pnl, 2),
        "total_profit_usd": round(total_profit, 2),
        "total_loss_usd": round(total_loss, 2),
        "total_fees_usd": round(total_fees_usd, 2),
        # Yazımlar hariç, yalnızca alım-satım kararlarından gelen sonuç.
        "total_write_off_usd": round(total_write_off_usd, 2),
        "write_off_count": write_off_count,
        "trading_realized_pnl_usd": round(total_realized_pnl - total_write_off_usd, 2),
        "closed_tx_count": total_closed,
        "winning_tx_count": win_count,
        "losing_tx_count": loss_count,
        "win_rate_pct": win_rate_pct,
        "profit_factor": profit_factor,
        "best_trade": best_trade,
        "worst_trade": worst_trade,
        "monthly_breakdown": monthly_list,
        "closed_transactions": sorted(closed_txs, key=lambda x: x.get("exit_date", ""), reverse=True)
    }


def format_excel_date(d_str):
    if not d_str:
        return ""
    try:
        d_clean = str(d_str).strip()
        if "-" in d_clean:
            parts = d_clean.split("-")
            if len(parts) == 3:
                if len(parts[0]) == 4:
                    return f"{parts[2].zfill(2)}.{parts[1].zfill(2)}.{parts[0]}"
                return f"{parts[0].zfill(2)}.{parts[1].zfill(2)}.{parts[2]}"
        if "/" in d_clean:
            parts = d_clean.split("/")
            if len(parts) == 3:
                if len(parts[2]) == 4:
                    return f"{parts[0].zfill(2)}.{parts[1].zfill(2)}.{parts[2]}"
                if len(parts[0]) == 4:
                    return f"{parts[2].zfill(2)}.{parts[1].zfill(2)}.{parts[0]}"
                return f"{parts[0].zfill(2)}.{parts[1].zfill(2)}.20{parts[2]}"
        return d_clean
    except Exception:
        return str(d_str)


# =============================================================
# FAZ 7: ÇOK SAYFALI PROFESYONEL EXCEL (.XLSX) DIŞA AKTARIMI
# =============================================================
def export_portfolio_excel(data, live_prices):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    metrics = calculate_portfolio_metrics(data, live_prices)
    coins = metrics.get("consolidated_coins", [])
    kpis = metrics.get("kpis", {})
    transactions = data.get("transactions", [])

    font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Segoe UI", size=9, italic=True, color="94A3B8")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
    font_regular = Font(name="Segoe UI", size=9.5, color="1E293B")
    font_code = Font(name="Consolas", size=9.5, color="0F172A")
    font_green = Font(name="Segoe UI", size=9.5, bold=True, color="059669")
    font_red = Font(name="Segoe UI", size=9.5, bold=True, color="DC2626")

    fill_title = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_header_slate = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_header_emerald = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
    fill_header_purple = PatternFill(start_color="581C87", end_color="581C87", fill_type="solid")
    fill_header_blue = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="CBD5E1")
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_total = Border(top=Side(border_style="thin", color="64748B"), bottom=Side(border_style="double", color="0F172A"))

    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # SHEET 1: Konsolide Portföy
    ws1 = wb.create_sheet(title="Konsolide Portfoy")
    ws1.views.sheetView[0].showGridLines = True

    ws1.merge_cells("A1:K1")
    ws1["A1"] = "COINTAKIP - CANLI KRIPTO PORTFOY VE POZISYON RAPORU"
    ws1["A1"].font = font_title
    ws1["A1"].fill = fill_title
    ws1["A1"].alignment = align_center
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells("A2:K2")
    ws1["A2"] = f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Toplam Varlik: ${kpis.get('total_kasa', 0):,.2f} | Serbest Nakit: ${kpis.get('usdt_cash', 0):,.2f} USDT"
    ws1["A2"].font = font_subtitle
    ws1["A2"].fill = fill_title
    ws1["A2"].alignment = align_center
    ws1.row_dimensions[2].height = 18

    headers1 = [
        "Varlık", "Borsa", "Kategori", "Miktar", "Ort. Maliyet ($)", 
        "Canlı Fiyat ($)", "Toplam Yatırım ($)", "Güncel Değer ($)", 
        "Net K/Z ($)", "Getiri (%)", "Portföy Payı (%)"
    ]
    for col_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header_slate
        cell.alignment = align_center
        cell.border = border_cell
    ws1.row_dimensions[4].height = 24

    row_curr = 5
    for c in coins:
        ws1.cell(row=row_curr, column=1, value=c.get("display_name", c.get("symbol"))).alignment = align_left
        ws1.cell(row=row_curr, column=2, value=c.get("exchange", "BINANCE")).alignment = align_center
        ws1.cell(row=row_curr, column=3, value=c.get("category", "Altcoin")).alignment = align_center
        
        c_qty = ws1.cell(row=row_curr, column=4, value=float(c.get("total_qty") or c.get("qty") or 0))
        c_qty.number_format = "#,##0.0000"
        c_qty.alignment = align_right

        c_cost = ws1.cell(row=row_curr, column=5, value=float(c.get("avg_cost", 0)))
        c_cost.number_format = "$#,##0.0000" if float(c.get("avg_cost", 0)) < 1 else "$#,##0.00"
        c_cost.alignment = align_right

        c_price = ws1.cell(row=row_curr, column=6, value=float(c.get("live_price", 0)))
        c_price.number_format = "$#,##0.0000" if float(c.get("live_price", 0)) < 1 else "$#,##0.00"
        c_price.alignment = align_right

        c_inv = ws1.cell(row=row_curr, column=7, value=float(c.get("total_invested", 0)))
        c_inv.number_format = "$#,##0.00"
        c_inv.alignment = align_right

        c_val = ws1.cell(row=row_curr, column=8, value=float(c.get("current_value", 0)))
        c_val.number_format = "$#,##0.00"
        c_val.alignment = align_right

        pnl_val = float(c.get("pnl_usd", 0))
        c_pnl = ws1.cell(row=row_curr, column=9, value=pnl_val)
        c_pnl.number_format = "+$#,##0.00;-$#,##0.00;$0.00"
        c_pnl.font = font_green if pnl_val >= 0 else font_red
        c_pnl.alignment = align_right

        pnl_pct_val = float(c.get("pnl_pct", 0)) / 100.0
        c_pct = ws1.cell(row=row_curr, column=10, value=pnl_pct_val)
        c_pct.number_format = "+0.0%;-0.0%;0.0%"
        c_pct.font = font_green if pnl_pct_val >= 0 else font_red
        c_pct.alignment = align_right

        share_val = float(c.get("portfolio_share_pct", 0)) / 100.0
        c_share = ws1.cell(row=row_curr, column=11, value=share_val)
        c_share.number_format = "0.0%"
        c_share.alignment = align_right

        for col_idx in range(1, 12):
            cell = ws1.cell(row=row_curr, column=col_idx)
            cell.border = border_cell
            if col_idx not in [9, 10]:
                cell.font = font_code if col_idx in [1, 2] else font_regular
            if row_curr % 2 == 0:
                cell.fill = fill_zebra

        ws1.row_dimensions[row_curr].height = 20
        row_curr += 1

    tot_row = row_curr
    ws1.cell(row=tot_row, column=1, value="GENEL TOPLAM").font = font_bold
    ws1.cell(row=tot_row, column=1).alignment = align_left
    ws1.cell(row=tot_row, column=7, value=f"=SUM(G5:G{tot_row-1})").number_format = "$#,##0.00"
    ws1.cell(row=tot_row, column=7).font = font_bold
    ws1.cell(row=tot_row, column=8, value=f"=SUM(H5:H{tot_row-1})").number_format = "$#,##0.00"
    ws1.cell(row=tot_row, column=8).font = font_bold
    ws1.cell(row=tot_row, column=9, value=f"=SUM(I5:I{tot_row-1})").number_format = "+$#,##0.00;-$#,##0.00;$0.00"
    ws1.cell(row=tot_row, column=9).font = font_bold
    ws1.cell(row=tot_row, column=11, value="100.0%").font = font_bold
    ws1.cell(row=tot_row, column=11).alignment = align_right

    for col_idx in range(1, 12):
        cell = ws1.cell(row=tot_row, column=col_idx)
        cell.border = border_total
        cell.fill = fill_total
    ws1.row_dimensions[tot_row].height = 24

    # SHEET 2: DCA Alım Defteri
    ws2 = wb.create_sheet(title="DCA Alim Defteri")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:J1")
    ws2["A1"] = "DCA ISLEM GECMISI VE ALIM DEFTERI"
    ws2["A1"].font = font_title
    ws2["A1"].fill = fill_title
    ws2["A1"].alignment = align_center
    ws2.row_dimensions[1].height = 30

    headers2 = ["Tarih", "Varlık", "Borsa", "Tür", "Miktar", "Birim Fiyat ($)", "Toplam Maliyet ($)", "Komisyon ($)", "Kategori", "Durum"]
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header_emerald
        cell.alignment = align_center
        cell.border = border_cell
    ws2.row_dimensions[3].height = 24

    row_curr = 4
    for t in transactions:
        t_coin = t.get("coin") or t.get("symbol") or "Bilinmeyen"
        t_qty = float(t.get("qty") or t.get("amount") or 0.0)
        t_cost = float(t.get("cost") or t.get("buy_price") or 0.0)
        t_tot = t_qty * t_cost
        t_fee = float(t.get("fee_usd") or 0.0)

        ws2.cell(row=row_curr, column=1, value=format_excel_date(t.get("date", ""))).alignment = align_center
        ws2.cell(row=row_curr, column=2, value=t_coin).alignment = align_left
        ws2.cell(row=row_curr, column=3, value=t.get("exchange", "BINANCE")).alignment = align_center
        ws2.cell(row=row_curr, column=4, value=t.get("type", "ALIM")).alignment = align_center

        c_amt = ws2.cell(row=row_curr, column=5, value=t_qty)
        c_amt.number_format = "#,##0.0000"
        c_amt.alignment = align_right

        c_bp = ws2.cell(row=row_curr, column=6, value=t_cost)
        c_bp.number_format = "$#,##0.0000" if t_cost < 1 else "$#,##0.00"
        c_bp.alignment = align_right

        c_tot_cell = ws2.cell(row=row_curr, column=7, value=t_tot)
        c_tot_cell.number_format = "$#,##0.00"
        c_tot_cell.alignment = align_right

        c_fee_cell = ws2.cell(row=row_curr, column=8, value=t_fee)
        c_fee_cell.number_format = "$#,##0.00"
        c_fee_cell.alignment = align_right

        ws2.cell(row=row_curr, column=9, value=t.get("category", "Altcoin")).alignment = align_center
        ws2.cell(row=row_curr, column=10, value=t.get("status", "Aktif")).alignment = align_center

        for col_idx in range(1, 11):
            cell = ws2.cell(row=row_curr, column=col_idx)
            cell.border = border_cell
            cell.font = font_regular
            if row_curr % 2 == 1:
                cell.fill = fill_zebra
        ws2.row_dimensions[row_curr].height = 20
        row_curr += 1

    # SHEET 3: Gerçekleşmiş K/Z & Kapanan İşlemler
    ws3 = wb.create_sheet(title="Gerceklesmis KZ")
    ws3.views.sheetView[0].showGridLines = True

    ws3.merge_cells("A1:J1")
    ws3["A1"] = "GERCEKLESMIS KAR/ZARAR VE ISLEM KARNESI"
    ws3["A1"].font = font_title
    ws3["A1"].fill = fill_title
    ws3["A1"].alignment = align_center
    ws3.row_dimensions[1].height = 30

    headers3 = ["Alış Tarihi", "Satış Tarihi", "Varlık", "Borsa", "Miktar", "Alış Fiyatı ($)", "Satış Fiyatı ($)", "Komisyon ($)", "Net K/Z ($)", "Getiri (%)"]
    for col_idx, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header_purple
        cell.alignment = align_center
        cell.border = border_cell
    ws3.row_dimensions[3].height = 24

    closed_txs = [t for t in transactions if t.get("status") == "Kapandı / İzleme" or t.get("realized_pnl_usd") is not None]
    row_curr = 4
    if not closed_txs:
        ws3.merge_cells("A4:J4")
        ws3["A4"] = "Henüz kapatılmış / kâr realize edilmiş bir işlem kaydı bulunmuyor."
        ws3["A4"].alignment = align_center
        ws3["A4"].font = font_subtitle
    else:
        for ct in closed_txs:
            c_coin = ct.get("coin") or ct.get("symbol") or "Bilinmeyen"
            c_qty_val = float(ct.get("qty") or ct.get("amount") or 0.0)
            c_buy_p = float(ct.get("cost") or ct.get("buy_price") or 0.0)
            c_cost_tot = c_qty_val * c_buy_p
            c_sell_p = float(ct.get("exit_price", c_buy_p))
            c_fee_usd = float(ct.get("fee_usd") or 0.0)
            
            pnl_val = float(ct.get("realized_pnl_usd", 0.0)) if ct.get("realized_pnl_usd") is not None else (((c_sell_p - c_buy_p) * c_qty_val) - c_fee_usd)

            ws3.cell(row=row_curr, column=1, value=format_excel_date(ct.get("date", ""))).alignment = align_center
            ws3.cell(row=row_curr, column=2, value=format_excel_date(ct.get("exit_date", ct.get("date", "")))).alignment = align_center
            ws3.cell(row=row_curr, column=3, value=c_coin).alignment = align_left
            ws3.cell(row=row_curr, column=4, value=ct.get("exchange", "BINANCE")).alignment = align_center

            c_amt = ws3.cell(row=row_curr, column=5, value=c_qty_val)
            c_amt.number_format = "#,##0.0000"
            c_amt.alignment = align_right

            c_bp = ws3.cell(row=row_curr, column=6, value=c_buy_p)
            c_bp.number_format = "$#,##0.0000" if c_buy_p < 1 else "$#,##0.00"
            c_bp.alignment = align_right

            c_sp = ws3.cell(row=row_curr, column=7, value=c_sell_p)
            c_sp.number_format = "$#,##0.0000" if c_sell_p < 1 else "$#,##0.00"
            c_sp.alignment = align_right

            c_fee_cell = ws3.cell(row=row_curr, column=8, value=c_fee_usd)
            c_fee_cell.number_format = "$#,##0.00"
            c_fee_cell.alignment = align_right

            c_rpnl = ws3.cell(row=row_curr, column=9, value=pnl_val)
            c_rpnl.number_format = "+$#,##0.00;-$#,##0.00;$0.00"
            c_rpnl.font = font_green if pnl_val >= 0 else font_red
            c_rpnl.alignment = align_right

            pnl_pct = (pnl_val / c_cost_tot) if c_cost_tot > 0 else 0.0
            c_rpct = ws3.cell(row=row_curr, column=10, value=pnl_pct)
            c_rpct.number_format = "+0.0%;-0.0%;0.0%"
            c_rpct.font = font_green if pnl_val >= 0 else font_red
            c_rpct.alignment = align_right

            for col_idx in range(1, 11):
                cell = ws3.cell(row=row_curr, column=col_idx)
                cell.border = border_cell
                if col_idx not in [9, 10]:
                    cell.font = font_regular
                if row_curr % 2 == 1:
                    cell.fill = fill_zebra
            ws3.row_dimensions[row_curr].height = 20
            row_curr += 1

    # SHEET 4: Borsa & Kasa Dağılımı
    ws4 = wb.create_sheet(title="Borsa ve Kasa")
    ws4.views.sheetView[0].showGridLines = True

    ws4.merge_cells("A1:F1")
    ws4["A1"] = "BORSA BAZLI SERMAYE VE SERBEST NAKIT KASALARI"
    ws4["A1"].font = font_title
    ws4["A1"].fill = fill_title
    ws4["A1"].alignment = align_center
    ws4.row_dimensions[1].height = 30

    headers4 = ["Borsa / Platform", "Spot Yatırım ($)", "Spot Güncel Değer ($)", "Serbest Nakit (USDT)", "Toplam Borsa Varlığı ($)", "Kasa Payı (%)"]
    for col_idx, h in enumerate(headers4, 1):
        cell = ws4.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header_blue
        cell.alignment = align_center
        cell.border = border_cell
    ws4.row_dimensions[3].height = 24

    ex_kpis = metrics.get("exchange_kpis", {})
    tot_equity = float(kpis.get("total_kasa", 0.0)) or 1.0
    row_curr = 4
    
    # Detail rows for each exchange (exclude 'ALL')
    for ex_name, ex_data in ex_kpis.items():
        if ex_name == "ALL":
            continue
        ex_spot_cost = float(ex_data.get("spot_invested", 0))
        ex_spot_val = float(ex_data.get("spot_current_value", 0))
        ex_cash = float(ex_data.get("usdt_cash", 0))
        ex_tot = float(ex_data.get("total_kasa", ex_spot_val + ex_cash))
        ex_share = (ex_tot / tot_equity) if tot_equity > 0 else 0.0

        ws4.cell(row=row_curr, column=1, value=ex_name).alignment = align_left
        ws4.cell(row=row_curr, column=1).font = font_bold

        c2 = ws4.cell(row=row_curr, column=2, value=ex_spot_cost)
        c2.number_format = "$#,##0.00"
        c2.alignment = align_right

        c3 = ws4.cell(row=row_curr, column=3, value=ex_spot_val)
        c3.number_format = "$#,##0.00"
        c3.alignment = align_right

        c4 = ws4.cell(row=row_curr, column=4, value=ex_cash)
        c4.number_format = "$#,##0.00"
        c4.alignment = align_right

        c5 = ws4.cell(row=row_curr, column=5, value=ex_tot)
        c5.number_format = "$#,##0.00"
        c5.font = font_bold
        c5.alignment = align_right

        c6 = ws4.cell(row=row_curr, column=6, value=ex_share)
        c6.number_format = "0.0%"
        c6.alignment = align_right

        for col_idx in range(1, 7):
            cell = ws4.cell(row=row_curr, column=col_idx)
            cell.border = border_cell
            if col_idx not in [1, 5]:
                cell.font = font_regular
            if row_curr % 2 == 1:
                cell.fill = fill_zebra
        ws4.row_dimensions[row_curr].height = 20
        row_curr += 1

    # Total Row for Sheet 4
    tot_row4 = row_curr
    ws4.cell(row=tot_row4, column=1, value="GENEL TOPLAM").font = font_bold
    ws4.cell(row=tot_row4, column=1).alignment = align_left
    ws4.cell(row=tot_row4, column=2, value=f"=SUM(B4:B{tot_row4-1})").number_format = "$#,##0.00"
    ws4.cell(row=tot_row4, column=2).font = font_bold
    ws4.cell(row=tot_row4, column=3, value=f"=SUM(C4:C{tot_row4-1})").number_format = "$#,##0.00"
    ws4.cell(row=tot_row4, column=3).font = font_bold
    ws4.cell(row=tot_row4, column=4, value=f"=SUM(D4:D{tot_row4-1})").number_format = "$#,##0.00"
    ws4.cell(row=tot_row4, column=4).font = font_bold
    ws4.cell(row=tot_row4, column=5, value=f"=SUM(E4:E{tot_row4-1})").number_format = "$#,##0.00"
    ws4.cell(row=tot_row4, column=5).font = font_bold
    ws4.cell(row=tot_row4, column=6, value="100.0%").font = font_bold
    ws4.cell(row=tot_row4, column=6).alignment = align_right

    for col_idx in range(1, 7):
        cell = ws4.cell(row=tot_row4, column=col_idx)
        cell.border = border_total
        cell.fill = fill_total
    ws4.row_dimensions[tot_row4].height = 24

    # Auto-fit column widths
    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row in [1, 2]:
                    continue
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()



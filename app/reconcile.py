"""
CoinTakip — Borsa Mutabakatı (FAZ F3)
=====================================================================

NE YAPAR, NE YAPMAZ
-------------------
Bu modül borsadan indirilen dışa aktarım dosyalarını okur ve kullanıcının
defteriyle KARŞILAŞTIRIR. **Deftere hiçbir şey yazmaz.**

Bu bilinçli bir mimari karardır. "Borsa geçmişini içe aktar, defteri yeniden
kur" yaklaşımı denendiğinde şu sorunlarla boğuşmak gerekiyordu: mükerrer
tespiti, kısmi dolumlar, BNB ile ödenen komisyonlar, dönüşümler, toz bakiyeler.
Bir hata maliyet tabanını bozar — yani sistemin tek gerçek değerini.

Bunun yerine: **maliyet tabanı kullanıcının elle girdiği hâliyle kalır, borsa
onu asla ezmez.** Sistem sadece "borsa ne diyor, defterin ne diyor, fark nerede"
sorusunu cevaplar. Ne aktarılacağına kullanıcı karar verir.

NEDEN API DEĞİL DOSYA
---------------------
Borsaların API'si geçmişi vermiyor: Binance ~2 yıl, MEXC 1 ay. Kullanıcının
web arayüzünden indirdiği dosyalar 2023-02'ye kadar iniyor — yani API'nin
ULAŞAMADIĞI derinlik. Ayrıca dosya okumak API anahtarı gerektirmediği için
anahtar güvenliği riski hiç doğmuyor.

DÜRÜSTLÜK KURALI
----------------
Rapor, "uyuşmuyor" ile "dosya o kadar geriye gitmiyor"u BİRBİRİNDEN AYIRMAK
ZORUNDADIR. Kapsam penceresinin dışında kalan bir pozisyonu "fark var" diye
göstermek, kullanıcıyı olmayan bir hatayı kovalamaya iter. Her borsa için
kapsam aralığı hesaplanır ve rapora yazılır.
"""

import csv
import glob
import hashlib
import os
import re
from datetime import datetime

from log_config import get_logger

logger = get_logger("reconcile")

# Dolara ~1:1 sabitli kotasyon varlıkları. Kullanıcının 430 Binance
# dolumunun tamamı USDT veya BUSD; yani USD dönüşümü sorunu bu veri setinde
# doğmuyor. Yine de liste açık tutuldu.
STABLE_QUOTES = {"USDT", "BUSD", "USDC", "FDUSD", "TUSD", "DAI"}

# Borsalar aynı varlığı farklı adlandırabiliyor. MEXC, Tether Gold'u
# `GOLD(XAUT)` diye yazıyor; defterde `XAUT` olarak duruyor. Takma ad tablosu
# olmadan mutabakat bunu "iki ayrı coin" sanır ve sahte fark üretir.
ASSET_ALIASES = {
    "GOLD(XAUT)": "XAUT",
    "XAUT0": "XAUT",
}

# Miktar karşılaştırmasında görece tolerans. Borsa ondalık yuvarlaması ve
# kullanıcının elle girdiği yuvarlanmış miktarlar birebir tutmaz.
QTY_TOLERANCE_PCT = 0.5
QTY_TOLERANCE_ABS = 1e-8


# ---------------------------------------------------------------------
# FAZ F5b — Binance hesap defteri (Transaction History) işlem sınıfları
# ---------------------------------------------------------------------
"""
Alım-satım dosyası hesabın TAMAMI DEĞİLDİR. Coin hesaba yalnızca satın alarak
girmez: airdrop, Launchpool, Convert, toz bakiye eritme, cüzdanlar arası taşıma
kanallarının hiçbiri `Spot Trade History` dosyasında görünmez. Bunlar okunmadığı
sürece yeniden kurulan bakiye gerçeği anlatmaz — ve daha kötüsü, anlattığını
sanır.

`Binance-Transaction-History-*.csv` hesabın tam defteridir; her bakiye hareketi
bir satırdır. Aşağıdaki tablolar hangi `Operation` değerinin ne anlama geldiğini
söyler. Tanınmayan bir işlem SESSİZCE ATLANMAZ, uyarı olarak rapora çıkar:
sessiz atlama tam da bu fazın düzelttiği hatanın kaynağıydı.
"""

# Bedelsiz girişler. Maliyetleri sıfırdır ve bu BİLİNEN bir sıfırdır —
# "maliyeti bilinmiyor" ile karıştırılmamalı, ikisi farklı şeydir.
TH_REWARD_OPS = {
    "earn - airdrop distribution",
    "launchpool airdrop - user claim distribution",
    "launchpool airdrop - system distribution",
    "airdrop assets", "distribution", "crypto box", "cash voucher",
    "campaign related reward", "mission reward distribution",
    "commission rebate", "referral kickback", "referral commission",
    "staking rewards", "eth 2.0 staking rewards", "bnb vault rewards",
    "simple earn flexible interest", "simple earn locked rewards",
    "savings interest", "savings distribution", "liquid swap rewards",
}

# İki (veya daha fazla) bacaklı dönüşümler: aynı zaman damgasında bir taraf
# eksi, bir taraf artı. Bacaklar eşleştirilerek maliyet çıkarılır.
TH_CONVERT_OPS = {
    "binance convert", "small assets exchange bnb", "bnb convert",
    "large otc trading", "stablecoins auto-conversion", "asset recovery",
    "token swap - redenomination/rebranding",
}

# Bakiyeyi değiştiren ama alım/satım olmayan taşımalar. Maliyet taşımazlar.
TH_TRANSFER_OPS = {
    "thirdparty wallet transfer", "inter-wallet transfer",
    "transfer between spot and um futures", "transfer between spot and cm futures",
    "transfer between spot and funding", "transfer between main and funding account",
    "transfer between main account and margin account",
    "launchpool subscription/redemption",
    "simple earn flexible subscription", "simple earn flexible redemption",
    "simple earn locked subscription", "simple earn locked redemption",
    "staking purchase", "staking redemption", "eth 2.0 staking",
    "isolated margin loan", "isolated margin repayment",
}

# Spot alım-satımının ayakları. `Spot Trade History` dosyası varsa bunlar
# ATLANIR — ikisini birden almak her işlemi iki kez saydırır. Dosya yoksa
# buradan üretilirler; o zaman tek kaynak budur.
TH_TRADE_OPS = {
    "transaction buy", "transaction sold", "transaction spend",
    "transaction revenue", "transaction related", "buy", "sell",
}
TH_FEE_OP = "transaction fee"

# Ayrı `Deposit-History` / `Withdraw-History` dosyaları varsa bunlar atlanır.
TH_DEPOSIT_OPS = {"deposit", "fiat deposit"}
TH_WITHDRAW_OPS = {"withdraw", "fiat withdraw"}

# Vadeli/margin hesabındaki hareketler spot bakiyeyi ilgilendirmez. Hesap
# sütunu bunu zaten ayırıyor; yalnızca 'Spot' satırları okunur.
TH_SPOT_ACCOUNT = "spot"


def export_root():
    """Dışa aktarım klasörü — proje kökündeki `borsa_exports/`."""
    from data_manager import BASE_DIR
    return os.path.join(BASE_DIR, "borsa_exports")


def normalize_asset(name):
    t = str(name or "").upper().strip()
    return ASSET_ALIASES.get(t, t)


_MIKTAR_RE = re.compile(r"^\s*(-?[\d.,]+)\s*([A-Z0-9()._-]*)\s*$", re.IGNORECASE)


def parse_amount_with_asset(text):
    """
    Binance/MEXC dışa aktarımları miktarı ve varlığı BİTİŞİK yazar:
    `"0.00089BTC"` → `(0.00089, "BTC")`, `"0USDT"` → `(0.0, "USDT")`.
    Varlık yoksa (düz sayı) ikinci değer boş döner.
    """
    if text is None:
        return 0.0, ""
    m = _MIKTAR_RE.match(str(text))
    if not m:
        return 0.0, ""
    ham = m.group(1).replace(",", "")
    try:
        deger = float(ham)
    except ValueError:
        return 0.0, ""
    return deger, normalize_asset(m.group(2))


def split_pair(pair, quote_hint=""):
    """`BTCUSDT` / `GOLD(XAUT)_USDT` → ("BTC","USDT") / ("XAUT","USDT")."""
    p = str(pair or "").upper().strip()
    if "_" in p:
        taban, _, kot = p.partition("_")
        return normalize_asset(taban), normalize_asset(kot)
    if quote_hint and p.endswith(quote_hint) and len(p) > len(quote_hint):
        return normalize_asset(p[:-len(quote_hint)]), normalize_asset(quote_hint)
    for q in sorted(STABLE_QUOTES | {"BTC", "ETH", "BNB", "TRY"}, key=len, reverse=True):
        if p.endswith(q) and len(p) > len(q):
            return normalize_asset(p[:-len(q)]), normalize_asset(q)
    return normalize_asset(p), ""


def _olay(exchange, zaman, kind, asset, qty, **ek):
    olay = {
        "exchange": exchange,
        "time": str(zaman or "")[:19],
        # TRADE | DEPOSIT | WITHDRAW | REWARD | FEE
        "kind": kind,
        "asset": normalize_asset(asset),
        "qty": float(qty or 0.0),        # işaretli: + giriş, − çıkış
        "quote_asset": "",
        "quote_qty": 0.0,
        "price": 0.0,
        "fee_asset": "",
        "fee_qty": 0.0,
        "usd_value": 0.0,
        "usd_known": False,
        # Bedelsiz giriş: maliyeti sıfır ve bu BİLİNEN bir sıfır.
        "zero_cost": False,
        "operation": "",
        "source": "",
    }
    olay.update(ek)
    return olay


# ---------------------------------------------------------------------
# Binance okuyucuları
# ---------------------------------------------------------------------
def load_binance_trades(path):
    olaylar, uyarilar = [], []
    with open(path, encoding="utf-8-sig", newline="") as f:
        for satir in csv.DictReader(f):
            try:
                miktar, taban = parse_amount_with_asset(satir.get("Executed"))
                tutar, kot = parse_amount_with_asset(satir.get("Amount"))
                komisyon, kom_varlik = parse_amount_with_asset(satir.get("Fee"))
                if not taban:
                    taban, kot2 = split_pair(satir.get("Pair"), kot)
                    kot = kot or kot2
                yon = str(satir.get("Side", "")).upper()
                isaret = 1.0 if yon.startswith("B") else -1.0

                # Komisyon coin'in KENDİSİNDEN alındıysa eline geçen miktar
                # o kadar azalır. Binance'in "Executed" sütunu komisyon
                # düşülmeden önceki dolum miktarıdır.
                net = miktar
                kom_varlik = normalize_asset(kom_varlik)
                if isaret > 0 and kom_varlik and kom_varlik == normalize_asset(taban):
                    net = miktar - komisyon

                olaylar.append(_olay(
                    "BINANCE", satir.get("Time"), "TRADE", taban, isaret * net,
                    quote_asset=kot, quote_qty=-isaret * tutar,
                    price=float(satir.get("Price") or 0.0),
                    fee_asset=kom_varlik, fee_qty=komisyon,
                    usd_value=tutar if kot in STABLE_QUOTES else 0.0,
                    usd_known=kot in STABLE_QUOTES,
                    source=os.path.basename(path),
                ))

                # Komisyon BAŞKA bir coinden ödendiyse (Binance'te çoğunlukla
                # BNB) o coinin bakiyesi azalır. Yukarıdaki satır yalnızca işlem
                # yapılan varlığı anlatır; komisyon coini ayrı bir çıkış olarak
                # yazılmazsa BNB bakiyesi olduğundan yüksek yeniden kurulur.
                # Komisyon işlem varlığının kendisindense zaten `net`ten düşüldü.
                if komisyon > 0 and kom_varlik and kom_varlik != normalize_asset(taban):
                    olaylar.append(_olay(
                        "BINANCE", satir.get("Time"), "FEE", kom_varlik, -komisyon,
                        operation="Trade Fee", source=os.path.basename(path),
                    ))
            except Exception as e:
                uyarilar.append(f"Binance işlem satırı okunamadı: {e}")
    return olaylar, uyarilar


def _binance_transfer(path, kind, isaret):
    olaylar, uyarilar = [], []
    with open(path, encoding="utf-8-sig", newline="") as f:
        for satir in csv.DictReader(f):
            try:
                if str(satir.get("Status", "")).lower() not in ("completed", "success", ""):
                    continue
                miktar = float(str(satir.get("Amount") or 0).replace(",", ""))
                olaylar.append(_olay(
                    "BINANCE", satir.get("Time"), kind, satir.get("Coin"),
                    isaret * miktar,
                    fee_qty=float(str(satir.get("Fee") or 0).replace(",", "") or 0),
                    fee_asset=normalize_asset(satir.get("Coin")),
                    source=os.path.basename(path),
                ))
            except Exception as e:
                uyarilar.append(f"Binance {kind} satırı okunamadı: {e}")
    return olaylar, uyarilar


def load_binance_deposits(path):
    return _binance_transfer(path, "DEPOSIT", 1.0)


def load_binance_withdrawals(path):
    return _binance_transfer(path, "WITHDRAW", -1.0)


def _cok_bacakli_hareket(zaman, kalemler, kaynak, operasyon):
    """
    Aynı zaman damgasını paylaşan çok bacaklı bir hareketi olaylara çevirir.

    Convert, toz eritme ve (alım-satım dosyası yoksa) spot dolumları hep aynı
    şekli taşır: bir tarafta çıkan varlık, diğer tarafta giren varlık, aynı
    saniyede. Dolar karşılığı ancak bacaklardan biri dolara sabitliyse
    BİLİNİR; değilse uydurmak yerine "bilinmiyor" denir.

    Aynı damgada birden fazla dolum olabilir (kısmi dolumlar tek emirdendir).
    Giren varlık tekse toplam stabil çıkış o dolumlara miktarları oranında
    dağıtılır — ağırlıklı ortalama maliyet doğru çıkar.
    """
    artilar = [(v, d) for v, d in kalemler if d > 0]
    eksiler = [(v, d) for v, d in kalemler if d < 0]
    stabil_giren = sum(d for v, d in artilar if v in STABLE_QUOTES)
    stabil_cikan = sum(-d for v, d in eksiler if v in STABLE_QUOTES)
    arti_varliklar = {v for v, _ in artilar if v not in STABLE_QUOTES}
    eksi_varliklar = {v for v, _ in eksiler if v not in STABLE_QUOTES}
    alinan_toplam = sum(d for v, d in artilar if v not in STABLE_QUOTES)
    verilen_toplam = sum(-d for v, d in eksiler if v not in STABLE_QUOTES)

    olaylar = []
    for varlik, degisim in artilar:
        if varlik in STABLE_QUOTES:
            tutar, bilinir = abs(degisim), True
        elif len(arti_varliklar) == 1 and stabil_cikan > 0 and alinan_toplam > 0:
            tutar, bilinir = stabil_cikan * (degisim / alinan_toplam), True
        else:
            tutar, bilinir = 0.0, False
        olaylar.append(_olay("BINANCE", zaman, "TRADE", varlik, degisim,
                             usd_value=tutar, usd_known=bilinir,
                             operation=operasyon, source=kaynak))
    for varlik, degisim in eksiler:
        if varlik in STABLE_QUOTES:
            tutar, bilinir = abs(degisim), True
        elif len(eksi_varliklar) == 1 and stabil_giren > 0 and verilen_toplam > 0:
            tutar, bilinir = stabil_giren * (-degisim / verilen_toplam), True
        else:
            tutar, bilinir = 0.0, False
        olaylar.append(_olay("BINANCE", zaman, "TRADE", varlik, degisim,
                             usd_value=tutar, usd_known=bilinir,
                             operation=operasyon, source=kaynak))
    return olaylar


def load_binance_transaction_history(path, skip_ops=frozenset()):
    """
    Binance hesap defteri — her bakiye hareketinin tek satır olduğu tam kayıt.

    NEDEN GEREKLİ
    -------------
    Alım-satım dosyası yalnızca spot emirleri anlatır. Airdrop, Launchpool,
    Convert, toz bakiye eritme ve cüzdanlar arası taşımalar orada YOKTUR; bu
    kanallardan gelen coinler görülmediği için yeniden kurulan bakiye eksik,
    bu kanallardan çıkanlar görülmediği için fazla çıkar. Gerçek veride
    ölçüldü: 21 önerinin 10'u bu yüzden yanlıştı.

    MÜKERRER TEHLİKESİ
    ------------------
    Bu dosya spot alım-satımları da içerir (`Transaction Buy/Sold/Spend/
    Revenue/Fee`) ve para yatırma/çekmeyi de. Kendi dosyaları varken bunları
    da almak her işlemi İKİ KEZ saydırır. `skip_ops` hangi işlem türünün
    başka bir dosyadan geldiğini söyler; kalanlar buradan üretilir. Böylece
    kullanıcı yalnızca bu dosyayı indirmişse de mutabakat çalışır.
    """
    olaylar, uyarilar = [], []
    gruplar = {}
    bilinmeyen = {}
    kaynak = os.path.basename(path)
    with open(path, encoding="utf-8-sig", newline="") as f:
        for satir in csv.DictReader(f):
            if str(satir.get("Account") or "").strip().lower() != TH_SPOT_ACCOUNT:
                continue
            ham_op = str(satir.get("Operation") or "").strip()
            op = ham_op.lower()
            if op in skip_ops:
                continue
            try:
                degisim = float(str(satir.get("Change") or 0).replace(",", ""))
            except (TypeError, ValueError):
                uyarilar.append(f"Binance hesap defteri satırı okunamadı "
                                f"({ham_op}): {satir.get('Change')!r}")
                continue
            if abs(degisim) < 1e-15:
                continue
            varlik = normalize_asset(satir.get("Coin"))
            zaman = str(satir.get("Time") or "")[:19]

            if op in TH_REWARD_OPS:
                # Bedelsiz geldi: maliyet sıfır ve bu bilinen bir sıfır.
                olaylar.append(_olay("BINANCE", zaman, "REWARD", varlik, degisim,
                                     usd_value=0.0, usd_known=True, zero_cost=True,
                                     operation=ham_op, source=kaynak))
            elif op == TH_FEE_OP:
                olaylar.append(_olay("BINANCE", zaman, "FEE", varlik, -abs(degisim),
                                     operation=ham_op, source=kaynak))
            elif op in TH_TRANSFER_OPS or op in TH_DEPOSIT_OPS or op in TH_WITHDRAW_OPS:
                kind = "DEPOSIT" if degisim > 0 else "WITHDRAW"
                olaylar.append(_olay("BINANCE", zaman, kind, varlik, degisim,
                                     operation=ham_op, source=kaynak))
            elif op in TH_CONVERT_OPS:
                gruplar.setdefault(("convert", zaman), []).append((varlik, degisim))
            elif op in TH_TRADE_OPS:
                gruplar.setdefault(("trade", zaman), []).append((varlik, degisim))
            else:
                # Tanınmayan işlem sessizce atlanmaz. Sessiz atlama F5b'de
                # düzeltilen hatanın ta kendisiydi.
                bilinmeyen[ham_op] = bilinmeyen.get(ham_op, 0) + 1

    for (tur, zaman) in sorted(gruplar):
        olaylar.extend(_cok_bacakli_hareket(
            zaman, gruplar[(tur, zaman)], kaynak,
            "Binance Convert" if tur == "convert" else "Spot Trade"))

    for ham_op, adet in sorted(bilinmeyen.items(), key=lambda kv: -kv[1]):
        uyarilar.append(
            f"Binance hesap defterinde tanınmayan işlem türü: '{ham_op}' "
            f"({adet} satır). Bu satırlar hesaba KATILMADI; bakiye eksik veya "
            "fazla çıkabilir."
        )
    return olaylar, uyarilar


# ---------------------------------------------------------------------
# MEXC okuyucuları
# ---------------------------------------------------------------------
def _xlsx_rows(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        satirlar = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not satirlar:
        return []
    basliklar = [str(c) if c is not None else "" for c in satirlar[0]]
    return [dict(zip(basliklar, r)) for r in satirlar[1:]
            if r and any(c is not None for c in r)]


def load_mexc_trades(path):
    olaylar, uyarilar = [], []
    for satir in _xlsx_rows(path):
        try:
            taban, kot = split_pair(satir.get("Pairs"))
            miktar = float(str(satir.get("Executed Amount") or 0).replace(",", ""))
            tutar = float(str(satir.get("Total") or 0).replace(",", ""))
            komisyon, kom_varlik = parse_amount_with_asset(satir.get("Fee"))
            yon = str(satir.get("Side", "")).upper()
            isaret = 1.0 if yon.startswith("B") else -1.0

            net = miktar
            if isaret > 0 and kom_varlik and normalize_asset(kom_varlik) == taban:
                net = miktar - komisyon

            olaylar.append(_olay(
                "MEXC", satir.get("Time"), "TRADE", taban, isaret * net,
                quote_asset=kot, quote_qty=-isaret * tutar,
                price=float(str(satir.get("Filled Price") or 0).replace(",", "")),
                fee_asset=kom_varlik, fee_qty=komisyon,
                usd_value=tutar if kot in STABLE_QUOTES else 0.0,
                usd_known=kot in STABLE_QUOTES,
                source=os.path.basename(path),
            ))
        except Exception as e:
            uyarilar.append(f"MEXC işlem satırı okunamadı: {e}")
    return olaylar, uyarilar


def load_mexc_statement(path):
    """
    MEXC ekstresinden YALNIZCA para yatırma/çekme alınır.

    'Spot Trading' satırları işlem dosyasında zaten var; ikisini birden almak
    her alımı iki kez saydırırdı. Vadeli hesap transferleri de spot bakiyeyi
    ilgilendirmediği için atlanır.
    """
    olaylar, uyarilar = [], []
    for satir in _xlsx_rows(path):
        try:
            tip = str(satir.get("Transaction Type") or "").strip().lower()
            if tip not in ("deposit", "withdraw"):
                continue
            miktar = float(str(satir.get("Quantity") or 0).replace(",", ""))
            kind = "DEPOSIT" if tip == "deposit" else "WITHDRAW"
            # Ekstre çıkışları zaten negatif yazıyor; işareti tipe göre sabitliyoruz.
            miktar = abs(miktar) * (1.0 if kind == "DEPOSIT" else -1.0)
            olaylar.append(_olay(
                "MEXC", satir.get("Creation Time(UTC+03:00)") or satir.get("Creation Time"),
                kind, satir.get("Crypto"), miktar,
                source=os.path.basename(path),
            ))
        except Exception as e:
            uyarilar.append(f"MEXC ekstre satırı okunamadı: {e}")
    return olaylar, uyarilar


# ---------------------------------------------------------------------
# Dosya keşfi ve yükleme
# ---------------------------------------------------------------------
# (dosya adı kalıbı, borsa, tür, okuyucu)
FILE_PATTERNS = [
    ("*Spot-Trade-History*.csv", "BINANCE", "trades", load_binance_trades),
    ("*Deposit-History*.csv", "BINANCE", "deposits", load_binance_deposits),
    ("*Withdraw-History*.csv", "BINANCE", "withdrawals", load_binance_withdrawals),
    ("*Transaction-History*.csv", "BINANCE", "transactions",
     load_binance_transaction_history),
    ("*Trade History*.xlsx", "MEXC", "trades", load_mexc_trades),
    ("*Statement*.xlsx", "MEXC", "statement", load_mexc_statement),
]


def discover_export_files(root=None):
    """Klasör yapısından bağımsız olarak tanınan dosyaları bulur."""
    root = root or export_root()
    bulunan = []
    if not os.path.isdir(root):
        return bulunan
    for kalip, borsa, tur, okuyucu in FILE_PATTERNS:
        for yol in glob.glob(os.path.join(root, "**", kalip), recursive=True):
            # 'Order History' dolum değil emirdir; 'Trade History' ile karışmasın.
            if "order history" in os.path.basename(yol).lower():
                continue
            bulunan.append({"path": yol, "name": os.path.basename(yol),
                            "exchange": borsa, "kind": tur, "_loader": okuyucu})
    bulunan.sort(key=lambda d: (d["exchange"], d["kind"], d["name"]))
    return bulunan


def _hesap_defteri_atlanacaklar(dosyalar):
    """
    Binance hesap defterinin hangi işlem türlerini ATLAMASI gerektiği.

    Bir hareketin tek bir kaynağı olmalı. Alım-satım, para yatırma ve çekme
    için ayrı dosyalar varsa onlar esastır (fiyat, pariteler ve durum bilgisi
    orada); hesap defteri o türleri atlar. Dosyalar yoksa atlama olmaz ve
    hesap defteri tek kaynak olarak devreye girer.
    """
    mevcut = {(d["exchange"], d["kind"]) for d in dosyalar}
    atla = set()
    if ("BINANCE", "trades") in mevcut:
        atla |= TH_TRADE_OPS | {TH_FEE_OP}
    if ("BINANCE", "deposits") in mevcut:
        atla |= TH_DEPOSIT_OPS
    if ("BINANCE", "withdrawals") in mevcut:
        atla |= TH_WITHDRAW_OPS
    return atla


def load_all_events(root=None):
    """(olaylar, kaynak_bilgisi, uyarilar)"""
    olaylar, kaynaklar, uyarilar = [], [], []
    dosyalar = discover_export_files(root)
    atlanacak = _hesap_defteri_atlanacaklar(dosyalar)
    for dosya in dosyalar:
        try:
            if dosya["kind"] == "transactions":
                yeni, uy = dosya["_loader"](dosya["path"], atlanacak)
            else:
                yeni, uy = dosya["_loader"](dosya["path"])
        except Exception as e:
            uyarilar.append(f"{dosya['name']} okunamadı: {e}")
            logger.warning("Mutabakat dosyası okunamadı (%s): %s", dosya["name"], e)
            continue
        olaylar.extend(yeni)
        uyarilar.extend(uy)
        tarihler = sorted(o["time"] for o in yeni if o["time"])
        kaynaklar.append({
            "name": dosya["name"], "exchange": dosya["exchange"],
            "kind": dosya["kind"], "rows": len(yeni),
            "first": tarihler[0][:10] if tarihler else None,
            "last": tarihler[-1][:10] if tarihler else None,
        })
    return olaylar, kaynaklar, uyarilar


def coverage_windows(olaylar):
    """
    Borsa başına kapsam aralığı.

    Rapor "uyuşmuyor" ile "dosya o kadar geriye gitmiyor"u ayırt edebilsin
    diye şart. MEXC dışa aktarımı 2024-10'da başlıyor; ondan önce alınmış bir
    coin dosyada YOK, ama bu bir tutarsızlık değil.
    """
    pencere = {}
    for o in olaylar:
        if not o["time"]:
            continue
        p = pencere.setdefault(o["exchange"], {"first": o["time"][:10], "last": o["time"][:10]})
        p["first"] = min(p["first"], o["time"][:10])
        p["last"] = max(p["last"], o["time"][:10])
    return pencere


# ---------------------------------------------------------------------
# Özetleme ve mutabakat
# ---------------------------------------------------------------------
def build_asset_summary(olaylar):
    """Varlık bazında borsa tarafının özeti."""
    ozet = {}
    for o in olaylar:
        a = o["asset"]
        if not a:
            continue
        s = ozet.setdefault(a, {
            "asset": a, "bought_qty": 0.0, "sold_qty": 0.0,
            "deposited_qty": 0.0, "withdrawn_qty": 0.0,
            "reward_qty": 0.0, "fee_qty": 0.0,
            "buy_cost_usd": 0.0, "buy_qty_usd_known": 0.0,
            "sell_proceeds_usd": 0.0, "trade_count": 0,
            "exchanges": set(), "first": None, "last": None,
            "usd_unknown_trades": 0,
        })
        s["exchanges"].add(o["exchange"])
        t = o["time"][:10] if o["time"] else None
        if t:
            s["first"] = t if s["first"] is None else min(s["first"], t)
            s["last"] = t if s["last"] is None else max(s["last"], t)

        if o["kind"] == "TRADE":
            s["trade_count"] += 1
            if o["qty"] >= 0:
                s["bought_qty"] += o["qty"]
                if o["usd_known"]:
                    s["buy_cost_usd"] += o["usd_value"]
                    s["buy_qty_usd_known"] += o["qty"]
                else:
                    s["usd_unknown_trades"] += 1
            else:
                s["sold_qty"] += -o["qty"]
                if o["usd_known"]:
                    s["sell_proceeds_usd"] += o["usd_value"]
                else:
                    s["usd_unknown_trades"] += 1
        elif o["kind"] == "DEPOSIT":
            s["deposited_qty"] += o["qty"]
        elif o["kind"] == "WITHDRAW":
            s["withdrawn_qty"] += -o["qty"]
        elif o["kind"] == "REWARD":
            # Bedelsiz giriş. Alım değildir — ortalama maliyeti bozmasın diye
            # `buy_cost_usd`'ye girmez, ama bakiyeyi artırır.
            s["reward_qty"] += o["qty"]
        elif o["kind"] == "FEE":
            s["fee_qty"] += -o["qty"]

    for s in ozet.values():
        s["exchanges"] = sorted(s["exchanges"])
        s["on_exchange_qty"] = (s["bought_qty"] - s["sold_qty"]
                                + s["deposited_qty"] - s["withdrawn_qty"]
                                + s["reward_qty"] - s["fee_qty"])
        s["acquired_qty"] = s["bought_qty"] + s["deposited_qty"] + s["reward_qty"]
        s["exchange_avg_cost"] = (s["buy_cost_usd"] / s["buy_qty_usd_known"]
                                  if s["buy_qty_usd_known"] > 0 else 0.0)
    return ozet


def ledger_summary(data):
    """Defterdeki aktif pozisyonların varlık bazında özeti."""
    from data_manager import ACTIVE_STATUS
    ozet = {}
    for tx in data.get("transactions", []):
        if tx.get("status") != ACTIVE_STATUS:
            continue
        ham = str(tx.get("coin") or "").upper().strip()
        varlik = normalize_asset(ham[:-4] if ham.endswith("USDT") and len(ham) > 4 else ham)
        if not varlik:
            continue
        qty = float(tx.get("qty") or 0.0)
        cost = float(tx.get("cost") or 0.0)
        s = ozet.setdefault(varlik, {"asset": varlik, "qty": 0.0, "invested": 0.0,
                                     "locations": set(), "lots": 0, "first_date": None})
        s["qty"] += qty
        s["invested"] += qty * cost
        s["locations"].add(str(tx.get("exchange") or "BINANCE").upper())
        s["lots"] += 1
        tarih = str(tx.get("date") or "")[:10]
        if tarih:
            s["first_date"] = tarih if s["first_date"] is None else min(s["first_date"], tarih)
    for s in ozet.values():
        s["locations"] = sorted(s["locations"])
        s["avg_cost"] = (s["invested"] / s["qty"]) if s["qty"] > 0 else 0.0
    return ozet


def _yakin(a, b):
    fark = abs(a - b)
    if fark <= QTY_TOLERANCE_ABS:
        return True
    olcek = max(abs(a), abs(b))
    return olcek > 0 and (fark / olcek * 100.0) <= QTY_TOLERANCE_PCT


def reconcile(data, root=None):
    """
    Borsa dosyalarıyla defteri karşılaştırır. **Deftere hiçbir şey yazmaz.**
    """
    olaylar, kaynaklar, uyarilar = load_all_events(root)
    pencere = coverage_windows(olaylar)
    borsa = build_asset_summary(olaylar)
    defter = ledger_summary(data)

    satirlar = []
    for varlik in sorted(set(borsa) | set(defter)):
        b = borsa.get(varlik)
        d = defter.get(varlik)
        d_qty = d["qty"] if d else 0.0
        b_kalan = b["on_exchange_qty"] if b else 0.0

        # --- Kapsam kanıtı ---
        # Dosyaların geçmişi kapsamadığını İSPATLAYAN iki sinyal var; ikisi de
        # varken "fark var" demek kullanıcıyı olmayan bir hatayı kovalamaya iter.
        konumlar = set(d["locations"]) if d else set()
        kapsanan_konum = bool(konumlar & set(pencere.keys()))

        # 1) Negatif borsa bakiyesi FİZİKSEL OLARAK İMKÂNSIZDIR. Çıkan miktar
        #    girenden fazlaysa, alım penceresinden önce olmuş demektir.
        negatif_kanit = b is not None and b_kalan < -QTY_TOLERANCE_ABS

        # 2) Defterdeki en eski lot, dosyaların başlangıcından öncesindeyse
        #    o alım dosyada zaten olamaz.
        pencere_basi = None
        if b and b["exchanges"]:
            adaylar = [pencere[x]["first"] for x in b["exchanges"] if x in pencere]
            pencere_basi = min(adaylar) if adaylar else None
        elif pencere:
            pencere_basi = min(p["first"] for p in pencere.values())
        tarih_kaniti = bool(d and d.get("first_date") and pencere_basi
                            and d["first_date"] < pencere_basi)

        # Stabilcoinler pozisyon değil NAKİTtir; cüzdan bakiyesinde takip
        # ediliyorlar. Pozisyon mutabakatında "USDT -3.789 eksik" demek gürültüdür.
        if varlik in STABLE_QUOTES:
            durum = "stablecoin"
            not_metni = ("Nakit birimi — pozisyon olarak takip edilmiyor. "
                         "Serbest nakit bakiyesi Cüzdan ekranından yönetilir.")
        elif b is None:
            if d and konumlar and not kapsanan_konum:
                durum = "uncovered_location"
                not_metni = (f"Bu varlık {', '.join(sorted(konumlar))} üzerinde tutuluyor "
                             "ve o konum için dışa aktarım dosyası yok. "
                             "Mutabakat yapılamaz — eksiklik değil, kapsam dışı.")
            else:
                durum = "only_ledger"
                not_metni = ("Bu varlık dışa aktarım dosyalarında hiç yok. "
                             "Dosyaların kapsamadığı bir tarihte alınmış veya "
                             "başka bir borsadan gelmiş olabilir.")
        elif negatif_kanit:
            durum = "coverage_gap"
            not_metni = (f"Borsa kayıtlarına göre bakiye eksiye düşüyor "
                         f"({b_kalan:,.4f}) — bu imkânsızdır. Alım işlemleri dosyanın "
                         f"başlangıcından ({pencere_basi}) önce yapılmış. "
                         "Karşılaştırma anlamsız, fark değil.")
        elif d is None or d_qty <= QTY_TOLERANCE_ABS:
            if b_kalan <= QTY_TOLERANCE_ABS:
                durum = "closed"
                not_metni = "Borsada girilip çıkılmış, defterde açık pozisyon yok."
            else:
                durum = "only_exchange"
                not_metni = "Borsa kayıtlarına göre bakiye var ama defterde yok."
        elif _yakin(b_kalan, d_qty):
            durum = "match"
            not_metni = ""
        elif tarih_kaniti:
            durum = "coverage_gap"
            not_metni = (f"Defterdeki en eski alım {d['first_date']}, dosya ise "
                         f"{pencere_basi} tarihinde başlıyor. Önceki işlemler dosyada "
                         "yok; fark bundan kaynaklanıyor olabilir.")
        elif b["withdrawn_qty"] > QTY_TOLERANCE_ABS:
            durum = "off_exchange"
            not_metni = (f"{b['withdrawn_qty']:,.8f} adet borsadan çekilmiş. "
                         "Defterdeki miktar cüzdan/başka borsa bakiyesini "
                         "yansıtıyor olabilir — bu bir tutarsızlık olmayabilir.")
        else:
            durum = "mismatch"
            not_metni = "Borsa kayıtları ile defter miktarı uyuşmuyor."

        satirlar.append({
            "asset": varlik,
            "exchanges": b["exchanges"] if b else [],
            "locations": d["locations"] if d else [],
            "bought_qty": b["bought_qty"] if b else 0.0,
            "sold_qty": b["sold_qty"] if b else 0.0,
            "deposited_qty": b["deposited_qty"] if b else 0.0,
            "withdrawn_qty": b["withdrawn_qty"] if b else 0.0,
            "reward_qty": b["reward_qty"] if b else 0.0,
            "fee_qty": b["fee_qty"] if b else 0.0,
            "on_exchange_qty": b_kalan,
            "acquired_qty": b["acquired_qty"] if b else 0.0,
            "ledger_qty": d_qty,
            "diff_qty": b_kalan - d_qty,
            "exchange_avg_cost": b["exchange_avg_cost"] if b else 0.0,
            "ledger_avg_cost": d["avg_cost"] if d else 0.0,
            "ledger_invested": d["invested"] if d else 0.0,
            "ledger_first_date": d.get("first_date") if d else None,
            "coverage_start": pencere_basi,
            "trade_count": b["trade_count"] if b else 0,
            "first_seen": b["first"] if b else None,
            "last_seen": b["last"] if b else None,
            "usd_unknown_trades": b["usd_unknown_trades"] if b else 0,
            "status": durum,
            "note": not_metni,
        })

    sira = {"mismatch": 0, "only_exchange": 1, "coverage_gap": 2, "off_exchange": 3,
            "only_ledger": 4, "uncovered_location": 5, "match": 6,
            "closed": 7, "stablecoin": 8}
    satirlar.sort(key=lambda r: (sira.get(r["status"], 9), r["asset"]))

    sayim = {}
    for r in satirlar:
        sayim[r["status"]] = sayim.get(r["status"], 0) + 1

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "export_root": root or export_root(),
        "files_found": len(kaynaklar),
        "sources": kaynaklar,
        "coverage": pencere,
        "event_count": len(olaylar),
        "rows": satirlar,
        "status_counts": sayim,
        "warnings": uyarilar[:50],
        # Bu rapor salt okunurdur; arayüz bunu kullanıcıya açıkça söyler.
        "read_only": True,
    }


# =====================================================================
# FAZ F5 — LOT YENİDEN KURULUMU (ÖNERİ ÜRETİMİ)
# =====================================================================
"""
NEDEN BU VAR
------------
F3 "borsa X diyor, defterin Y diyor" der ve orada durur. Ama kullanıcının
defteri bir İŞLEM GÜNLÜĞÜ değildir: sisteme geçtiği gün elinde kalan
pozisyonları elle girmiştir. Yani "hangi işlemi girmeyi unuttum" sorusunun
defter tarafında cevabı yoktur — girilmemiş bir işlem değil, hiç girilmemiş
bir geçmiş söz konusudur.

Cevap borsa dosyasındadır. Dosya her alımı ve satımı tarihiyle taşıdığı için
FIFO yürütülerek "bugün elde ne kalmış olması gerektiği" gerçek tarih ve
gerçek fiyatlarla yeniden kurulabilir. Kullanıcının hatırlamasına gerek yok.

Bu bölüm YALNIZCA ÖNERİ ÜRETİR. Yazma işi `data_manager.apply_rebuild`
tarafındadır ve pozisyon başına açık onay ister. F3'ün salt okunurluğu
bozulmadı: `reconcile()` hâlâ hiçbir şey yazmaz, bu fonksiyonlar da yazmaz.

DÜRÜSTLÜK KURALI (F3'ten devam)
-------------------------------
Kapsamı kanıtlanamayan bir pozisyon için öneri VERİLMEZ. Yanlış bir maliyet
tabanı, eksik bir maliyet tabanından daha zararlıdır: kullanıcı ona güvenir.
Öneriyi engelleyen üç durum var ve üçü de rapora ayrı ayrı yazılır.

F5b DÜZELTMESİ — KANIT TEK YÖNLÜYDÜ
-----------------------------------
İlk sürüm "bakiye hiç eksiye düşmüyorsa dosya geçmişi kapsıyordur" diyordu ve
engeli olmayan satırları yeşil `ready` rozetiyle gösteriyordu. Bu çıkarım
geçersizdi: negatif bakiyenin imkânsızlığı yalnızca *görünmeyen bir satış
olmadığını* söyler, *görünmeyen bir alım olmadığını* söylemez. Pencereden önce
alınıp hiç satılmamış bir bakiye hiçbir iz bırakmaz.

İki düzeltme yapıldı. Birincisi, Binance hesap defteri (`Transaction History`)
okunmaya başlandı: airdrop, Launchpool, Convert ve toz eritme kanalları artık
görünüyor. İkincisi, `ready` durumu kaldırıldı — hiçbir satır kullanıcı
borsadaki gerçek bakiyeyi girmeden uygulanamaz (`evaluate_verified_qty`).
"""

# Yeniden kurulmuş lotların defterdeki maliyet yöntemi etiketi.
REBUILD_COST_METHOD = "Borsa Kaydı (mutabakat)"


def ledger_positions(data):
    """
    Defterdeki aktif pozisyonlar, **(varlık, konum) çifti** bazında.

    `ledger_summary` varlığı borsalar üstünden toplar; düzeltme için bu yetmez.
    Binance'teki TIA ile MEXC'teki TIA ayrı maliyet tabanı taşır ve uygulamanın
    temel vaadi budur — ikisini birlikte ezmek o ayrımı yok ederdi.
    """
    from data_manager import ACTIVE_STATUS, normalize_location
    ozet = {}
    for tx in data.get("transactions", []):
        if tx.get("status") != ACTIVE_STATUS:
            continue
        ham = str(tx.get("coin") or "").upper().strip()
        varlik = normalize_asset(ham[:-4] if ham.endswith("USDT") and len(ham) > 4 else ham)
        if not varlik:
            continue
        konum = normalize_location(tx.get("exchange"))
        s = ozet.setdefault((varlik, konum), {
            "asset": varlik, "location": konum, "coin": ham,
            "qty": 0.0, "invested": 0.0, "lots": 0,
            "first_date": None, "tx_ids": [],
        })
        qty = float(tx.get("qty") or 0.0)
        s["qty"] += qty
        s["invested"] += qty * float(tx.get("cost") or 0.0)
        s["lots"] += 1
        s["tx_ids"].append(int(tx.get("id", 0) or 0))
        tarih = str(tx.get("date") or "")[:10]
        if tarih:
            s["first_date"] = tarih if s["first_date"] is None else min(s["first_date"], tarih)
    for s in ozet.values():
        s["avg_cost"] = (s["invested"] / s["qty"]) if s["qty"] > 0 else 0.0
    return ozet


def _birim_maliyet(olay):
    """
    Bir alım olayının birim maliyeti.

    `usd_value / qty` tercih edilir çünkü `qty` komisyon düşülmüş NET miktardır:
    100 token alıp 0.1'ini komisyona verdiyseniz 99.9 tokene o paranın tamamını
    ödemişsinizdir. Ham `price` sütunu bunu kaçırır. Kotasyon dolara sabitli
    değilse maliyet BİLİNMİYOR sayılır — uydurmak yerine söylemek doğrudur.
    """
    # Bedelsiz giriş (airdrop, Launchpool, kupon): maliyeti gerçekten sıfır.
    # Bu "bilinmiyor" değildir — ikisini karıştırmak ya sahte kâr yazdırır
    # ya da bilinen bir maliyeti engel sanıp öneriyi boşuna bloke eder.
    if olay.get("zero_cost"):
        return 0.0, True
    if not olay.get("usd_known"):
        return 0.0, False
    qty = float(olay.get("qty") or 0.0)
    tutar = abs(float(olay.get("usd_value") or 0.0))
    if qty > 0 and tutar > 0:
        return tutar / qty, True
    fiyat = float(olay.get("price") or 0.0)
    return (fiyat, True) if fiyat > 0 else (0.0, False)


def fifo_rebuild(olaylar):
    """
    Tek bir (varlık, borsa) çifti için kalan lotları FIFO ile yeniden kurar.

    `(lotlar, tanı)` döndürür. Tanı, önerinin güvenilir olup olmadığına karar
    vermek için gereken sinyalleri taşır; bunlar gizlenmez, rapora çıkar.

    GERÇEKLEŞMİŞ K/Z NEDEN BURADA HESAPLANIYOR
    ------------------------------------------
    FIFO'da hayatta kalan lotlar en SON alımlardır. Düşen bir coinde bunlar
    genellikle en ucuz alımlardır; dolayısıyla yeniden kurulmuş pozisyonun
    maliyeti, kullanıcının konsolide ortalamasından belirgin biçimde düşük
    çıkar. Aradaki fark kaybolmuş değildir — geçmişteki satışlarda
    GERÇEKLEŞMİŞTİR.

    Kullanıcının defterinde hiç satış kaydı olmadığı için o zarar hiçbir yerde
    durmuyor. Yalnızca açık lotları düzeltip bunu atlarsak pozisyon ucuzlar,
    tablo iyileşir ve sistem gerçekte olduğundan KÂRLI görünür. Bu, düzeltmenin
    amacının tam tersidir. Bu yüzden kapanmış tur-işlemlerin gerçekleşmiş K/Z'si
    de burada hesaplanır ve uygulanırken deftere tek bir özet kayıt olarak
    yazılır.

    Çekilen miktarın taşıdığı maliyet ayrı tutulur: o coinler satılmadı, başka
    bir konuma gitti. Kullanıcı onları oraya bu maliyetle girebilsin diye
    rapora yazılır.
    """
    lotlar = []
    fazla_satis = 0.0
    cekilen = 0.0
    cekilen_maliyet = 0.0
    komisyon_qty = 0.0
    komisyon_maliyet = 0.0
    odul_qty = 0.0
    yatirilan = 0.0
    satilan_qty = 0.0
    satis_hasilati = 0.0
    satilan_maliyet = 0.0
    son_satis_tarihi = None
    kz_bilinmeyen = 0
    # Aynı gün içinde önce giriş sonra çıkış işlensin; ters sıra sahte
    # "açığa satış" üretir ve kapsam boşluğu sanılır.
    sirali = sorted(olaylar, key=lambda o: (o.get("time") or "",
                                            0 if float(o.get("qty") or 0) >= 0 else 1))
    for o in sirali:
        qty = float(o.get("qty") or 0.0)
        tarih = (o.get("time") or "")[:10]
        if qty > 0:
            if o.get("kind") == "DEPOSIT":
                yatirilan += qty
                maliyet, bilinir = 0.0, False
            else:
                if o.get("kind") == "REWARD":
                    odul_qty += qty
                maliyet, bilinir = _birim_maliyet(o)
            lotlar.append({"date": tarih, "qty": qty, "cost": maliyet,
                           "cost_known": bilinir, "kind": o.get("kind")})
        elif qty < 0:
            satis_mi = o.get("kind") == "TRADE"
            kalan = -qty
            tuketilen_maliyet = 0.0
            maliyet_tam = True
            while kalan > 1e-12 and lotlar:
                alinan = min(kalan, lotlar[0]["qty"])
                if lotlar[0]["cost_known"]:
                    tuketilen_maliyet += alinan * lotlar[0]["cost"]
                else:
                    maliyet_tam = False
                lotlar[0]["qty"] -= alinan
                kalan -= alinan
                if lotlar[0]["qty"] <= 1e-12:
                    lotlar.pop(0)
            if kalan > 1e-9:
                fazla_satis += kalan
                maliyet_tam = False

            if satis_mi:
                satilan_qty += -qty
                satilan_maliyet += tuketilen_maliyet
                son_satis_tarihi = max(son_satis_tarihi or "", tarih) or None
                if o.get("usd_known") and maliyet_tam:
                    satis_hasilati += abs(float(o.get("usd_value") or 0.0))
                else:
                    kz_bilinmeyen += 1
            elif o.get("kind") == "FEE":
                # Komisyon bir satış değildir; hasılatı yoktur. Çekim de
                # değildir — coin başka bir konuma gitmedi, harcandı. Ayrı
                # tutulmazsa kullanıcıya olmayan bir transfer uyarısı çıkar.
                komisyon_qty += -qty
                komisyon_maliyet += tuketilen_maliyet
            else:
                cekilen += -qty
                cekilen_maliyet += tuketilen_maliyet

    lotlar = [l for l in lotlar if l["qty"] > 1e-12]
    return lotlar, {
        "oversold_qty": fazla_satis,
        "withdrawn_qty": cekilen,
        "withdrawn_cost_usd": cekilen_maliyet,
        "fee_qty": komisyon_qty,
        "fee_cost_usd": komisyon_maliyet,
        "reward_qty": odul_qty,
        "deposited_qty": yatirilan,
        "unknown_cost_qty": sum(l["qty"] for l in lotlar if not l["cost_known"]),
        "zero_cost_qty": sum(l["qty"] for l in lotlar
                             if l["cost_known"] and l["cost"] <= 0),
        "event_count": len(sirali),
        "realized_qty": satilan_qty,
        "realized_proceeds_usd": satis_hasilati,
        "realized_cost_usd": satilan_maliyet,
        "realized_pnl_usd": satis_hasilati - satilan_maliyet,
        "realized_last_date": son_satis_tarihi,
        # Tek bir satışın bile K/Z'si çıkarılamıyorsa toplam güvenilmez;
        # yaklaşık bir rakamı kesinmiş gibi deftere yazmayız.
        "realized_known": kz_bilinmeyen == 0 and satilan_qty > 1e-12,
        "realized_unknown_trades": kz_bilinmeyen,
    }


def lot_signature(lotlar):
    """
    Önerinin parmak izi.

    Kullanıcı öneriyi gördükten sonra klasöre yeni bir dosya koyarsa öneri
    değişir; eski ekranı onaylamak sessizce BAŞKA bir şeyi uygular. İmza
    uyuşmazsa uygulama reddedilir ve kullanıcıdan listeyi yenilemesi istenir.
    """
    ham = "|".join(f"{l['date']}:{round(float(l['qty']), 10):.10f}:"
                   f"{round(float(l['cost']), 12):.12f}" for l in lotlar)
    return hashlib.sha1(ham.encode("utf-8")).hexdigest()[:16]


def _kz_zaten_defterde(data, varlik, borsa):
    """
    Bu (varlık, konum) için defterde zaten gerçekleşmiş K/Z kaydı var mı?

    Varsa borsanın kapanmış turlarını bir kez daha yazmak MÜKERRER olur. Kural
    kaba tutuldu — tek bir satış kaydı bile varsa yazmıyoruz — çünkü hata yönü
    önemli: eksik kalmış bir K/Z kullanıcıyı arayışa iter, uydurulmuş bir K/Z
    ise yanlış bir rakama güvendirir. İkincisi daha zararlıdır.
    """
    from data_manager import normalize_location
    for tx in data.get("transactions", []):
        ham = str(tx.get("coin") or "").upper().strip()
        v = normalize_asset(ham[:-4] if ham.endswith("USDT") and len(ham) > 4 else ham)
        if v != varlik or normalize_location(tx.get("exchange")) != borsa:
            continue
        if tx.get("exit_price") is not None or tx.get("realized_pnl_usd") is not None:
            return True
    return False


def evaluate_verified_qty(satir, verified_qty):
    """
    Kullanıcının borsa ekranından okuyup girdiği GERÇEK bakiyeyi hakem yapar.

    NEDEN VAR (FAZ F5b)
    -------------------
    F5'in kapsam kanıtı şuydu: "bakiye hiç eksiye düşmüyorsa dosya geçmişi
    kapsıyordur." Bu çıkarım GEÇERSİZ. Negatif bakiyenin imkânsızlığı yalnızca
    *"dosyada görünmeyen bir SATIŞ yok"* der; *"dosyada görünmeyen bir ALIM
    yok"* demez. Dosya penceresinden önce alınmış ve hiç satılmamış bir bakiye
    hiçbir iz bırakmaz — ve sistem onu "fazladan girilmiş" sanıp silmeyi
    önerir. Kullanıcının MEXC'teki 562.66 BCCOIN'i tam olarak böyle 238.78'e
    düşürülmek istendi.

    Dosyalar tek başına hangi tarafın haklı olduğunu ASLA söyleyemez. Söyleyen
    tek şey borsadaki güncel gerçek bakiyedir; kullanıcı onu ekrandan okuyup
    girebilir. Karar kuralı basit:

    - gerçek ≈ önerilen  → defter yanlış, düzeltme uygulanabilir
    - gerçek ≈ defter    → dosya eksik, **dokunulmaz**
    - ikisi de değil     → üçüncü bir eksik var, önce o bulunmalı

    Mutlak bir tolerans yerine "hangisine daha yakın" ölçütü kullanılır: iki
    aday birbirine yakınsa sabit bir yüzde ikisini de kabul eder ve doğrulama
    hükmünü yitirirdi.
    """
    try:
        v = float(verified_qty)
    except (TypeError, ValueError):
        return {"ok": False, "verdict": "invalid",
                "message": "Girilen bakiye bir sayı değil."}
    if v < 0:
        return {"ok": False, "verdict": "invalid",
                "message": "Bakiye negatif olamaz."}

    onerilen = float(satir.get("proposed_qty") or 0.0)
    defterde = float(satir.get("ledger_qty") or 0.0)
    varlik = satir.get("asset", "")

    # Miktarlar zaten örtüşüyorsa (fark yalnızca maliyet tabanındaysa)
    # doğrulama tek bir soruya iner: miktar gerçekten bu mu?
    if _yakin(onerilen, defterde):
        if _yakin(v, onerilen):
            return {"ok": True, "verdict": "matches_proposal",
                    "message": f"Borsadaki bakiye ({v:,.8f} {varlik}) öneriyle uyuşuyor."}
        return {"ok": False, "verdict": "matches_neither",
                "message": (f"Girilen bakiye ({v:,.8f}) ne defterdeki ne de hesaplanan "
                            f"miktarla ({onerilen:,.8f}) uyuşuyor. Düzeltme uygulanmadı.")}

    oneriye_uzaklik = abs(v - onerilen)
    deftere_uzaklik = abs(v - defterde)

    if oneriye_uzaklik < deftere_uzaklik and _yakin(v, onerilen):
        return {"ok": True, "verdict": "matches_proposal",
                "message": (f"Borsadaki gerçek bakiye ({v:,.8f} {varlik}) borsa "
                            f"kayıtlarından hesaplanan miktarla uyuşuyor; fark defterden "
                            "kaynaklanıyor. Düzeltme uygulanabilir.")}

    if deftere_uzaklik < oneriye_uzaklik:
        return {"ok": False, "verdict": "matches_ledger",
                "message": (f"Borsadaki gerçek bakiye ({v:,.8f} {varlik}) DEFTERDEKİ "
                            f"miktarla uyuşuyor — yani defteriniz doğru, eksik olan dosya. "
                            f"Bu {varlik} muhtemelen dışa aktarım penceresinden "
                            f"({satir.get('coverage_start') or 'dosya başlangıcı'}) önce "
                            "alınmış ve hiç satılmamış; öyle bir bakiye hiçbir dosyada iz "
                            "bırakmaz. Düzeltme UYGULANMADI — defteriniz olduğu gibi kaldı.")}

    return {"ok": False, "verdict": "matches_neither",
            "message": (f"Girilen bakiye ({v:,.8f} {varlik}) ne defterdeki "
                        f"({defterde:,.8f}) ne de hesaplanan ({onerilen:,.8f}) miktarla "
                        "uyuşuyor. Üçüncü bir kaynak eksik olabilir (başka bir cüzdan, "
                        "kilitli bakiye, kapsam dışı bir borsa). Düzeltme uygulanmadı.")}


def build_rebuild_plan(data, root=None):
    """
    Her (varlık, borsa) çifti için düzeltme önerisi üretir. **Yazma yok.**
    """
    olaylar, kaynaklar, uyarilar = load_all_events(root)
    pencere = coverage_windows(olaylar)
    defter = ledger_positions(data)

    # Olayları çift bazında grupla — tek geçiş yeter.
    gruplar = {}
    for o in olaylar:
        varlik = o.get("asset")
        if not varlik or varlik in STABLE_QUOTES:
            continue
        gruplar.setdefault((varlik, o["exchange"]), []).append(o)

    # Defterde olup borsada hiç görünmeyen pozisyonlar da değerlendirilmeli:
    # "borsa sıfır diyor" da bir bilgidir. Ama yalnızca KAPSANAN borsalar için.
    ciftler = set(gruplar)
    for (varlik, konum) in defter:
        if varlik not in STABLE_QUOTES and konum in pencere:
            ciftler.add((varlik, konum))

    satirlar = []
    for varlik, borsa in sorted(ciftler):
        alt_olaylar = gruplar.get((varlik, borsa), [])
        lotlar, tani = fifo_rebuild(alt_olaylar)
        d = defter.get((varlik, borsa))

        onerilen_qty = sum(l["qty"] for l in lotlar)
        onerilen_maliyet = sum(l["qty"] * l["cost"] for l in lotlar)
        defter_qty = d["qty"] if d else 0.0
        defter_maliyet = d["invested"] if d else 0.0
        kapsam_basi = pencere.get(borsa, {}).get("first")

        coin_adi = d["coin"] if d else f"{varlik}USDT"
        pos_key = f"{coin_adi}@{borsa}"

        # --- Öneriyi ENGELLEYEN durumlar ---
        engeller = []
        if tani["oversold_qty"] > 1e-9:
            engeller.append(
                f"Borsa kayıtlarında {tani['oversold_qty']:,.8f} adet karşılıksız çıkış var — "
                f"alım dosyanın başlangıcından ({kapsam_basi}) önce yapılmış. "
                "Dosya bu varlığın geçmişini tam kapsamıyor."
            )
        if tani["unknown_cost_qty"] > 1e-9:
            engeller.append(
                f"Kalan lotların {tani['unknown_cost_qty']:,.8f} adedi dışarıdan gelen bir "
                "yatırma; borsa dosyası bunun maliyetini bilmiyor. Sıfır maliyet yazmak "
                "sahte kâr üretirdi."
            )
        if d and d.get("first_date") and kapsam_basi and d["first_date"] < kapsam_basi:
            engeller.append(
                f"Defterdeki en eski kayıt {d['first_date']}, dosya ise {kapsam_basi} "
                "tarihinde başlıyor. Öncesi dosyada yok."
            )
        if not lotlar and not d:
            continue  # ne borsada ne defterde bir şey var

        # --- Öneriyi ZAYIFLATAN ama engellemeyen durumlar ---
        # Deftere geçmemiş bir gerçekleşmiş K/Z var mı? Yalnızca hesabı tam
        # çıkarılabiliyorsa ve defterde bu pozisyonun hiç satış kaydı yoksa.
        kz_var = tani["realized_known"] and abs(tani["realized_pnl_usd"]) > 0.01
        kz_defterde = _kz_zaten_defterde(data, varlik, borsa)
        yazilacak_kz = kz_var and not kz_defterde

        # ETKİ ile İKAZ ayrı tutuluyor. "Şu olacak" bir uyarı değildir; ikisini
        # aynı listeye koymak her satırı uyarılı gösterir ve uyarı anlamını
        # yitirir — kullanıcı gerçekten dikkat etmesi gerekeni fark edemez.
        etkiler = []
        ikazlar = []

        if yazilacak_kz:
            isaret = "kâr" if tani["realized_pnl_usd"] > 0 else "zarar"
            etkiler.append(
                f"Kapanmış {tani['realized_qty']:,.8f} adetlik alım-satımın "
                f"${abs(tani['realized_pnl_usd']):,.2f} gerçekleşmiş {isaret}ı deftere "
                "tek bir özet kayıt olarak geçecek. Şu an hiçbir yerde görünmüyor."
            )
        if d is None:
            etkiler.append("Defterde bu konumda böyle bir pozisyon yok; düzeltme yeni bir "
                           "pozisyon açacak.")
        if not lotlar and d:
            etkiler.append("Borsa kayıtlarına göre bu pozisyondan tamamen çıkılmış. "
                           "Düzeltme defterdeki kaydı kapatır, nakit üretmez.")

        fark_qty = onerilen_qty - defter_qty
        if fark_qty < -1e-9:
            # Küçültme yönü TEHLİKELİ yöndür. Dosya penceresinden önce alınmış
            # ve hiç satılmamış bir bakiye hiçbir dosyada iz bırakmaz; hesaba
            # da girmez ve burada "fazladan girilmiş" gibi görünür.
            ikazlar.append(
                f"Öneri pozisyonu {abs(fark_qty):,.8f} adet KÜÇÜLTÜYOR. Dosya "
                f"{kapsam_basi or 'bilinmeyen bir tarihte'} başlıyor; o tarihten önce "
                f"alınmış ve hiç satılmamış bir bakiye dosyada görünmez ve bu hesaba "
                f"girmez. Uygulamadan önce borsadaki gerçek {varlik} bakiyesine bakın."
            )
        if tani.get("zero_cost_qty", 0.0) > 1e-9:
            ikazlar.append(
                f"Kalan lotların {tani['zero_cost_qty']:,.8f} adedi bedelsiz geldi "
                "(airdrop, Launchpool, kupon). Maliyeti sıfır yazılacak — bu doğrudur, "
                "ama satıldığında tutarın tamamı kâr sayılır."
            )
        if tani.get("fee_qty", 0.0) > 1e-9:
            etkiler.append(
                f"{tani['fee_qty']:,.8f} adet işlem komisyonu olarak ödenmiş "
                f"(taşıdığı maliyet ${tani.get('fee_cost_usd', 0.0):,.2f}); "
                "hesaptan düşüldü."
            )
        if tani["withdrawn_qty"] > 1e-9:
            ikazlar.append(
                f"Bu borsadan {tani['withdrawn_qty']:,.8f} adet çekilmiş "
                f"(taşıdığı maliyet ${tani['withdrawn_cost_usd']:,.2f}). Öneri yalnızca "
                f"{borsa} bakiyesini anlatır; çekilen miktar cüzdanınızda duruyorsa onu "
                "ayrı bir konum olarak bu maliyetle eklemelisiniz."
            )
        if kz_var and kz_defterde:
            ikazlar.append(
                f"Borsa kayıtlarındaki ${abs(tani['realized_pnl_usd']):,.2f} gerçekleşmiş K/Z "
                "deftere YAZILMAYACAK: bu pozisyonun zaten satış kaydı var ve iki kez saymak "
                "tabloyu bozardı. Yalnızca açık lotlar düzeltilir."
            )
        elif tani["realized_unknown_trades"] > 0:
            ikazlar.append(
                f"{tani['realized_unknown_trades']} satışın dolar karşılığı çıkarılamadı "
                "(dolara sabitli olmayan kotasyon veya kapsam dışı alım). Gerçekleşmiş K/Z "
                "deftere yazılmayacak — açık lotlar yine de düzeltilir."
            )

        # "Aynı" demek için açık lotların tutması yetmez: deftere geçmemiş bir
        # gerçekleşmiş K/Z varsa uygulanacak bir şey hâlâ vardır.
        ayni = (d is not None and _yakin(onerilen_qty, defter_qty)
                and abs(onerilen_maliyet - defter_maliyet) < 0.01
                and not yazilacak_kz)

        # FAZ F5b — DOĞRULANMAMIŞ HİÇBİR SATIR "UYGULANABİLİR" DEĞİLDİR.
        #
        # Eskiden engeli olmayan her satır yeşil `ready` rozetiyle çıkıyordu.
        # O rozet, dosyaların geçmişi kapsadığı KANITLANMIŞ gibi bir güven
        # veriyordu; oysa kanıt tek yönlüydü (bkz. `evaluate_verified_qty`).
        # Artık uygulanabilirliğin tek kanıtı kullanıcının borsa ekranından
        # okuyup girdiği gerçek bakiyedir. `ready` durumu bu yüzden plandan
        # tamamen kaldırıldı — üretilebilseydi yine hak edilmemiş olurdu.
        if engeller:
            durum = "blocked"
        elif ayni:
            durum = "identical"
        elif ikazlar:
            durum = "caution"
        else:
            durum = "needs_check"
        dogrulama_gerekli = durum in ("caution", "needs_check")

        satirlar.append({
            "asset": varlik,
            "exchange": borsa,
            "coin": coin_adi,
            "pos_key": pos_key,
            "ledger_qty": defter_qty,
            "ledger_invested": defter_maliyet,
            "ledger_avg_cost": d["avg_cost"] if d else 0.0,
            "ledger_lots": d["lots"] if d else 0,
            "ledger_first_date": d.get("first_date") if d else None,
            "proposed_qty": onerilen_qty,
            "proposed_invested": onerilen_maliyet,
            "proposed_avg_cost": (onerilen_maliyet / onerilen_qty) if onerilen_qty > 0 else 0.0,
            "proposed_lots": [{"date": l["date"], "qty": l["qty"], "cost": l["cost"]}
                              for l in lotlar],
            "diff_qty": fark_qty,
            "diff_invested": onerilen_maliyet - defter_maliyet,
            "coverage_start": kapsam_basi,
            "trade_count": tani["event_count"],
            "withdrawn_qty": tani["withdrawn_qty"],
            "withdrawn_cost_usd": tani["withdrawn_cost_usd"],
            "fee_qty": tani.get("fee_qty", 0.0),
            "fee_cost_usd": tani.get("fee_cost_usd", 0.0),
            "reward_qty": tani.get("reward_qty", 0.0),
            "zero_cost_qty": tani.get("zero_cost_qty", 0.0),
            "realized_qty": tani["realized_qty"],
            "realized_proceeds_usd": tani["realized_proceeds_usd"],
            "realized_cost_usd": tani["realized_cost_usd"],
            "realized_pnl_usd": tani["realized_pnl_usd"],
            "realized_last_date": tani["realized_last_date"],
            "realized_known": tani["realized_known"],
            # Uygulama bu bayrağa bakar: K/Z özeti yazılacak mı, yazılmayacak mı.
            "will_book_realized": yazilacak_kz,
            "status": durum,
            "blockers": engeller,
            "warnings": ikazlar,
            "effects": etkiler,
            # Uygulama için borsadaki gerçek bakiye şart mı? Şu an her
            # uygulanabilir satır için şarttır; alan, ileride salt-okunur API
            # bakiyeyi kendisi getirdiğinde anlamını koruyabilsin diye var.
            "verify_required": dogrulama_gerekli,
            "verify_prompt": (
                f"{borsa} hesabınızdaki güncel {varlik} bakiyesini yazın. "
                "Dosyalar hangi tarafın haklı olduğunu tek başına söyleyemez; "
                "bunu yalnızca borsadaki gerçek bakiye söyler."
            ),
            "signature": lot_signature(lotlar),
        })

    sira = {"needs_check": 0, "caution": 1, "blocked": 2, "identical": 3}
    satirlar.sort(key=lambda r: (sira.get(r["status"], 4), -abs(r["diff_invested"]), r["asset"]))

    sayim = {}
    for r in satirlar:
        sayim[r["status"]] = sayim.get(r["status"], 0) + 1

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "export_root": root or export_root(),
        "files_found": len(kaynaklar),
        "sources": kaynaklar,
        "coverage": pencere,
        "rows": satirlar,
        "status_counts": sayim,
        "warnings": uyarilar[:50],
        # Plan üretimi de salt okunurdur; yazma ancak açık onayla olur.
        "read_only": True,
    }

"""
NET BAŞA BAŞ testleri.

Bu metrik gerçek bir kullanım hatasından doğdu: kullanıcı konsolide tabloda
ARB için "Ort: $0.231711" gördü ve bunu "bu fiyatta zararımdan kurtulurum"
diye okudu. Oysa o sayı yalnızca ELDEKİ açık lotların maliyetiydi; daha önce
kapanmış 1,133 ARB alım-satımındaki $376.77 zarardan haberi yoktu. Gerçek
kurtulma fiyatı iki katından fazlaydı.

Buradaki testlerin taşıdığı üç ayrı yük var:

  1. Yeni sayının DOĞRU olduğunu göstermek (kullanıcının Excel'iyle birebir).
  2. Eski sayının DEĞİŞMEDİĞİNİ göstermek — `avg_cost` ve açık K/Z aynı
     kalmalı; yeni metrik onların yerine geçmez, yanlarına eklenir.
  3. Sayının anlamsız çıktığı halleri (negatif fiyat, sıfır pozisyon)
     fiyat gibi göstermediğimizi kanıtlamak.
"""

import copy

import pytest

import data_manager as dm


# ===========================================================================
# GERÇEK VERİ: ARB
#
# Aşağıdaki sayılar kullanıcının gerçek defterinden ve kendi Excel belgesinden
# geliyor. Uydurma değiller; bu yüzden bir regresyon burada kırıldığında
# gerçekten yanlış bir şey olmuş demektir.
# ===========================================================================

# Mutabakatla borsa kayıtlarından kurulan dört açık lot
ARB_ACIK_LOTLAR = [
    (209.5774, 0.392),
    (264.0,    0.3825),
    (171.8,    0.2968),
    (800.9,    0.1261),
]
ARB_MIKTAR = 1446.2774
ARB_ORT_MALIYET = 0.231711          # kullanıcının Excel'iyle aynı

# Kapanmış 1,133.2 ARB'nin mutabakat özeti
ARB_KAPALI_QTY = 1133.2
ARB_KAPALI_BIRIM_MALIYET = 1.078270436993
ARB_SATIS_HASILATI = 845.12758
ARB_KOMISYON_USDT = 0.101092        # Binance'ten; deftere içe aktarılmadı

ARB_ACIK_MALIYET = sum(q * c for q, c in ARB_ACIK_LOTLAR)
ARB_TOPLAM_ALIS = ARB_ACIK_MALIYET + ARB_KAPALI_QTY * ARB_KAPALI_BIRIM_MALIYET


def _arb_defteri(komisyon_dahil=False):
    """ARB'nin gerçek defterdeki şekli: 4 açık lot + 1 mutabakat özeti.

    `komisyon_dahil=True` olduğunda, gerçekleşmiş K/Z'ye Binance'in kestiği
    0.101092 USDT komisyon da işlenir. Gerçek defterde bu tutar YOK
    (mutabakat komisyonları içe aktarmadı) — o yüzden iki hâli de test edilir:
    biri bugünkü verimizin ürettiği sayıyı, diğeri komisyon verisi geldiğinde
    kullanıcının Excel'indeki sayının çıktığını doğrular.
    """
    txs = []
    tx_id = 1
    for qty, cost in ARB_ACIK_LOTLAR:
        txs.append({
            "id": tx_id, "date": "2025-02-25", "coin": "ARBUSDT",
            "exchange": "BINANCE", "qty": qty, "cost": cost,
            "status": "Aktif", "rebuild_in_id": 1, "category": "Layer 2",
        })
        tx_id += 1

    komisyon = ARB_KOMISYON_USDT if komisyon_dahil else 0.0
    realized = ARB_SATIS_HASILATI - ARB_KAPALI_QTY * ARB_KAPALI_BIRIM_MALIYET - komisyon

    txs.append({
        "id": tx_id, "date": "2023-02-07", "coin": "ARBUSDT",
        "exchange": "BINANCE", "qty": ARB_KAPALI_QTY,
        "cost": ARB_KAPALI_BIRIM_MALIYET,
        "status": "Kapandı / İzleme", "close_reason": "rebuild_realized",
        "exit_date": "2025-08-22",
        "exit_price": ARB_SATIS_HASILATI / ARB_KAPALI_QTY,
        "exit_value": ARB_SATIS_HASILATI,
        "realized_pnl_usd": realized,
        "fee_usd": komisyon, "fee_asset": "USDT",
        "rebuild_in_id": 1, "category": "Layer 2", "type": "MUTABAKAT",
    })

    return {
        "wallets": {"usdt_cash": 0.0, "exchange_cash": {}},
        "transactions": txs,
        "next_tx_id": tx_id + 1,
        "targets": {},
    }


class TestArbGercekVeri:

    def test_ort_maliyet_degismedi(self):
        """Yeni metrik eskisinin YERİNE geçmiyor; eski sayı bozulmamalı."""
        d = _arb_defteri()
        m = dm.calculate_portfolio_metrics(d, {})
        arb = [c for c in m["consolidated_coins"] if c["symbol"] == "ARBUSDT"][0]

        assert arb["total_qty"] == pytest.approx(ARB_MIKTAR, abs=1e-6)
        assert arb["avg_cost"] == pytest.approx(ARB_ORT_MALIYET, abs=5e-7)

    def test_net_basabas_komisyonsuz(self):
        """Bugünkü verimizin üretebildiği sayı."""
        bb = dm.calculate_net_breakeven(_arb_defteri())["ARBUSDT"]

        beklenen = (ARB_TOPLAM_ALIS - ARB_SATIS_HASILATI) / ARB_MIKTAR
        assert bb["net_breakeven_price"] == pytest.approx(beklenen, rel=1e-12)
        assert bb["net_breakeven_price"] == pytest.approx(0.4922199, abs=1e-6)

    def test_net_basabas_komisyon_dahil_kullanicinin_excel_sayisi(self):
        """Komisyon verisi olduğunda kullanıcının Excel'indeki sayı çıkmalı."""
        bb = dm.calculate_net_breakeven(_arb_defteri(komisyon_dahil=True))["ARBUSDT"]
        assert bb["net_breakeven_price"] == pytest.approx(0.4922898207, abs=1e-9)

    def test_ort_maliyet_ile_net_basabas_bagimsiz_ve_farkli(self):
        """İkisi birbirinden bağımsız; karıştırılırsa test burada patlar."""
        d = _arb_defteri()
        arb = [c for c in dm.calculate_portfolio_metrics(d, {})["consolidated_coins"]
               if c["symbol"] == "ARBUSDT"][0]

        assert arb["avg_cost"] == pytest.approx(0.231711, abs=5e-7)
        assert arb["net_breakeven"]["price"] == pytest.approx(0.4922199, abs=1e-6)
        # Kullanıcının yaşadığı yanılgının ölçüsü: iki katından fazla.
        assert arb["net_breakeven"]["price"] > arb["avg_cost"] * 2

    def test_dokum_toplamlari_tutarli(self):
        d = _arb_defteri()
        bb = dm.calculate_net_breakeven(d)["ARBUSDT"]

        assert bb["open_cost_usd"] == pytest.approx(ARB_ACIK_MALIYET, rel=1e-12)
        assert bb["realized_pnl_usd"] == pytest.approx(
            ARB_SATIS_HASILATI - ARB_KAPALI_QTY * ARB_KAPALI_BIRIM_MALIYET, rel=1e-12)
        assert bb["net_capital_at_risk"] == pytest.approx(
            ARB_TOPLAM_ALIS - ARB_SATIS_HASILATI, rel=1e-12)
        assert bb["history_quality"] == dm.BB_KALITE_MUTABIK
        assert bb["realized_tx_count"] == 1


class TestOzdeslik:
    """(alışlar − satışlar)/miktar  ≡  (açık maliyet − gerçekleşmiş)/miktar

    Fonksiyon ikinci yolu kullanıyor. Bu test birinci yolu bağımsız olarak
    hesaplayıp ikisinin aynı çıktığını gösterir; formül bozulursa yakalanır.
    """

    def test_iki_yol_ayni_sonucu_verir(self):
        d = _arb_defteri()
        bb = dm.calculate_net_breakeven(d)["ARBUSDT"]

        alislar = ARB_TOPLAM_ALIS
        satislar = ARB_SATIS_HASILATI
        ham_yol = (alislar - satislar) / ARB_MIKTAR

        assert bb["net_breakeven_price"] == pytest.approx(ham_yol, rel=1e-12)


class TestAnlamsizSonuclar:
    """Matematiğin çalıştığı ama sayının gösterilemeyeceği haller."""

    def test_sermaye_cikarildi_negatif_fiyat_uretmez(self):
        """Gerçek örnek: 0.0035 ENS kalmış, +$97 gerçekleşmiş → −$27,472."""
        d = {"transactions": [
            {"id": 1, "coin": "ENSUSDT", "exchange": "BINANCE",
             "qty": 0.0035, "cost": 6.56, "status": "Aktif"},
            {"id": 2, "coin": "ENSUSDT", "exchange": "BINANCE",
             "qty": 30.63, "cost": 14.824281527914, "status": "Kapandı / İzleme",
             "exit_price": 17.991178583088, "realized_pnl_usd": 97.0,
             "close_reason": "rebuild_realized"},
        ]}
        bb = dm.calculate_net_breakeven(d)["ENSUSDT"]

        assert bb["net_breakeven_state"] == dm.BB_DURUM_SERMAYE_CIKTI
        assert bb["net_breakeven_price"] is None, "Negatif fiyat asla gösterilmemeli"
        assert bb["net_capital_at_risk"] < 0

    def test_tam_sifir_net_risk_de_sermaye_cikarildi_sayilir(self):
        """Sınır: net risk tam 0. Başa baş fiyatı 0 olurdu; anlamsız."""
        d = {"transactions": [
            {"id": 1, "coin": "XUSDT", "exchange": "BINANCE",
             "qty": 10.0, "cost": 5.0, "status": "Aktif"},
            {"id": 2, "coin": "XUSDT", "exchange": "BINANCE",
             "qty": 10.0, "cost": 5.0, "status": "Kapandı / İzleme",
             "realized_pnl_usd": 50.0},
        ]}
        bb = dm.calculate_net_breakeven(d)["XUSDT"]
        assert bb["net_capital_at_risk"] == pytest.approx(0.0)
        assert bb["net_breakeven_state"] == dm.BB_DURUM_SERMAYE_CIKTI
        assert bb["net_breakeven_price"] is None

    def test_acik_pozisyon_yoksa_fiyat_yok(self):
        d = {"transactions": [
            {"id": 1, "coin": "YUSDT", "exchange": "BINANCE",
             "qty": 100.0, "cost": 2.0, "status": "Kapandı / İzleme",
             "exit_price": 1.0, "realized_pnl_usd": -100.0},
        ]}
        bb = dm.calculate_net_breakeven(d)["YUSDT"]
        assert bb["net_breakeven_state"] == dm.BB_DURUM_POZISYON_YOK
        assert bb["net_breakeven_price"] is None

    def test_qty_sifir_aktif_kayit_sifira_bolmez(self):
        """Gerçek veride qty=0.0 olan Aktif kayıtlar var (id 35/41)."""
        d = {"transactions": [
            {"id": 1, "coin": "SOLUSDT", "exchange": "BINANCE",
             "qty": 0.0, "cost": 200.0, "status": "Aktif"},
        ]}
        bb = dm.calculate_net_breakeven(d)["SOLUSDT"]
        assert bb["net_breakeven_state"] == dm.BB_DURUM_POZISYON_YOK
        assert bb["net_breakeven_price"] is None


class TestGecmisiOlmayanCoin:
    """Geçmiş satış yoksa sonuç ort. maliyete EŞİT çıkar.

    Bu doğru ama tehlikeli: kullanıcı "demek ki bu coinde geçmiş zararım yok"
    diye okuyabilir. Oysa doğrusu "geçmişini bilmiyoruz". Bu yüzden kalite
    etiketi zorunlu — ARB'de yaşanan yanılgının tersi burada üretilmesin.
    """

    def test_esitlik_ve_etiket(self):
        d = {"transactions": [
            {"id": 1, "coin": "APTUSDT", "exchange": "BINANCE",
             "qty": 44.14, "cost": 6.91882, "status": "Aktif"},
        ]}
        bb = dm.calculate_net_breakeven(d)["APTUSDT"]

        assert bb["net_breakeven_price"] == pytest.approx(6.91882, rel=1e-12)
        assert bb["history_quality"] == dm.BB_KALITE_GECMIS_YOK
        assert bb["realized_tx_count"] == 0


class TestKapsamDisiKayitlar:
    """Hangi kayıtların hesaba GİRMEDİĞİ, girenler kadar önemli."""

    def test_mutabakatla_kapatilan_eski_lotlar_iki_kez_sayilmaz(self):
        """`rebuild` kayıtları düzeltilen ESKİ hâl; yerine özet duruyor.

        Gerçek defterde ARB'nin 8 tane `rebuild` kaydı var. Sayılsalardı aynı
        geçmiş iki kez düşülür ve başa baş yapay olarak yükselirdi.
        """
        temiz = _arb_defteri()
        kirli = copy.deepcopy(temiz)
        kirli["transactions"].extend([
            {"id": 900 + i, "coin": "ARBUSDT", "exchange": "BINANCE",
             "qty": q, "cost": c, "status": "Kapandı / İzleme",
             "close_reason": "rebuild", "rebuild_out_id": 1,
             "exit_date": "2026-08-29"}
            for i, (q, c) in enumerate([(800.9, 0.1261), (171.8, 0.2968), (60.0, 0.86)])
        ])

        a = dm.calculate_net_breakeven(temiz)["ARBUSDT"]
        b = dm.calculate_net_breakeven(kirli)["ARBUSDT"]
        assert a["net_breakeven_price"] == pytest.approx(b["net_breakeven_price"], rel=1e-12)
        assert a["realized_tx_count"] == b["realized_tx_count"] == 1

    def test_rebuild_kaydi_exit_price_tasisa_bile_sayilmaz(self):
        """Savunma: bir düzeltme kaydına yanlışlıkla çıkış fiyatı yazılırsa."""
        d = {"transactions": [
            {"id": 1, "coin": "ZUSDT", "exchange": "BINANCE",
             "qty": 100.0, "cost": 1.0, "status": "Aktif"},
            {"id": 2, "coin": "ZUSDT", "exchange": "BINANCE",
             "qty": 50.0, "cost": 1.0, "status": "Kapandı / İzleme",
             "close_reason": "rebuild", "exit_price": 0.5,
             "realized_pnl_usd": -25.0},
        ]}
        bb = dm.calculate_net_breakeven(d)["ZUSDT"]
        assert bb["realized_tx_count"] == 0
        assert bb["net_breakeven_price"] == pytest.approx(1.0)

    def test_transfer_satis_degildir(self):
        """Varlık başka konuma taşındı; para geri gelmedi, zarar da yok."""
        d = {"transactions": [
            {"id": 1, "coin": "RDNTUSDT", "exchange": "METAMASK",
             "qty": 3722.0, "cost": 0.0137, "status": "Kapandı / İzleme",
             "transfer_out_id": 2, "exit_price": 0.02},
            {"id": 2, "coin": "RDNTUSDT", "exchange": "GATE.IO",
             "qty": 3722.0, "cost": 0.0137, "status": "Aktif",
             "transfer_in_id": 2},
        ]}
        bb = dm.calculate_net_breakeven(d)["RDNTUSDT"]
        assert bb["realized_tx_count"] == 0
        assert bb["net_breakeven_price"] == pytest.approx(0.0137, rel=1e-12)

    def test_yazim_gercek_zarardir_ve_sayilir(self):
        """Değersizleşen coin yazıldığında para gerçekten kaybedildi."""
        d = {"transactions": [
            {"id": 1, "coin": "WUSDT", "exchange": "BINANCE",
             "qty": 100.0, "cost": 1.0, "status": "Aktif"},
            {"id": 2, "coin": "WUSDT", "exchange": "BINANCE",
             "qty": 200.0, "cost": 1.0, "status": "Kapandı / İzleme",
             "close_reason": "write_off", "exit_price": 0.0,
             "realized_pnl_usd": -200.0},
        ]}
        bb = dm.calculate_net_breakeven(d)["WUSDT"]
        # 100 birim maliyet + 200 yazılan zarar = 300, 100 adete bölünür.
        assert bb["net_capital_at_risk"] == pytest.approx(300.0)
        assert bb["net_breakeven_price"] == pytest.approx(3.0)

    def test_usdt_kaydi_atlanir(self):
        d = {"transactions": [
            {"id": 1, "coin": "USDT", "exchange": "BINANCE",
             "qty": 500.0, "cost": 1.0, "status": "Aktif"},
        ]}
        assert "USDT" not in dm.calculate_net_breakeven(d)


class TestKonumBirlestirme:
    """Metrik SEMBOL bazlıdır, konum bazlı değil.

    Gerçek örnek — TIA: zarar MEXC'te gerçekleşti. Konum bazlı hesap tüm
    zararı MEXC satırına yıkıp Binance satırını olduğundan iyi gösteriyordu
    (2.91 / 4.68). Doğrusu ikisi birlikte: 3.07.
    """

    def _tia(self):
        return {
            "wallets": {"usdt_cash": 0.0, "exchange_cash": {}},
            "transactions": [
                {"id": 1, "coin": "TIAUSDT", "exchange": "BINANCE",
                 "qty": 80.0, "cost": 2.8, "status": "Aktif"},
                {"id": 2, "coin": "TIAUSDT", "exchange": "MEXC",
                 "qty": 40.41, "cost": 3.43, "status": "Aktif"},
                {"id": 3, "coin": "TIAUSDT", "exchange": "MEXC",
                 "qty": 36.18, "cost": 4.529849640685, "status": "Kapandı / İzleme",
                 "exit_price": 4.343606965174, "realized_pnl_usd": -6.74,
                 "close_reason": "rebuild_realized"},
            ],
            "targets": {},
        }

    def test_tek_fiyat_uretilir_ve_konumlar_listelenir(self):
        bb = dm.calculate_net_breakeven(self._tia())["TIAUSDT"]

        toplam_qty = 80.0 + 40.41
        toplam_maliyet = 80.0 * 2.8 + 40.41 * 3.43
        beklenen = (toplam_maliyet + 6.74) / toplam_qty

        assert bb["net_breakeven_price"] == pytest.approx(beklenen, rel=1e-12)
        assert bb["multi_location"] is True
        assert bb["locations"] == ["BINANCE", "MEXC"]

    def test_iki_satir_da_ayni_fiyati_gosterir(self):
        """Satırlar konum bazlı ama fiyat sembol geneli — kasıtlı."""
        m = dm.calculate_portfolio_metrics(self._tia(), {})
        tia = [c for c in m["consolidated_coins"] if c["symbol"] == "TIAUSDT"]

        assert len(tia) == 2, "İki konum iki satır üretmeli"
        fiyatlar = {round(c["net_breakeven"]["price"], 10) for c in tia}
        assert len(fiyatlar) == 1, "Aynı sembolün iki satırı farklı başa baş göstermemeli"

    def test_konum_bazli_hesap_farkli_cikardi(self):
        """Yanlış yaklaşımın gerçekten farklı sonuç verdiğini kanıtlar.

        Bu test olmasa 'sembol bazlı' kararı sadece bir yorum olurdu.
        """
        bb = dm.calculate_net_breakeven(self._tia())["TIAUSDT"]
        sadece_binance = (80.0 * 2.8) / 80.0            # zarar görmez
        sadece_mexc = (40.41 * 3.43 + 6.74) / 40.41     # zararın tamamını yer

        assert sadece_binance < bb["net_breakeven_price"] < sadece_mexc


class TestGecmisKalitesi:

    def test_uc_durum(self):
        d = {"transactions": [
            # mutabakattan gelen
            {"id": 1, "coin": "AUSDT", "exchange": "BINANCE", "qty": 10.0,
             "cost": 1.0, "status": "Aktif", "rebuild_in_id": 1},
            # deftere elle girilmiş satış
            {"id": 2, "coin": "BUSDT", "exchange": "BINANCE", "qty": 10.0,
             "cost": 1.0, "status": "Aktif"},
            {"id": 3, "coin": "BUSDT", "exchange": "BINANCE", "qty": 5.0,
             "cost": 1.0, "status": "Kapandı / İzleme", "exit_price": 1.5,
             "realized_pnl_usd": 2.5},
            # hiç geçmişi olmayan
            {"id": 4, "coin": "CUSDT", "exchange": "BINANCE", "qty": 10.0,
             "cost": 1.0, "status": "Aktif"},
        ]}
        t = dm.calculate_net_breakeven(d)
        assert t["AUSDT"]["history_quality"] == dm.BB_KALITE_MUTABIK
        assert t["BUSDT"]["history_quality"] == dm.BB_KALITE_DEFTER
        assert t["CUSDT"]["history_quality"] == dm.BB_KALITE_GECMIS_YOK

    def test_mutabakat_satis_olmasa_da_mutabik_sayilir(self):
        """Borsa geçmişi okundu ve satış çıkmadı — bu da bir bilgidir."""
        d = {"transactions": [
            {"id": 1, "coin": "MAVUSDT", "exchange": "BINANCE", "qty": 100.0,
             "cost": 0.067418, "status": "Aktif", "rebuild_in_id": 1},
        ]}
        bb = dm.calculate_net_breakeven(d)["MAVUSDT"]
        assert bb["history_quality"] == dm.BB_KALITE_MUTABIK
        assert bb["realized_tx_count"] == 0


class TestRealizedHesabi:
    """`realized_pnl_usd` yoksa `calculate_realized_metrics` ile aynı sırayla
    türetilmeli; iki yerde iki farklı tanım aynı defterden farklı toplam üretir.
    """

    def test_exit_price_uzerinden_turetilir(self):
        d = {"transactions": [
            {"id": 1, "coin": "KUSDT", "exchange": "BINANCE", "qty": 10.0,
             "cost": 2.0, "status": "Aktif"},
            {"id": 2, "coin": "KUSDT", "exchange": "BINANCE", "qty": 10.0,
             "cost": 2.0, "status": "Kapandı / İzleme", "exit_price": 1.0},
        ]}
        bb = dm.calculate_net_breakeven(d)["KUSDT"]
        assert bb["realized_pnl_usd"] == pytest.approx(-10.0)
        assert bb["net_breakeven_price"] == pytest.approx(3.0)

    def test_komisyon_realized_hesabina_girer(self):
        d = {"transactions": [
            {"id": 1, "coin": "KUSDT", "exchange": "BINANCE", "qty": 10.0,
             "cost": 2.0, "status": "Aktif"},
            {"id": 2, "coin": "KUSDT", "exchange": "BINANCE", "qty": 10.0,
             "cost": 2.0, "status": "Kapandı / İzleme", "exit_price": 1.0,
             "fee_usd": 1.5},
        ]}
        bb = dm.calculate_net_breakeven(d)["KUSDT"]
        assert bb["realized_pnl_usd"] == pytest.approx(-11.5)
        assert bb["net_breakeven_price"] == pytest.approx(3.15)

    def test_exit_value_exit_price_uzerinde_onceliklidir(self):
        d = {"transactions": [
            {"id": 1, "coin": "KUSDT", "exchange": "BINANCE", "qty": 10.0,
             "cost": 2.0, "status": "Aktif"},
            {"id": 2, "coin": "KUSDT", "exchange": "BINANCE", "qty": 10.0,
             "cost": 2.0, "status": "Kapandı / İzleme", "exit_price": 1.0,
             "exit_value": 12.0},
        ]}
        bb = dm.calculate_net_breakeven(d)["KUSDT"]
        assert bb["realized_pnl_usd"] == pytest.approx(-8.0)


class TestKonsolideSatiraIlistirme:

    def test_alanlar_satirda_var(self, ornek_portfoy):
        m = dm.calculate_portfolio_metrics(ornek_portfoy, {})
        for c in m["consolidated_coins"]:
            assert "net_breakeven" in c
            assert c["net_breakeven"]["state"] in (
                dm.BB_DURUM_OK, dm.BB_DURUM_SERMAYE_CIKTI, dm.BB_DURUM_POZISYON_YOK)
            # Düz alanlar da olmalı — Excel/AI iç içe sözlük okumasın.
            assert "net_breakeven_price" in c
            assert "net_breakeven_state" in c

    def test_toplam_kz_dokumu_tutarli(self):
        """symbol_total_pnl == symbol_open_pnl + realized_pnl"""
        d = _arb_defteri()
        arb = [c for c in dm.calculate_portfolio_metrics(d, {})["consolidated_coins"]
               if c["symbol"] == "ARBUSDT"][0]
        bb = arb["net_breakeven"]

        assert bb["symbol_total_pnl_usd"] == pytest.approx(
            bb["symbol_open_pnl_usd"] + bb["realized_pnl_usd"], abs=0.01)

    def test_cok_konumlu_sembolde_dokum_tum_konumlari_kapsar(self):
        d = {
            "wallets": {"usdt_cash": 0.0, "exchange_cash": {}},
            "transactions": [
                {"id": 1, "coin": "ETHUSDT", "exchange": "BINANCE",
                 "qty": 1.0, "cost": 2500.0, "status": "Aktif"},
                {"id": 2, "coin": "ETHUSDT", "exchange": "METAMASK",
                 "qty": 2.0, "cost": 2000.0, "status": "Aktif"},
            ],
            "targets": {},
        }
        satirlar = [c for c in dm.calculate_portfolio_metrics(d, {})["consolidated_coins"]
                    if c["symbol"] == "ETHUSDT"]
        assert len(satirlar) == 2
        for c in satirlar:
            bb = c["net_breakeven"]
            # Satırın kendi miktarı değil, sembolün toplamı
            assert bb["symbol_qty"] == pytest.approx(3.0)
            assert bb["symbol_open_cost_usd"] == pytest.approx(6500.0)
            assert bb["multi_location"] is True

    def test_portfoy_metrikleri_bozulmadi(self, ornek_portfoy):
        """Geriye dönük uyumluluk: mevcut alanlar aynı kalmalı."""
        onceki = copy.deepcopy(ornek_portfoy)
        m = dm.calculate_portfolio_metrics(ornek_portfoy, {})

        assert ornek_portfoy == onceki, "Hesap veriyi değiştirmemeli"
        btc = [c for c in m["consolidated_coins"] if c["symbol"] == "BTCUSDT"][0]
        assert btc["total_qty"] == pytest.approx(0.02)
        assert btc["avg_cost"] == pytest.approx(85000.0)
        assert btc["total_invested"] == pytest.approx(1700.0)


class TestExcelDisaAktarim:
    """Excel'e sütun eklerken sonraki 6 sütunun indeksi kaydı.

    Bu sayfanın hiç testi yoktu; kayma sessizce yanlış başlık altına yanlış
    sayı yazabilirdi. Artık başlık sırası ve toplam formüllerinin hangi
    sütuna baktığı açıkça doğrulanıyor.
    """

    def _sayfa(self, data):
        import io as _io
        import openpyxl
        icerik = dm.export_portfolio_excel(data, {})
        wb = openpyxl.load_workbook(_io.BytesIO(icerik))
        return wb["Konsolide Portfoy"]

    def test_baslik_sirasi(self):
        ws = self._sayfa(_arb_defteri())
        basliklar = [ws.cell(row=4, column=i).value for i in range(1, 13)]
        assert basliklar == [
            "Varlık", "Borsa", "Kategori", "Miktar", "Ort. Maliyet ($)",
            "Net Başa Baş ($)", "Canlı Fiyat ($)", "Toplam Yatırım ($)",
            "Güncel Değer ($)", "Net K/Z ($)", "Getiri (%)", "Portföy Payı (%)",
        ]

    def test_degerler_dogru_sutunda(self):
        ws = self._sayfa(_arb_defteri())
        assert ws.cell(row=5, column=5).value == pytest.approx(ARB_ORT_MALIYET, abs=5e-7)
        assert ws.cell(row=5, column=6).value == pytest.approx(0.4922199, abs=1e-6)
        assert ws.cell(row=5, column=8).value == pytest.approx(ARB_ACIK_MALIYET, rel=1e-9)

    def test_toplam_formulleri_dogru_sutunu_topluyor(self):
        ws = self._sayfa(_arb_defteri())
        tot = 6  # 4 başlık satırı + 1 veri satırı
        assert ws.cell(row=tot, column=1).value == "GENEL TOPLAM"
        assert ws.cell(row=tot, column=8).value == "=SUM(H5:H5)"
        assert ws.cell(row=tot, column=9).value == "=SUM(I5:I5)"
        assert ws.cell(row=tot, column=10).value == "=SUM(J5:J5)"
        assert ws.cell(row=tot, column=12).value == "100.0%"

    def test_sermaye_cikarildi_sayi_yerine_kelime_yazilir(self):
        """0 yazmak 'başa baş sıfır dolar' gibi okunurdu."""
        d = {"wallets": {"usdt_cash": 0.0, "exchange_cash": {}}, "targets": {},
             "transactions": [
                 {"id": 1, "coin": "ENSUSDT", "exchange": "BINANCE",
                  "qty": 0.0035, "cost": 6.56, "status": "Aktif"},
                 {"id": 2, "coin": "ENSUSDT", "exchange": "BINANCE",
                  "qty": 30.63, "cost": 14.82, "status": "Kapandı / İzleme",
                  "exit_price": 17.99, "realized_pnl_usd": 97.0,
                  "close_reason": "rebuild_realized"},
             ]}
        ws = self._sayfa(d)
        assert ws.cell(row=5, column=6).value == "Sermaye cikarildi"


class TestYapayZekaBaglami:
    """Model yanlış başabaş sayısına dayanarak tavsiye vermemeli."""

    def test_gecmisi_olan_coin_net_basabas_ile_gonderilir(self, monkeypatch):
        import ai_service
        d = _arb_defteri()
        monkeypatch.setattr(ai_service, "load_portfolio", lambda: d)
        monkeypatch.setattr(ai_service.price_service, "get_prices", lambda: {})

        ctx = ai_service.AIFinancialAdvisor().get_portfolio_context()
        arb = [c for c in ctx["coins"] if c["symbol"] == "ARBUSDT"][0]

        assert arb["net_breakeven"] is not None
        assert arb["net_breakeven"]["net_breakeven_price"] == pytest.approx(
            0.4922199, abs=1e-6)
        assert "acik lotlari" in arb["net_breakeven"]["note"]

    def test_gecmisi_olmayan_coinde_gurultu_yok(self, monkeypatch):
        import ai_service
        d = {"wallets": {"usdt_cash": 0.0, "exchange_cash": {}}, "targets": {},
             "transactions": [
                 {"id": 1, "coin": "APTUSDT", "exchange": "BINANCE",
                  "qty": 44.14, "cost": 6.91882, "status": "Aktif"},
             ]}
        monkeypatch.setattr(ai_service, "load_portfolio", lambda: d)
        monkeypatch.setattr(ai_service.price_service, "get_prices", lambda: {})

        ctx = ai_service.AIFinancialAdvisor().get_portfolio_context()
        apt = [c for c in ctx["coins"] if c["symbol"] == "APTUSDT"][0]
        assert apt["net_breakeven"] is None

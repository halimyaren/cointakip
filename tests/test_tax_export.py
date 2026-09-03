r"""
CoinTakip — Vergi-Hazır Dışa Aktarım Testleri

En kritik güvenceler:

1. **Hiçbir satır sessizce düşmez.** Defterdeki her kayıt tam olarak dört
   kümeden birine girer: gerçekleşmiş / eksik veri / kapsam dışı. Toplamları
   defterin tamamını vermelidir. Bir vergi dosyasının en tehlikeli hatası,
   kaydı sessizce atmaktır — bu hem kazancı hem zararı eksik gösterir.

2. **Transfer satış değildir.** Bir varlığı kendi cüzdanına taşımak
   gerçekleşmiş K/Z üretmez; vergi tablosunda satır olarak görünemez.

3. **Mutabakat kapanışı satış değildir.** `close_reason == "rebuild"` bir
   hatalı kayıt düzeltmesidir, bir elden çıkarma değil.

4. **TRY yoktur ve dosya bunu kendi üstünde yazar.** Kur uygulamak üç ayrı
   karar gerektirir ve yanlış kur, doğru veriden yanlış beyan üretir.

5. **`calculate_realized_metrics` ile aynı ölçüt.** İki yerin farklı cevap
   vermesi, ekrandaki K/Z ile dosyadaki K/Z'nin tutmaması demektir.

6. **Dışa aktarım deftere hiçbir şey YAZMAZ.**

Testler ağa çıkmaz ve gerçek veriye dokunmaz.
"""

import io

import pytest

import data_manager as dm
import tax_export as tx


# =============================================================
# Ortak defter — her olay türünden en az bir kayıt
# =============================================================
@pytest.fixture
def defter():
    return {
        "transactions": [
            # 1) Düz satış, kâr
            {
                "id": 1, "coin": "ETHUSDT", "exchange": "BINANCE",
                "qty": 0.5, "cost": 2000.0, "status": dm.CLOSED_STATUS,
                "date": "2025-11-02", "exit_date": "2026-03-14",
                "exit_price": 3000.0, "exit_value": 1500.0,
                "fee_usd": 1.5, "realized_pnl_usd": 498.5,
                "cost_method": "FIFO",
            },
            # 2) Yazım (değersizleşme) — hasılatı yok, zarar
            {
                "id": 2, "coin": "SCM", "exchange": "METAMASK",
                "qty": 1000.0, "cost": 0.4, "status": dm.CLOSED_STATUS,
                "date": "2024-06-01", "exit_date": "2026-01-09",
                "exit_price": 0.0, "exit_value": 0.0,
                "realized_pnl_usd": -400.0,
                "close_reason": "write_off", "write_off_reason": "rug",
            },
            # 3) Hâlâ açık pozisyon — kapsam dışı
            {
                "id": 3, "coin": "BTCUSDT", "exchange": "BINANCE",
                "qty": 0.1, "cost": 30000.0, "status": dm.ACTIVE_STATUS,
                "date": "2025-01-01",
            },
            # 4) Mutabakat kapanışı — hatalı kayıt düzeltmesi, satış değil
            {
                "id": 4, "coin": "ARBUSDT", "exchange": "BINANCE",
                "qty": 800.0, "cost": 0.12, "status": dm.CLOSED_STATUS,
                "date": "2026-08-22", "close_reason": "rebuild",
            },
            # 5) Transfer — satış değil
            {
                "id": 5, "coin": "BNBUSDT", "exchange": "MEXC",
                "qty": 2.0, "cost": 500.0, "status": dm.CLOSED_STATUS,
                "type": "TRANSFER", "date": "2026-02-02",
            },
            # 6) Kapanmış ama çıkış fiyatı yok — EKSİK VERİ
            {
                "id": 6, "coin": "AXSUSDT", "exchange": "BINANCE",
                "qty": 43.25, "cost": 1.179, "status": dm.CLOSED_STATUS,
                "date": "2026-08-22",
            },
            # 7) Mutabakat özeti — toplu sonuç, tek işlem değil
            {
                "id": 7, "coin": "IMXUSDT", "exchange": "BINANCE",
                "qty": 100.0, "cost": 1.0, "status": dm.CLOSED_STATUS,
                "type": "MUTABAKAT", "date": "2025-02-01",
                "exit_date": "2025-09-09", "exit_price": 1.2,
                "exit_value": 120.0, "realized_pnl_usd": 20.0,
                "close_reason": "rebuild_realized",
            },
        ]
    }


# =============================================================
# 1. KÜMELERE AYIRMA — hiçbir kayıt kaybolmaz
# =============================================================
class TestKumelereAyirma:

    def test_her_kayit_tam_olarak_bir_kumeye_girer(self, defter):
        g = tx.build_tax_rows(defter)
        toplam = len(g["rows"]) + len(g["gaps"]) + len(g["excluded"])
        assert toplam == len(defter["transactions"]) == 7

    def test_gerceklesmis_olaylar(self, defter):
        g = tx.build_tax_rows(defter)
        assert {r["Kayıt No"] for r in g["rows"]} == {1, 2, 7}

    def test_eksik_veri_atilmaz_ayri_listelenir(self, defter):
        g = tx.build_tax_rows(defter)
        assert [r["Kayıt No"] for r in g["gaps"]] == [6]
        assert g["gaps"][0]["Neden Dışarıda"] == tx.EKSIK_CIKIS_YOK

    def test_transfer_satis_sayilmaz(self, defter):
        g = tx.build_tax_rows(defter)
        transfer = next(r for r in g["excluded"] if r["Kayıt No"] == 5)
        assert transfer["Neden Kapsam Dışı"] == tx.DISI_TRANSFER
        assert 5 not in {r["Kayıt No"] for r in g["rows"]}

    def test_mutabakat_kapanisi_satis_sayilmaz(self, defter):
        g = tx.build_tax_rows(defter)
        rebuild = next(r for r in g["excluded"] if r["Kayıt No"] == 4)
        assert rebuild["Neden Kapsam Dışı"] == tx.DISI_REBUILD

    def test_acik_pozisyon_kapsam_disi(self, defter):
        g = tx.build_tax_rows(defter)
        acik = next(r for r in g["excluded"] if r["Kayıt No"] == 3)
        assert acik["Neden Kapsam Dışı"] == tx.DISI_ACIK

    def test_bos_defter_patlamaz(self):
        g = tx.build_tax_rows({"transactions": []})
        assert g["rows"] == [] and g["gaps"] == [] and g["excluded"] == []
        assert g["totals"]["olay_sayisi"] == 0

    def test_none_defter_patlamaz(self):
        assert tx.build_tax_rows(None)["rows"] == []


# =============================================================
# 2. OLAY TÜRLERİ — satış, yazım ve toplu özet ayrışır
# =============================================================
class TestOlayTurleri:

    def test_satis_yazim_ve_mutabakat_ayri_isaretlenir(self, defter):
        g = tx.build_tax_rows(defter)
        turler = {r["Kayıt No"]: r["Olay Türü"] for r in g["rows"]}
        assert turler[1] == tx.OLAY_SATIS
        assert turler[2] == tx.OLAY_YAZIM
        assert turler[7] == tx.OLAY_MUTABAKAT

    def test_yazim_nedeni_aciklamaya_yazilir(self, defter):
        g = tx.build_tax_rows(defter)
        yazim = next(r for r in g["rows"] if r["Kayıt No"] == 2)
        assert dm.WRITE_OFF_REASONS["rug"] in yazim["Açıklama"]

    def test_toplu_ozet_kapsam_beyaninda_uyari_uretir(self, defter):
        g = tx.build_tax_rows(defter)
        notlar = " ".join(tx.coverage_notes(g))
        assert "Mutabakat özeti" in notlar
        assert "TOPLU" in notlar


# =============================================================
# 3. SAYILAR — ekrandaki K/Z ile dosyadaki K/Z aynı olmalı
# =============================================================
class TestSayilar:

    def test_kayitli_kz_yeniden_hesaplanmaz(self, defter):
        g = tx.build_tax_rows(defter)
        satis = next(r for r in g["rows"] if r["Kayıt No"] == 1)
        assert satis[f"Gerçekleşmiş K/Z ({tx.PARA_BIRIMI})"] == 498.5

    def test_kz_yoksa_komisyon_dusulerek_hesaplanir(self):
        defter = {"transactions": [{
            "id": 1, "coin": "SOLUSDT", "exchange": "BINANCE",
            "qty": 10.0, "cost": 100.0, "status": dm.CLOSED_STATUS,
            "date": "2026-01-01", "exit_date": "2026-06-01",
            "exit_price": 150.0, "exit_value": 1500.0, "fee_usd": 5.0,
        }]}
        satir = tx.build_tax_rows(defter)["rows"][0]
        # (1500 - 1000) - 5
        assert satir[f"Gerçekleşmiş K/Z ({tx.PARA_BIRIMI})"] == 495.0

    def test_toplamlar_satirlarla_tutar(self, defter):
        g = tx.build_tax_rows(defter)
        alan = f"Gerçekleşmiş K/Z ({tx.PARA_BIRIMI})"
        assert g["totals"]["toplam_kz"] == round(sum(r[alan] for r in g["rows"]), 2)
        assert g["totals"]["toplam_kar"] == 518.5    # 498.5 + 20.0
        assert g["totals"]["toplam_zarar"] == 400.0

    def test_kar_ve_zarar_ayri_toplanir(self, defter):
        g = tx.build_tax_rows(defter)
        t = g["totals"]
        assert t["toplam_kz"] == round(t["toplam_kar"] - t["toplam_zarar"], 2)

    def test_sembol_temel_hale_getirilir(self, defter):
        g = tx.build_tax_rows(defter)
        satis = next(r for r in g["rows"] if r["Kayıt No"] == 1)
        assert satis["Varlık"] == "ETH"
        # Deftere geri izlenebilirlik için ham sembol de taşınır
        assert satis["Kayıt Sembolü"] == "ETHUSDT"

    def test_bozuk_sayi_patlatmaz(self):
        defter = {"transactions": [{
            "id": 1, "coin": "X", "exchange": "BINANCE",
            "qty": "abc", "cost": None, "status": dm.CLOSED_STATUS,
            "date": "2026-01-01", "exit_price": 1.0, "exit_value": "",
        }]}
        satir = tx.build_tax_rows(defter)["rows"][0]
        assert satir["Miktar"] == 0.0


# =============================================================
# 4. DÖNEM SÜZGECİ
# =============================================================
class TestDonemSuzgeci:

    def test_yillar_cikis_tarihinden_derlenir(self, defter):
        g = tx.build_tax_rows(defter)
        assert g["available_years"] == ["2026", "2025"]

    def test_yil_secilince_yalnizca_o_yil_gelir(self, defter):
        g = tx.build_tax_rows(defter, "2026")
        assert {r["Kayıt No"] for r in g["rows"]} == {1, 2}
        assert g["year"] == "2026"

    def test_alis_yili_degil_cikis_yili_esas_alinir(self, defter):
        # #1 2025'te alındı, 2026'da satıldı → 2025'te görünmemeli
        g = tx.build_tax_rows(defter, "2025")
        assert {r["Kayıt No"] for r in g["rows"]} == {7}

    def test_kayitsiz_yil_bos_doner_ama_patlamaz(self, defter):
        g = tx.build_tax_rows(defter, "1999")
        assert g["rows"] == []
        assert g["totals"]["toplam_kz"] == 0.0

    def test_eksik_ve_kapsam_disi_yila_gore_suzulmez(self, defter):
        """Amaçları 'şu yıl ne oldu' değil, 'defterin tamamı nereye düştü'."""
        hepsi = tx.build_tax_rows(defter)
        yillik = tx.build_tax_rows(defter, "2026")
        assert len(yillik["gaps"]) == len(hepsi["gaps"])
        assert len(yillik["excluded"]) == len(hepsi["excluded"])


# =============================================================
# 5. KAPSAM BEYANI — dosyanın ne OLMADIĞINI söylemesi
# =============================================================
class TestKapsamBeyani:

    def test_vergi_hesabi_olmadigi_yazar(self, defter):
        notlar = " ".join(tx.coverage_notes(tx.build_tax_rows(defter)))
        assert "VERGİ BEYANI VEYA VERGİ HESABI DEĞİLDİR" in notlar

    def test_try_uygulanmadigi_yazar(self, defter):
        notlar = " ".join(tx.coverage_notes(tx.build_tax_rows(defter)))
        assert "TRY kuru UYGULANMAMIŞTIR" in notlar

    def test_eksik_veri_varsa_acikca_uyarir(self, defter):
        notlar = " ".join(tx.coverage_notes(tx.build_tax_rows(defter)))
        assert "DİKKAT" in notlar and "çıkış" in notlar

    def test_eksik_veri_yoksa_uyari_cikmaz(self):
        defter = {"transactions": [{
            "id": 1, "coin": "ETHUSDT", "exchange": "BINANCE",
            "qty": 1.0, "cost": 100.0, "status": dm.CLOSED_STATUS,
            "date": "2026-01-01", "exit_date": "2026-02-01",
            "exit_price": 200.0, "exit_value": 200.0,
        }]}
        notlar = " ".join(tx.coverage_notes(tx.build_tax_rows(defter)))
        assert "DİKKAT" not in notlar

    def test_dort_kume_toplami_beyanda_gecer(self, defter):
        notlar = " ".join(tx.coverage_notes(tx.build_tax_rows(defter)))
        assert "Defterdeki 7 kaydın tamamı" in notlar

    def test_hicbir_yerde_para_birimi_try_degil(self, defter):
        g = tx.build_tax_rows(defter)
        assert g["currency"] == "USD"
        for satir in g["rows"]:
            assert not any("TRY" in str(b) for b in satir.keys())


# =============================================================
# 6. DOSYA ÜRETİMİ
# =============================================================
class TestDosyaUretimi:

    def test_csv_uretilir_ve_bom_tasir(self, defter):
        ham = tx.export_tax_csv(defter)
        assert isinstance(ham, bytes) and len(ham) > 0
        # Excel'in Türkçe karakterleri doğru açması için BOM şart
        assert ham.startswith(b"\xef\xbb\xbf")

    def test_csv_tum_bolumleri_icerir(self, defter):
        metin = tx.export_tax_csv(defter).decode("utf-8-sig")
        assert "GERÇEKLEŞMİŞ İŞLEMLER" in metin
        assert "EKSİK VERİ — İNCELENMELİ" in metin
        assert "KAPSAM DIŞI" in metin
        assert "TOPLAMLAR" in metin

    def test_csv_eksik_kaydin_id_sini_tasir(self, defter):
        metin = tx.export_tax_csv(defter).decode("utf-8-sig")
        assert tx.EKSIK_CIKIS_YOK in metin
        assert "AXS" in metin

    def test_xlsx_dort_sayfa_uretir(self, defter):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(tx.export_tax_excel(defter)))
        assert wb.sheetnames == [
            "Ozet", "Gerceklesmis Islemler", "Eksik Veri", "Kapsam Disi"
        ]

    def test_xlsx_satir_sayisi_veriyle_tutar(self, defter):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(tx.export_tax_excel(defter)))
        g = tx.build_tax_rows(defter)
        assert wb["Gerceklesmis Islemler"].max_row == len(g["rows"]) + 1  # + başlık
        assert wb["Eksik Veri"].max_row == len(g["gaps"]) + 1
        assert wb["Kapsam Disi"].max_row == len(g["excluded"]) + 1

    def test_bos_sayfa_sessiz_kalmaz_aciklama_yazar(self):
        import openpyxl
        defter = {"transactions": []}
        wb = openpyxl.load_workbook(io.BytesIO(tx.export_tax_excel(defter)))
        assert "gerçekleşmiş işlem kaydı yok" in str(wb["Gerceklesmis Islemler"]["A1"].value)

    def test_xlsx_bos_defterde_de_uretilir(self):
        assert len(tx.export_tax_excel({"transactions": []})) > 0

    def test_yil_secili_dosya_da_uretilir(self, defter):
        assert len(tx.export_tax_excel(defter, "2026")) > 0
        assert len(tx.export_tax_csv(defter, "2026")) > 0


# =============================================================
# 7. DIŞA AKTARIM DEFTERE YAZMAZ
# =============================================================
class TestSaltOkunurluk:

    def test_defter_degismez(self, defter):
        import copy
        onceki = copy.deepcopy(defter)
        tx.build_tax_rows(defter)
        tx.export_tax_csv(defter)
        tx.export_tax_excel(defter)
        tx.tax_summary(defter)
        assert defter == onceki


# =============================================================
# 8. ÖZET UCU — arayüz indirmeden önce ne göreceğini bilir
# =============================================================
class TestOzet:

    def test_ozet_sayilari_govdeyle_tutar(self, defter):
        ozet = tx.tax_summary(defter)
        g = tx.build_tax_rows(defter)
        assert ozet["total_events"] == len(g["rows"])
        assert ozet["gap_count"] == len(g["gaps"])
        assert ozet["excluded_count"] == len(g["excluded"])
        assert ozet["ledger_tx_count"] == 7

    def test_yil_bazli_sayim(self, defter):
        ozet = tx.tax_summary(defter)
        assert ozet["year_counts"] == {"2026": 2, "2025": 1}

    def test_ozet_dosya_uretmez_ve_para_birimi_usd(self, defter):
        ozet = tx.tax_summary(defter)
        assert ozet["currency"] == "USD"
        assert "rows" not in ozet


# =============================================================
# 9. API UÇLARI
# =============================================================
class TestApiUclari:

    @pytest.fixture
    def istemci(self):
        from fastapi.testclient import TestClient
        import main
        return TestClient(main.app)

    def test_ozet_ucu_calisir(self, istemci):
        r = istemci.get("/api/export/tax/summary")
        assert r.status_code == 200
        assert "gap_count" in r.json()

    def test_xlsx_indirilir(self, istemci):
        r = istemci.get("/api/export/tax")
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]
        assert "CoinTakip_Vergi_TumYillar.xlsx" in r.headers["content-disposition"]

    def test_csv_indirilir(self, istemci):
        r = istemci.get("/api/export/tax?format=csv")
        assert r.status_code == 200
        assert r.headers["content-type"].startswith("text/csv")

    def test_yil_dosya_adina_yansir(self, istemci):
        r = istemci.get("/api/export/tax?year=2026")
        assert r.status_code == 200
        assert "CoinTakip_Vergi_2026.xlsx" in r.headers["content-disposition"]

    def test_gecersiz_bicim_reddedilir(self, istemci):
        assert istemci.get("/api/export/tax?format=pdf").status_code == 400

    def test_serbest_metin_yil_reddedilir(self, istemci):
        """Yıl doğrudan dosya adına giriyor; serbest metin kabul edilmez."""
        assert istemci.get("/api/export/tax?year=../../etc").status_code == 400
        assert istemci.get("/api/export/tax?year=20xx").status_code == 400
